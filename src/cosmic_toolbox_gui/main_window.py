"""PySide6 main window for configuring and executing the access analysis."""

from __future__ import annotations
import math
import time
from dataclasses import replace
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import TYPE_CHECKING

import cartopy.crs as ccrs
import numpy as np
from PySide6.QtCore import QObject, QThread, QTimer, Signal, Qt
from PySide6.QtGui import QIcon
from PySide6.QtWidgets import (
    QCheckBox,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QMainWindow,
    QMessageBox,
    QPushButton,
    QTabWidget,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
    QSizePolicy,
    QWidget,
    QFileDialog,
)
from cosmic_toolbox.models import (
    AnalysisConfig,
    AnalysisOptions,
    AnalysisResult,
    AnalysisSummary,
    DerivedAccessResult,
    GroundStationConfig,
    OrbitConfig,
    PassStatistic,
    PropagatedEphemeris,
    PropagationConfig,
    ScenarioConfig,
    StationSummary,
)
from cosmic_toolbox import ToolboxFacade
from cosmic_toolbox.services import pass_geometry
from cosmic_toolbox.services.cached_access_recompute import (
    analysis_result_from_components,
    derive_access_results_from_ephemeris,
)
from cosmic_toolbox_gui.workers import (
    AnalysisWorker,
)
from cosmic_toolbox.services.scenario_package_io import (
    export_cached_trajectory_package,
    import_cached_trajectory_package,
)
from cosmic_toolbox_gui.plot_helpers import PlotHelpersMixin
from cosmic_toolbox_gui.tabs import (
    GroundTabMixin,
    LinkBudgetTabMixin,
    MissionTabMixin,
    OrbitSummaryTabMixin,
    VisualizationTabMixin,
)
from cosmic_toolbox_gui.widgets.range_slider import RangeSlider

if TYPE_CHECKING:
    import numpy as np
    from matplotlib.backends.backend_qtagg import FigureCanvasQTAgg
    from cosmic_toolbox_gui.opengl import GlobeWidget
    from pyqtgraph import PlotWidget
    from PySide6.QtGui import QColor
    from PySide6.QtWidgets import (
        QComboBox,
        QListWidget,
        QProgressBar,
        QSlider,
    )
    from cosmic_toolbox_gui.widgets.range_slider import RangeSlider



class GroundStationApp(
    GroundTabMixin,
    MissionTabMixin,
    VisualizationTabMixin,
    LinkBudgetTabMixin,
    OrbitSummaryTabMixin,
    PlotHelpersMixin,
    QMainWindow,
):
    """Main PyQt window that exposes the configuration controls."""

    @staticmethod
    def _configure_responsive_tab_widget(tab_widget: QTabWidget) -> None:
        """Keep long tab labels from forcing the main window wider."""
        tab_widget.setDocumentMode(True)
        tab_widget.setUsesScrollButtons(True)
        tab_bar = tab_widget.tabBar()
        tab_bar.setElideMode(Qt.TextElideMode.ElideRight)
        tab_bar.setUsesScrollButtons(True)
        tab_bar.setExpanding(False)

    def __init__(self) -> None:
        super().__init__()
        self.setWindowTitle("Cosmic Comms Toolbox")
        icon_path = (
            Path(__file__).resolve().parents[2] / "resources" / "img" / "menu_icon.png"
        )
        if icon_path.exists():
            self.setWindowIcon(QIcon(str(icon_path)))
        self._station_presets: list[GroundStationConfig] = []
        self._is_dirty = True
        self._derived_outputs_stale = False
        self._suspend_dirty_tracking = False
        self._current_config: AnalysisConfig | None = None
        self._last_result: AnalysisResult | None = None
        self._last_ephemeris: PropagatedEphemeris | None = None
        self._last_derived_access: DerivedAccessResult | None = None
        self._last_ephemeris_signature: tuple | None = None
        self._cached_recompute_message = "Run a mission scenario to cache a reusable trajectory."
        self._station_color_map: dict[str, QColor] = {}
        self.station_table: QTableWidget | None = None
        self.select_all_button: QPushButton | None = None
        self.clear_selection_button: QPushButton | None = None
        self.map_canvas: FigureCanvasQTAgg | None = None
        self.map_axes = None
        self.map_projection = ccrs.PlateCarree()
        self.map_status_label: QLabel | None = None
        self._map_station_artists: list = []
        self.analysis_context_label: QLabel | None = None
        self.contact_link_close_checkbox: QCheckBox | None = None
        self.analysis_cache_status_label: QLabel | None = None
        self.ground_station_recompute_hint_label: QLabel | None = None
        self.ground_station_recompute_button: QPushButton | None = None
        self.refresh_cached_access_button: QPushButton | None = None
        self.export_cached_package_button: QPushButton | None = None
        self.import_cached_package_button: QPushButton | None = None
        self.pie_canvas: FigureCanvasQTAgg | None = None
        self.pie_axes = None
        self.visual_globe_widget: GlobeWidget | None = None
        self.mission_globe_widget: GlobeWidget | None = None
        self.mission_globe_status_label: QLabel | None = None
        self._analysis_thread: QThread | None = None
        self._analysis_worker: AnalysisWorker | None = None
        self._pending_config: AnalysisConfig | None = None
        self.run_progress: QProgressBar | None = None
        self.stop_button: QPushButton | None = None
        # Progress ETA estimation (wall-clock based).
        self._run_progress_last_t: float | None = None
        self._run_progress_last_pct: float | None = None
        self._run_progress_rate_ewma: float | None = None  # percent per second
        self._run_progress_start_t: float | None = None
        self._run_progress_start_pct: float | None = None
        self.link_budget_station_combo: QComboBox | None = None
        self.link_budget_table: QTableWidget | None = None
        self.link_budget_summary_label: QLabel | None = None
        self._link_budget_auto_enabled = False
        self._latest_access_series: dict | None = None
        self._link_budget_rate_curve: tuple[np.ndarray, np.ndarray] | None = None
        self._latest_station_rate_series: dict[str, np.ndarray] | None = None
        self._latest_combined_rate_series: np.ndarray | None = None
        self._downlink_total_label: QLabel | None = None
        self._downlink_per_orbit_label: QLabel | None = None
        self.pass_volume_plot = None
        self.daily_volume_plot = None
        self.pass_bin_label: QLabel | None = None
        self.daily_bin_label: QLabel | None = None
        self._pass_volume_bin_width: float | None = None
        self._daily_volume_bin_width: float | None = None
        self._last_pass_volume_samples = None
        self._last_daily_volume_samples = None
        self._active_station_lookup: dict[str, GroundStationConfig] = {}
        self.visual_station_tabs: QTabWidget | None = None
        self.visual_pass_status_label: QLabel | None = None
        self.visual_play_button: QPushButton | None = None
        self.visual_time_slider: QSlider | None = None
        self.visual_speed_slider: QSlider | None = None
        self.visual_view_tabs: QTabWidget | None = None
        self.visual_graph_combo: QComboBox | None = None
        self.visual_graph_plot: PlotWidget | None = None
        self._visual_graph_data: dict[str, np.ndarray] | None = None
        self._visual_station_lists: dict[str, QListWidget] = {}
        self._visual_selected_pass: PassStatistic | None = None
        self._visual_pass_track: list | None = None
        self._visual_animation_index: int = 0
        self._visual_station_ecef: tuple[float, float, float] | None = None
        self._visual_contact_window: tuple[datetime, datetime] | None = None
        # Debounce timer for the elevation-filter slider: fires once the user
        # stops dragging instead of on every intermediate tick.
        self._elevation_filter_debounce_timer = QTimer(self)
        self._elevation_filter_debounce_timer.setSingleShot(True)
        self._elevation_filter_debounce_timer.setInterval(150)
        self._elevation_filter_debounce_timer.timeout.connect(  # type: ignore[attr-defined]
            self._apply_contact_statistics_filter
        )
        self._visual_animation_timer = QTimer(self)
        self._visual_animation_timer.setInterval(33)
        self._visual_animation_timer.timeout.connect(
            self._advance_visualization_animation
        )
        self._visual_animation_timer.setSingleShot(False)
        self._visual_base_step = 0.25
        self._visual_speed_multiplier = 1.0
        self._visual_anim_fraction = 0.0
        self._visual_frame_mode: str = "ECI"
        self._visual_reference_epoch: datetime | None = None
        self._visual_freq_listener_connected = False
        self.mission_frame_combo: QComboBox | None = None
        self.mission_window_slider: QSlider | None = None
        self.mission_window_label: QLabel | None = None
        self._mission_frame_mode: str = "ECI"
        self._mission_reference_epoch: datetime | None = None
        self._mission_orbit_period_s: float = 0.0
        self._mission_window_fractions: tuple[float, float] = (0.0, 1.0)
        self._mission_ground_track: PropagatedEphemeris | None = None
        self._mission_globe_refresh_timer = QTimer(self)
        self._mission_globe_refresh_timer.setSingleShot(True)
        self._mission_globe_refresh_timer.setInterval(120)
        self._mission_globe_refresh_timer.timeout.connect(self._refresh_mission_globe)
        # Optional standalone AoA figure (Matplotlib) shown at the end of each run.
        self._aoa_window: QMainWindow | None = None
        self._aoa_canvas = None
        self._aoa_axes = None
        self._filtered_contact_result: AnalysisResult | None = None
        self._build_ui()
        self._update_cached_access_ui_state()

    def _set_run_button_state(self, state: str) -> None:
        """Update run button text and color based on state."""
        if not hasattr(self, "run_button"):
            return
        palette = {
            "dirty": ("Run Analysis", "#FFA500", "#FFD580"),
            "running": ("Running…", "#1E90FF", "#87CEFA"),
            "success": ("Completed", "#2E8B57", "#90EE90"),
        }
        text, color, hover = palette.get(state, palette["dirty"])
        self.run_button.setText(text)
        self.run_button.setStyleSheet(
            f"QPushButton {{ background-color: {color}; border-radius: 4px; padding: 6px; }}"
            f"QPushButton:hover {{ background-color: {hover}; }}"
        )

    def _mark_dirty(self) -> None:
        """Mark configuration as needing re-run."""
        if self._suspend_dirty_tracking:
            return
        self._is_dirty = True
        self._derived_outputs_stale = True
        self._set_run_button_state("dirty")
        self._update_cached_access_ui_state()

    def _mark_derived_outputs_stale(self, message: str | None = None) -> None:
        """Mark cached access outputs as stale while preserving reusable ephemeris."""
        if self._suspend_dirty_tracking:
            return
        if self._last_ephemeris is None:
            self._mark_dirty()
            return
        self._derived_outputs_stale = True
        if message is not None:
            self._cached_recompute_message = message
        self._update_cached_access_ui_state()

    def _ephemeris_signature_for_config(self, config: AnalysisConfig) -> tuple:
        """Return a stable key for the subset of config that affects propagation."""
        propagation = config.propagation
        return (
            config.scenario.start_time,
            config.scenario.end_time,
            config.orbit.semi_major_axis_km,
            config.orbit.eccentricity,
            config.orbit.inclination_deg,
            config.orbit.raan_deg,
            config.orbit.arg_perigee_deg,
            config.orbit.mean_anomaly_deg,
            propagation.propagator_type,
            propagation.sample_step_seconds,
            propagation.attitude_mode,
            propagation.enable_contact_attitude_switching,
            propagation.drag_coefficient,
            propagation.apply_mean_orbit_correction,
        )

    def _is_cached_recompute_supported_for_config(self, config: AnalysisConfig) -> bool:
        """Return True when cached access recompute is supported for this config."""
        return bool(config.options.compute_ground_station_passes)

    def _collect_run_inputs(self) -> tuple[AnalysisConfig, list[GroundStationConfig]]:
        """Build the current config and active station list from the UI."""
        perform_ground_passes = getattr(self, "ground_pass_checkbox", None)
        ground_pass_enabled = (
            perform_ground_passes.isChecked()
            if perform_ground_passes is not None
            else True
        )
        stations: list[GroundStationConfig] = []
        primary_station: GroundStationConfig | None = None
        if ground_pass_enabled:
            stations = self._collect_active_stations_for_run()
            if not stations:
                raise ValueError("Select at least one enabled ground station.")
            primary_station = stations[0]
        config = self._build_config_from_inputs(
            primary_station=primary_station,
            enable_ground_pass_analysis=ground_pass_enabled,
        )
        return config, stations

    def _update_cached_access_ui_state(self) -> None:
        """Refresh cached-trajectory status labels and refresh button state."""
        has_ephemeris = self._last_ephemeris is not None
        status = self._cached_recompute_message
        if self._is_dirty:
            status = "Full rerun required before cached recompute can be used."
        elif has_ephemeris and self._derived_outputs_stale:
            status = self._cached_recompute_message
        elif has_ephemeris:
            status = "Ephemeris cached for recompute. Derived access outputs are current."
        if self.analysis_cache_status_label is not None:
            self.analysis_cache_status_label.setText(status)
        if self.ground_station_recompute_hint_label is not None:
            self.ground_station_recompute_hint_label.setText(status)
        can_recompute = has_ephemeris and self._derived_outputs_stale and not self._is_dirty
        if self.ground_station_recompute_button is not None:
            self.ground_station_recompute_button.setEnabled(can_recompute)
        if self.refresh_cached_access_button is not None:
            self.refresh_cached_access_button.setEnabled(can_recompute)
        if self.export_cached_package_button is not None:
            self.export_cached_package_button.setEnabled(has_ephemeris)
        if self.import_cached_package_button is not None:
            self.import_cached_package_button.setEnabled(True)

    def _collect_scenario_package_settings(self) -> dict:
        """Capture the current toolbox UI settings for sidecar export."""
        return {
            "orbit": {
                "altitude_km": float(self.altitude_input.value()),
                "eccentricity": float(self.ecc_input.value()),
                "inclination_deg": float(self.inc_input.value()),
                "raan_deg": float(self.raan_input.value()),
                "arg_perigee_deg": float(self.argp_input.value()),
                "mean_anomaly_deg": float(self.mean_anom_input.value()),
                "sso_enabled": bool(self.sso_checkbox.isChecked()),
            },
            "scenario": {
                "start_time_utc": self._qdatetime_to_utc(self.start_datetime).isoformat(),
                "end_time_utc": self._qdatetime_to_utc(self.end_datetime).isoformat(),
            },
            "propagation_ui": {
                "propagator_index": int(self.propagator_combo.currentIndex()),
                "sample_step_seconds": int(self.sample_step_input.value()),
                "ground_pass_enabled": bool(self.ground_pass_checkbox.isChecked()),
                "contact_elevation_deg": float(self.contact_attitude_elevation_input.value()),
                "comms_pointing_mode_index": int(self.comms_pointing_mode_combo.currentIndex()),
                "comms_pointing_aoa_limit_deg": float(self.comms_pointing_aoa_limit_input.value()),
            },
        }

    def _collect_station_package_payload(self) -> tuple[list[dict], list[str]]:
        """Capture station definitions and enabled selections for export."""
        stations = [
            {
                "name": station.name,
                "latitude_deg": float(station.latitude_deg),
                "longitude_deg": float(station.longitude_deg),
                "altitude_m": float(station.altitude_m),
                **(
                    {"horizon_mask_path": station.horizon_mask_path}
                    if getattr(station, "horizon_mask_path", None)
                    else {}
                ),
                **(
                    {"supplier": station.supplier}
                    if getattr(station, "supplier", "")
                    else {}
                ),
            }
            for station in getattr(self, "_station_presets", [])
        ]
        enabled_names: list[str] = []
        for idx in getattr(self, "_get_enabled_station_indices", lambda: [])():
            if 0 <= idx < len(self._station_presets):
                enabled_names.append(self._station_presets[idx].name)
        return stations, enabled_names

    def _apply_imported_scenario_package(self, package) -> tuple[AnalysisConfig, list[GroundStationConfig]]:
        """Apply imported scenario settings and stations to the UI."""
        settings = package.settings
        orbit = settings["orbit"]
        scenario = settings["scenario"]
        propagation_ui = settings["propagation_ui"]

        widgets_to_block = [
            self.altitude_input,
            self.ecc_input,
            self.inc_input,
            self.raan_input,
            self.argp_input,
            self.mean_anom_input,
            self.sso_checkbox,
            self.start_datetime,
            self.end_datetime,
            self.propagator_combo,
            self.sample_step_input,
            self.ground_pass_checkbox,
            self.contact_attitude_elevation_input,
            self.comms_pointing_mode_combo,
            self.comms_pointing_aoa_limit_input,
        ]
        self._suspend_dirty_tracking = True
        for widget in widgets_to_block:
            widget.blockSignals(True)
        try:
            self.sso_checkbox.setChecked(bool(orbit["sso_enabled"]))
            self.altitude_input.setValue(float(orbit["altitude_km"]))
            self.ecc_input.setValue(float(orbit["eccentricity"]))
            self.inc_input.setValue(float(orbit["inclination_deg"]))
            self.raan_input.setValue(float(orbit["raan_deg"]))
            self.argp_input.setValue(float(orbit["arg_perigee_deg"]))
            self.mean_anom_input.setValue(float(orbit["mean_anomaly_deg"]))
            self.start_datetime.setDateTime(
                datetime.fromisoformat(str(scenario["start_time_utc"]))
            )
            self.end_datetime.setDateTime(
                datetime.fromisoformat(str(scenario["end_time_utc"]))
            )
            self.propagator_combo.setCurrentIndex(int(propagation_ui["propagator_index"]))
            self.sample_step_input.setValue(int(propagation_ui["sample_step_seconds"]))
            self.ground_pass_checkbox.setChecked(bool(propagation_ui["ground_pass_enabled"]))
            self.contact_attitude_elevation_input.setValue(
                float(propagation_ui["contact_elevation_deg"])
            )
            self.comms_pointing_mode_combo.setCurrentIndex(
                int(propagation_ui["comms_pointing_mode_index"])
            )
            self.comms_pointing_aoa_limit_input.setValue(
                float(propagation_ui["comms_pointing_aoa_limit_deg"])
            )
        finally:
            for widget in widgets_to_block:
                widget.blockSignals(False)

        imported_stations = [
            GroundStationConfig(
                name=str(item["name"]),
                latitude_deg=float(item["latitude_deg"]),
                longitude_deg=float(item["longitude_deg"]),
                altitude_m=float(item["altitude_m"]),
                horizon_mask_path=item.get("horizon_mask_path") or None,
                supplier=item.get("supplier") or "",
            )
            for item in package.stations
        ]
        self._station_presets = imported_stations
        self._populate_station_table(self._station_presets)
        enabled_set = set(package.enabled_station_names)
        if self.station_table is not None:
            self.station_table.blockSignals(True)
            try:
                for row, station in enumerate(self._station_presets):
                    item = self.station_table.item(row, 0)
                    if item is not None:
                        item.setCheckState(
                            Qt.CheckState.Checked
                            if station.name in enabled_set
                            else Qt.CheckState.Unchecked
                        )
            finally:
                self.station_table.blockSignals(False)
        try:
            self._apply_sso_if_enabled()
            self._update_station_map()
            self._refresh_link_budget_station_list()
        finally:
            self._suspend_dirty_tracking = False
        config, stations = self._collect_run_inputs()
        return config, stations

    def _build_contact_analysis_tab(self) -> QTabWidget:
        """Create the Contact Analysis tab stack."""
        contact_tabs = QTabWidget()
        contact_tabs.setTabPosition(QTabWidget.TabPosition.North)
        self._configure_responsive_tab_widget(contact_tabs)

        # Build the individual components
        gs_stats_tab = self._build_analysis_tabs()
        visualization_tab = self._build_visualization_tab()
        static_link_budget_tab = self._build_static_link_budget_tab()
        dynamic_link_budget_tab = self._build_link_budget_tab()
        self._attach_visualization_frequency_listener()

        # Add them as sub-tabs
        contact_tabs.addTab(gs_stats_tab, "Contact Statistics")
        contact_tabs.addTab(visualization_tab, "Pass Visualization")
        contact_tabs.addTab(static_link_budget_tab, "Static Link Budget Tool")
        contact_tabs.addTab(dynamic_link_budget_tab, "Dynamic Link Budget Tool")

        return contact_tabs

    def _build_ui(self) -> None:
        """Compose the widgets and layouts making up the window."""
        central_widget = QWidget(self)
        self.setCentralWidget(central_widget)
        root_layout = QVBoxLayout(central_widget)
        self.main_tabs = QTabWidget()
        self._configure_responsive_tab_widget(self.main_tabs)
        root_layout.addWidget(self.main_tabs)
        ground_station_tab = self._build_ground_station_tab()
        mission_tab = self._build_mission_analysis_tab()
        contact_analysis_tab = self._build_contact_analysis_tab()
        orbit_summary_tab = self._build_orbit_summary_tab()
        self.main_tabs.addTab(ground_station_tab, "Ground Station Selection")
        self.main_tabs.addTab(mission_tab, "Mission Configuration")
        self.main_tabs.addTab(contact_analysis_tab, "Contact Analysis")
        self.main_tabs.addTab(orbit_summary_tab, "Orbit Summary")



    def _build_analysis_tabs(self) -> QTabWidget:
        """Create the Statistics/Contact Analysis tab stack."""
        self.summary_label = QLabel("No analysis run yet.")
        self.summary_label.setWordWrap(True)
        self.summary_label.setSizePolicy(
            QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Preferred
        )
        self.summary_label.setMinimumWidth(0)
        self.analysis_context_label = QLabel("Active stations: manual entry")
        self.analysis_context_label.setWordWrap(True)
        elevation_filter_row = QHBoxLayout()
        elevation_filter_row.addWidget(QLabel("Elevation filter (deg):"))
        self.contact_elevation_slider = RangeSlider(Qt.Orientation.Horizontal)
        self.contact_elevation_slider.setRange(0, 90)
        self.contact_elevation_slider.setSingleStep(1)
        self.contact_elevation_slider.setValues(0, 90)
        self.contact_elevation_slider.valuesChanged.connect(
            self._handle_contact_elevation_filter_changed
        )  # type: ignore[arg-type]
        elevation_filter_row.addWidget(self.contact_elevation_slider, stretch=1)
        self.contact_elevation_range_label = QLabel("0° → 90°")
        elevation_filter_row.addWidget(self.contact_elevation_range_label)
        self.contact_link_close_checkbox = QCheckBox(
            "Trim pass AOS/LOS to dynamic link close"
        )
        self.contact_link_close_checkbox.setChecked(True)
        self.contact_link_close_checkbox.stateChanged.connect(
            self._handle_contact_link_close_filter_changed
        )  # type: ignore[arg-type]
        self.results_table = QTableWidget(0, 9)
        self.results_table.setHorizontalHeaderLabels(
            [
                "Pass",
                "Station",
                "AOS (UTC)",
                "LOS (UTC)",
                "Duration (min)",
                "Max Elev (deg)",
                "Max S/C Slew Rate (deg/s)",
                "Data Volume (Gbit)",
                "Link Close (°)",
            ]
        )
        self.results_table.setMinimumWidth(0)
        self.results_table.setSizePolicy(
            QSizePolicy.Policy.Ignored, QSizePolicy.Policy.Expanding
        )
        header = self.results_table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.ResizeMode.Stretch)

        # Create export buttons
        export_layout = QHBoxLayout()
        export_csv_button = QPushButton("Export to NPZ")
        export_csv_button.clicked.connect(lambda: self._export_results_table("npz"))
        export_xlsx_button = QPushButton("Export to XLSX")
        export_xlsx_button.clicked.connect(lambda: self._export_results_table("xlsx"))
        export_layout.addWidget(export_csv_button)
        export_layout.addWidget(export_xlsx_button)
        export_layout.addStretch()

        stats_tab = QWidget()
        stats_layout = QVBoxLayout(stats_tab)
        stats_layout.addWidget(self.summary_label)
        stats_layout.addWidget(self.analysis_context_label)
        stats_layout.addLayout(elevation_filter_row)
        stats_layout.addWidget(self.contact_link_close_checkbox)
        stats_layout.addLayout(export_layout)
        stats_layout.addWidget(self.results_table, stretch=1)
        tabs = QTabWidget()
        tabs.setTabPosition(QTabWidget.TabPosition.North)
        self._configure_responsive_tab_widget(tabs)
        tabs.addTab(stats_tab, "Statistics")
        for title, widget in self._create_contact_plot_tabs():
            tabs.addTab(widget, title)
        return tabs

    def _update_contact_elevation_label(self, lower: int, upper: int) -> None:
        self.contact_elevation_range_label.setText(f"{lower}° → {upper}°")

    def _initialize_contact_elevation_filter(self) -> None:
        self.contact_elevation_slider.blockSignals(True)
        try:
            self.contact_elevation_slider.setRange(0, 90)
            self.contact_elevation_slider.setValues(0, 90)
        finally:
            self.contact_elevation_slider.blockSignals(False)
        self._update_contact_elevation_label(0, 90)

    def _handle_contact_elevation_filter_changed(self, lower: int, upper: int) -> None:
        self._update_contact_elevation_label(lower, upper)
        # Restart the debounce timer so the heavy recompute only fires once the
        # user stops dragging the slider rather than on every intermediate tick.
        self._elevation_filter_debounce_timer.start()

    def _handle_contact_link_close_filter_changed(self, _state: int) -> None:
        self._apply_contact_statistics_filter()

    def _use_link_close_for_contact_windows(self) -> bool:
        checkbox = getattr(self, "contact_link_close_checkbox", None)
        return bool(checkbox is not None and checkbox.isChecked())

    def _filter_contact_passes(
        self, passes: list[PassStatistic]
    ) -> list[PassStatistic]:
        lower = float(self.contact_elevation_slider.lowerValue())
        upper = float(self.contact_elevation_slider.upperValue())
        filtered: list[PassStatistic] = []
        for item in passes:
            max_elev = float(item.max_elevation_deg)
            if not math.isfinite(max_elev):
                continue
            if lower <= max_elev <= upper:
                filtered.append(item)
        return filtered

    def _build_contact_summary_from_passes(
        self, passes: list[PassStatistic]
    ) -> AnalysisSummary:
        if self._current_config is None:
            raise ValueError("Missing scenario configuration for summary.")
        scenario = self._current_config.scenario
        scenario_duration_s = float(
            (scenario.end_time - scenario.start_time).total_seconds()
        )
        if not passes:
            return AnalysisSummary(
                total_passes=0,
                total_access_minutes=0.0,
                coverage_percent=0.0,
                avg_duration_minutes=0.0,
                min_duration_minutes=0.0,
                max_duration_minutes=0.0,
            )
        durations = np.array([p.duration_minutes for p in passes], dtype=float)
        total_access_minutes = float(np.sum(durations))
        coverage_percent = (
            100.0 * (total_access_minutes * 60.0) / scenario_duration_s
            if scenario_duration_s > 0.0
            else 0.0
        )
        return AnalysisSummary(
            total_passes=len(passes),
            total_access_minutes=total_access_minutes,
            coverage_percent=coverage_percent,
            avg_duration_minutes=float(np.mean(durations)),
            min_duration_minutes=float(np.min(durations)),
            max_duration_minutes=float(np.max(durations)),
        )

    def _build_contact_result_from_passes(
        self, result: AnalysisResult, passes: list[PassStatistic]
    ) -> AnalysisResult:
        filtered_passes = sorted(passes, key=lambda item: item.aos)
        summary = self._build_contact_summary_from_passes(filtered_passes)
        station_groups: dict[str, list[PassStatistic]] = {}
        for item in filtered_passes:
            station_name = item.station_name or "Ground Station"
            station_groups.setdefault(station_name, []).append(item)
        station_summaries = self._summaries_from_groups(station_groups)
        if result.station_summaries:
            order = [entry.station_name for entry in result.station_summaries]

            def _order_key(entry: StationSummary) -> int:
                return (
                    order.index(entry.station_name)
                    if entry.station_name in order
                    else len(order)
                )

            station_summaries.sort(key=_order_key)
        return replace(
            result,
            passes=filtered_passes,
            summary=summary,
            station_summaries=station_summaries,
        )

    def _build_geometric_filtered_contact_result(
        self, result: AnalysisResult
    ) -> AnalysisResult:
        return self._build_contact_result_from_passes(
            result, self._filter_contact_passes(result.passes)
        )

    def _filter_contact_passes_by_link_close(
        self, passes: list[PassStatistic]
    ) -> list[PassStatistic]:
        if not self._latest_access_series:
            raise ValueError("Link-close filtering requires cached access-series data.")
        if not getattr(self, "_latest_station_rate_series", None):
            raise ValueError("Link-close filtering requires dynamic station rate series.")
        if self._current_config is None:
            raise ValueError("Link-close filtering requires current scenario data.")
        time_axis = np.asarray(
            self._latest_access_series.get("time_seconds", []), dtype=float
        )
        if time_axis.ndim != 1 or time_axis.size < 2:
            raise ValueError("Cached access-series timeline is invalid for filtering.")
        station_rate_lookup = self._latest_station_rate_series or {}
        scenario_start = self._current_config.scenario.start_time
        filtered_passes: list[PassStatistic] = []
        for entry in passes:
            station_name = entry.station_name
            if not station_name:
                raise ValueError(
                    "Link-close filtering requires a station name on every pass."
                )
            rates = station_rate_lookup.get(station_name)
            if rates is None:
                # Station removed from the list or rates not rebuilt yet — omit from
                # link-close-filtered output instead of failing the whole refresh.
                continue
            rate_series = np.asarray(rates, dtype=float)
            bounds = pass_geometry.resolve_positive_rate_bounds(
                time_axis=time_axis,
                rates=rate_series,
                start_sec=(entry.aos - scenario_start).total_seconds(),
                end_sec=(entry.los - scenario_start).total_seconds(),
            )
            if bounds is None:
                continue
            trimmed_start, trimmed_end = bounds
            filtered_passes.append(
                replace(
                    entry,
                    aos=scenario_start + timedelta(seconds=trimmed_start),
                    los=scenario_start + timedelta(seconds=trimmed_end),
                    duration_minutes=(trimmed_end - trimmed_start) / 60.0,
                )
            )
        return filtered_passes

    def _build_filtered_contact_result(self, result: AnalysisResult) -> AnalysisResult:
        filtered_result = self._build_geometric_filtered_contact_result(result)
        if not self._use_link_close_for_contact_windows():
            return filtered_result
        return self._build_contact_result_from_passes(
            filtered_result,
            self._filter_contact_passes_by_link_close(filtered_result.passes),
        )

    def _refresh_contact_statistics_view(
        self, filtered_result: AnalysisResult, *, update_downlink_summary: bool
    ) -> None:
        self._filtered_contact_result = filtered_result
        self._populate_results_table(filtered_result)
        self._update_summary_label(filtered_result)
        self._update_plots(filtered_result)
        if update_downlink_summary:
            self._update_downlink_summary()

    def _apply_contact_statistics_filter(self) -> None:
        result = getattr(self, "_last_result", None)
        if result is None:
            return
        if self._use_link_close_for_contact_windows():
            # Rate series depends on LB params and antenna geometry — not on the
            # elevation filter bounds.  Only recompute if the cache is absent;
            # callers that invalidate LB params (e.g. input changes, LUT upload)
            # must clear _latest_station_rate_series themselves to force a rebuild.
            if self._latest_station_rate_series is None:
                self._refresh_data_volume_rate_series()
        filtered_result = self._build_filtered_contact_result(result)
        self._refresh_contact_statistics_view(
            filtered_result, update_downlink_summary=True
        )

    def _get_geometric_contact_statistics_result(self) -> AnalysisResult | None:
        result = getattr(self, "_last_result", None)
        if result is None:
            return None
        return self._build_geometric_filtered_contact_result(result)

    def _get_contact_statistics_result(self) -> AnalysisResult | None:
        filtered = getattr(self, "_filtered_contact_result", None)
        if filtered is not None:
            return filtered
        return getattr(self, "_last_result", None)

    def _apply_analysis_result(
        self,
        *,
        config: AnalysisConfig,
        result: AnalysisResult,
        stations: list[GroundStationConfig],
        update_ephemeris_cache: bool,
    ) -> None:
        """Apply a result payload to the UI and cache state."""
        self._current_config = config
        self._active_station_lookup = {station.name: station for station in stations}
        if self._current_config is not None:
            self._visual_reference_epoch = self._current_config.scenario.start_time
        self._last_result = result
        self._filtered_contact_result = None
        if update_ephemeris_cache:
            self._last_ephemeris = result.ephemeris
            self._last_ephemeris_signature = self._ephemeris_signature_for_config(config)
        self._last_derived_access = result.derived_access
        self._is_dirty = False
        self._derived_outputs_stale = False
        if self._is_cached_recompute_supported_for_config(config):
            self._cached_recompute_message = (
                "Ephemeris cached. Station and comms changes can refresh access without a full rerun."
            )
        else:
            self._cached_recompute_message = (
                "Cached recompute is unavailable because ground-station analysis is disabled."
            )
        self._store_access_series(result)
        self._initialize_contact_elevation_filter()
        self._apply_contact_statistics_filter()
        self._set_run_button_state("success")
        self._mission_reference_epoch = self._visual_reference_epoch
        self._mission_orbit_period_s = float(result.orbit_period_seconds or 0.0)
        self._mission_ground_track = result.ephemeris
        self._mission_window_fractions = (0.0, 1.0)
        if self.mission_window_slider:
            self.mission_window_slider.setValue(
                self.mission_window_slider.maximum()
            )
        self._update_mission_window_label()
        self._refresh_mission_globe()
        self._refresh_visualization_pass_tabs(result)
        self._update_downlink_summary()
        self._update_orbit_summary(result)
        if hasattr(self, "_apply_sso_if_enabled"):
            self._apply_sso_if_enabled()  # type: ignore[attr-defined]
        self._update_cached_access_ui_state()

    def _refresh_cached_access_from_inputs(self, *, silent: bool) -> bool:
        """Try to rebuild access outputs from the cached trajectory."""
        try:
            if self._last_ephemeris is None or self._last_ephemeris_signature is None:
                raise ValueError("No cached ephemeris is available for recompute.")
            if silent:
                if self._is_dirty:
                    self._update_cached_access_ui_state()
                    return False
                perform_ground_passes = getattr(self, "ground_pass_checkbox", None)
                ground_pass_enabled = (
                    perform_ground_passes.isChecked()
                    if perform_ground_passes is not None
                    else True
                )
                if ground_pass_enabled:
                    enabled_indices = getattr(self, "_get_enabled_station_indices", lambda: [])()
                    if not getattr(self, "_station_presets", None) or not enabled_indices:
                        self._update_cached_access_ui_state()
                        return False
            config, stations = self._collect_run_inputs()
            if not self._is_cached_recompute_supported_for_config(config):
                raise ValueError(
                    "Cached recompute is only supported when ground-station analysis is enabled."
                )
            if self._ephemeris_signature_for_config(config) != self._last_ephemeris_signature:
                raise ValueError(
                    "Current orbit or dynamics settings no longer match the cached trajectory. Run the mission scenario again."
                )
            derived_access = derive_access_results_from_ephemeris(
                ephemeris=self._last_ephemeris,
                config=config,
                stations=stations,
            )
            result = analysis_result_from_components(
                ephemeris=self._last_ephemeris,
                derived_access=derived_access,
            )
        except ValueError as exc:
            if not silent:
                QMessageBox.warning(self, "Cached Recompute Unavailable", str(exc))
            self._update_cached_access_ui_state()
            return False
        self._apply_analysis_result(
            config=config,
            result=result,
            stations=stations,
            update_ephemeris_cache=False,
        )
        return True

    def _handle_refresh_cached_access_clicked(self) -> None:
        """Recompute access outputs from the cached trajectory."""
        self._refresh_cached_access_from_inputs(silent=False)

    def _handle_export_cached_package_clicked(self) -> None:
        """Export the cached trajectory package as a compressed NPZ plus lean JSON sidecar."""
        if self._last_ephemeris is None:
            QMessageBox.warning(
                self,
                "No Cached Trajectory",
                "Run or import a scenario first so there is cached ephemeris to export.",
            )
            return
        outputs_dir = Path(__file__).resolve().parents[2] / "outputs"
        outputs_dir.mkdir(exist_ok=True)
        timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
        default_path = outputs_dir / f"cached_trajectory_{timestamp}.npz"
        filename, _ = QFileDialog.getSaveFileName(
            self,
            "Export Cached Trajectory Package",
            str(default_path),
            "NPZ files (*.npz);;All files (*.*)",
        )
        if not filename:
            return
        stations, enabled_station_names = self._collect_station_package_payload()
        try:
            npz_path, sidecar_path = export_cached_trajectory_package(
                ephemeris=self._last_ephemeris,
                settings=self._collect_scenario_package_settings(),
                stations=stations,
                enabled_station_names=enabled_station_names,
                output_path=Path(filename),
            )
        except Exception as exc:
            QMessageBox.critical(
                self,
                "Export Failed",
                f"Failed to export cached trajectory package:\n{exc}",
            )
            raise
        QMessageBox.information(
            self,
            "Export Successful",
            f"Trajectory exported to:\n{npz_path}\n\nSidecar exported to:\n{sidecar_path}",
        )

    def _handle_import_cached_package_clicked(self) -> None:
        """Import a cached trajectory package from an NPZ file or legacy OEM package."""
        outputs_dir = Path(__file__).resolve().parents[2] / "outputs"
        outputs_dir.mkdir(exist_ok=True)
        filename, _ = QFileDialog.getOpenFileName(
            self,
            "Import Cached Trajectory Package",
            str(outputs_dir),
            "Trajectory packages (*.npz *.oem *.json);;All files (*.*)",
        )
        if not filename:
            return
        try:
            package = import_cached_trajectory_package(filename)
            config, stations = self._apply_imported_scenario_package(package)
            derived_access = derive_access_results_from_ephemeris(
                ephemeris=package.ephemeris,
                config=config,
                stations=stations,
            )
            result = analysis_result_from_components(
                ephemeris=package.ephemeris,
                derived_access=derived_access,
            )
        except Exception as exc:
            QMessageBox.critical(
                self,
                "Import Failed",
                f"Failed to import cached trajectory package:\n{exc}",
            )
            raise
        self._apply_analysis_result(
            config=config,
            result=result,
            stations=stations,
            update_ephemeris_cache=True,
        )

    def _handle_run_clicked(self) -> None:
        """Build the analysis configuration and execute it."""
        self._set_run_button_state("running")
        try:
            config, stations = self._collect_run_inputs()
        except ValueError as exc:  # pragma: no cover - GUI validation
            QMessageBox.warning(self, "Invalid Input", str(exc))
            self._set_run_button_state("dirty")
            return
        self._active_station_lookup = {station.name: station for station in stations}
        self._clear_visualization_display("Running analysis…")
        self._show_run_progress_ui()
        self._start_analysis_worker(config, stations)

    def _show_run_progress_ui(self) -> None:
        """Display the in-button progress bar state."""
        # Reset ETA estimator for a fresh run.
        self._run_progress_last_t = None
        self._run_progress_last_pct = None
        self._run_progress_rate_ewma = None
        self._run_progress_start_t = None
        self._run_progress_start_pct = None
        if self.run_progress:
            self.run_progress.setValue(0)
            self.run_progress.setFormat("0% - --:-- remaining")
            self.run_progress.show()
        self.run_button.setEnabled(False)
        self.run_button.hide()
        if getattr(self, "stop_button", None) is not None:
            self.stop_button.setEnabled(True)
            self.stop_button.show()

    def _hide_run_progress_ui(self) -> None:
        """Restore the run button once analysis completes."""
        if self.run_progress:
            self.run_progress.hide()
        # Reset ETA estimator after completion/cancellation.
        self._run_progress_last_t = None
        self._run_progress_last_pct = None
        self._run_progress_rate_ewma = None
        self._run_progress_start_t = None
        self._run_progress_start_pct = None
        self.run_button.show()
        self.run_button.setEnabled(True)
        if getattr(self, "stop_button", None) is not None:
            self.stop_button.setEnabled(False)

    def _start_analysis_worker(
        self, config: AnalysisConfig, stations: list[GroundStationConfig]
    ) -> None:
        """Kick off the background analysis worker."""
        self._pending_config = config
        self._analysis_thread = QThread(self)
        self._analysis_worker = AnalysisWorker(config, stations)
        self._analysis_worker.moveToThread(self._analysis_thread)
        self._analysis_thread.started.connect(self._analysis_worker.run)
        self._analysis_worker.progress.connect(self._update_run_progress)
        self._analysis_worker.finished.connect(self._handle_analysis_success)
        self._analysis_worker.error.connect(self._handle_analysis_error)
        self._analysis_worker.finished.connect(self._analysis_thread.quit)
        self._analysis_worker.error.connect(self._analysis_thread.quit)
        self._analysis_worker.finished.connect(self._analysis_worker.deleteLater)
        self._analysis_worker.error.connect(self._analysis_worker.deleteLater)
        self._analysis_thread.finished.connect(self._cleanup_analysis_thread)
        self._analysis_thread.start()

    def _handle_stop_clicked(self) -> None:
        """Request cancellation of the currently running analysis, if any."""
        worker = getattr(self, "_analysis_worker", None)
        if worker is not None:
            worker.request_cancel()

    def _cleanup_analysis_thread(self) -> None:
        """Release worker references after the thread stops."""
        self._analysis_thread = None
        self._analysis_worker = None

    def _update_run_progress(self, value: float) -> None:
        """Update the progress indicator."""
        if not self.run_progress:
            return
        clamped = max(0, min(100, int(value)))
        now = time.monotonic()

        # Start-of-run reference (first time we see a non-zero progress update).
        if clamped > 0 and self._run_progress_start_t is None:
            self._run_progress_start_t = now
            self._run_progress_start_pct = float(clamped)

        # Update EWMA progress rate (percent per second).
        if self._run_progress_last_t is None or self._run_progress_last_pct is None:
            self._run_progress_last_t = now
            self._run_progress_last_pct = float(clamped)
        else:
            dt = float(now - self._run_progress_last_t)
            dp = float(clamped) - float(self._run_progress_last_pct)
            # Only learn from forward progress.
            if dt > 0.0 and dp > 0.0:
                inst_rate = dp / dt
                if self._run_progress_rate_ewma is None:
                    self._run_progress_rate_ewma = inst_rate
                else:
                    alpha = 0.25
                    self._run_progress_rate_ewma = (
                        alpha * inst_rate + (1.0 - alpha) * self._run_progress_rate_ewma
                    )
            self._run_progress_last_t = now
            self._run_progress_last_pct = float(clamped)

        # Estimate remaining wall-clock time.
        eta = "--:--"
        if clamped >= 100:
            eta = "00:00"
        else:
            rem_s: float | None = None

            # Prefer a stable average-rate ETA from the run start.
            if (
                self._run_progress_start_t is not None
                and self._run_progress_start_pct is not None
            ):
                elapsed = float(now - self._run_progress_start_t)
                progressed = float(clamped) - float(self._run_progress_start_pct)
                if elapsed >= 0.5 and progressed > 0.0:
                    avg_rate = progressed / elapsed  # %/s
                    if avg_rate > 1e-6:
                        rem_s = float((100.0 - float(clamped)) / avg_rate)

            # Fallback: EWMA instantaneous rate.
            if rem_s is None:
                rate = self._run_progress_rate_ewma
                if rate is not None and rate > 1e-6 and clamped > 0:
                    rem_s = float((100.0 - float(clamped)) / rate)

            if rem_s is not None:
                if rem_s < 0.0 or not (rem_s == rem_s):  # NaN check
                    rem_s = 0.0
                # Avoid showing 00:00 unless actually complete: round up to the next second.
                total_seconds = int(rem_s + 0.999)
                if total_seconds == 0:
                    total_seconds = 1
                mm = total_seconds // 60
                ss = total_seconds % 60
                eta = f"{mm:02d}:{ss:02d}"

        self.run_progress.setValue(clamped)
        self.run_progress.setFormat(f"{clamped:.0f}% - {eta} remaining")

    def _handle_analysis_success(self, result: AnalysisResult) -> None:
        """Handle successful completion of the analysis."""
        config = self._pending_config
        self._pending_config = None
        if config is None:
            raise ValueError("Missing pending config for completed analysis.")
        stations = list(self._active_station_lookup.values())
        self._apply_analysis_result(
            config=config,
            result=result,
            stations=stations,
            update_ephemeris_cache=True,
        )
        self._update_run_progress(100.0)
        self._hide_run_progress_ui()

    def _show_instantaneous_aoa_figure(self, result: AnalysisResult) -> None:
        """Show a standalone Matplotlib figure of instantaneous AoA at full resolution."""
        times_s = getattr(result, "timeline_seconds", None)
        aoa_deg = getattr(result, "angle_of_attack_deg", None)
        if not times_s or not aoa_deg:
            return
        if len(times_s) != len(aoa_deg):
            return

        # Import lazily to avoid matplotlib startup cost unless requested.
        from matplotlib.backends.backend_qtagg import FigureCanvasQTAgg
        from matplotlib.figure import Figure

        from matplotlib.backends.backend_qtagg import NavigationToolbar2QT

        if getattr(self, "_aoa_window", None) is None:
            win = QMainWindow(self)
            win.setWindowTitle("Instantaneous AoA (all points)")
            central = QWidget(win)
            layout = QVBoxLayout(central)
            fig = Figure(figsize=(8, 4.5), tight_layout=True)
            canvas = FigureCanvasQTAgg(fig)
            layout.addWidget(canvas)
            toolbar = NavigationToolbar2QT(canvas, win)
            layout.addWidget(toolbar)
            win.setCentralWidget(central)

            self._aoa_window = win
            self._aoa_canvas = canvas
            self._aoa_axes = fig.add_subplot(111)

            def _clear_refs(*_args) -> None:
                self._aoa_window = None
                self._aoa_canvas = None
                self._aoa_axes = None

            win.destroyed.connect(_clear_refs)  # type: ignore[arg-type]

        ax = self._aoa_axes
        canvas = self._aoa_canvas
        win = self._aoa_window
        if ax is None or canvas is None or win is None:
            return

        times_h = [float(t) / 3600.0 for t in times_s]
        # Plot all points (no UI downsampling) to support debugging.
        ax.clear()
        ax.plot(times_h, aoa_deg, color="#8e24aa", linewidth=0.8)
        ax.set_title("Instantaneous Angle of Attack (full resolution)")
        ax.set_xlabel("Time since start (h)")
        ax.set_ylabel("AoA (deg)")
        ax.grid(True, alpha=0.25)
        canvas.draw_idle()

        win.show()
        win.raise_()
        win.activateWindow()

    def _handle_analysis_error(self, message: str) -> None:
        """Handle errors raised by the analysis worker."""
        self._pending_config = None
        self._set_run_button_state("dirty")
        self._hide_run_progress_ui()
        self._update_cached_access_ui_state()
        # Treat user-initiated cancellation as a non-fatal condition.
        if message.strip().lower().startswith("analysis cancelled"):
            return
        QMessageBox.critical(self, "Analysis Error", message)

    def _build_config_from_inputs(
        self,
        primary_station: GroundStationConfig | None,
        *,
        enable_ground_pass_analysis: bool,
    ) -> AnalysisConfig:
        """Translate widget state into a strongly-typed config."""
        start_dt = self._qdatetime_to_utc(self.start_datetime)
        end_dt = self._qdatetime_to_utc(self.end_datetime)
        if end_dt <= start_dt:
            raise ValueError("End time must be later than the start time.")
        if enable_ground_pass_analysis and primary_station is None:
            raise ValueError(
                "Select at least one ground station when ground-station analysis is enabled."
            )
        ground = primary_station
        altitude_km = float(self.altitude_input.value()) if getattr(self, "altitude_input", None) is not None else 0.0
        ecc = float(self.ecc_input.value())
        # Derive SMA from geodetic altitude above WGS-84.
        sma_km = float(getattr(self, "_WGS84_EQUATORIAL_RADIUS_KM", 6378.137)) + altitude_km

        # If SSO is enabled, override inclination using orbital mechanics.
        inc_deg = float(self.inc_input.value())
        if getattr(self, "sso_checkbox", None) is not None and self.sso_checkbox.isChecked():
            computed = self._compute_sso_inclination_deg(altitude_km=altitude_km, eccentricity=ecc)  # type: ignore[attr-defined]
            if computed is not None:
                inc_deg = float(computed)
                # Keep the UI in sync with the applied inclination.
                try:
                    self.inc_input.blockSignals(True)
                    self.inc_input.setValue(inc_deg)
                finally:
                    self.inc_input.blockSignals(False)
        orbit = OrbitConfig(
            semi_major_axis_km=sma_km,
            eccentricity=ecc,
            inclination_deg=inc_deg,
            raan_deg=self.raan_input.value(),
            arg_perigee_deg=self.argp_input.value(),
            mean_anomaly_deg=self.mean_anom_input.value(),
        )
        comms_pointing_widget = getattr(self, "comms_pointing_mode_combo", None)
        if comms_pointing_widget is None:
            raise ValueError("Comms pointing mode control is not available.")
        comms_pointing_label = str(comms_pointing_widget.currentText()).strip()
        if comms_pointing_label == "Prograde Pointing":
            comms_pointing_mode = "prograde_pointing"
        elif comms_pointing_label == "Free to Roll":
            comms_pointing_mode = "free_to_roll"
        elif comms_pointing_label == "Constrained AoA":
            comms_pointing_mode = "constrained_aoa"
        else:
            raise ValueError(f"Unsupported comms pointing mode: {comms_pointing_label!r}")
        comms_pointing_aoa_widget = getattr(self, "comms_pointing_aoa_limit_input", None)
        if comms_pointing_aoa_widget is None:
            raise ValueError("Comms AoA limit control is not available.")
        comms_pointing_aoa_limit_deg = float(comms_pointing_aoa_widget.value())
        # The propagated attitude remains in the nominal prograde convention.
        # Station steering for link-budget antenna geometry is applied separately
        # according to the selected comms pointing mode.
        attitude_mode = "prograde"
        enable_contact_switching = False

        propagator_widget = getattr(self, "propagator_combo", None)
        propagator_label = (
            str(propagator_widget.currentText()).strip() if propagator_widget is not None else "Brouwer-Lyddane"
        )
        if propagator_label == "Brouwer-Lyddane":
            propagator_type = "brouwer_lyddane"
        elif propagator_label == "Keplerian":
            propagator_type = "keplerian"
        else:
            raise ValueError(f"Unsupported propagator selection: {propagator_label!r}")

        propagation = PropagationConfig(
            propagator_type=propagator_type,
            min_elevation_deg=0.0,
            contact_elevation_deg=float(
                getattr(self, "contact_attitude_elevation_input", None).value()
            )
            if getattr(self, "contact_attitude_elevation_input", None) is not None
            else 10.0,
            sample_step_seconds=float(self.sample_step_input.value()),
            attitude_mode=attitude_mode,
            enable_contact_attitude_switching=enable_contact_switching,
            comms_pointing_mode=comms_pointing_mode,
            comms_pointing_aoa_limit_deg=comms_pointing_aoa_limit_deg,
            # BrouwerLyddane's closed-form mean-orbit correction (applied in
            # access_analysis) replaces the retired numerical mean-IC solver.
            apply_mean_orbit_correction=True,
        )
        scenario = ScenarioConfig(start_time=start_dt, end_time=end_dt)
        options = AnalysisOptions(
            compute_ground_station_passes=enable_ground_pass_analysis
        )
        return AnalysisConfig(
            ground_station=ground,
            orbit=orbit,
            propagation=propagation,
            scenario=scenario,
            options=options,
        )

    def _qdatetime_to_utc(self, widget) -> datetime:
        """Convert a QDateTime widget value into a timezone-aware datetime."""
        qdt = widget.dateTime()
        py_dt = None
        for attr in ("toPyDateTime", "toPython"):
            converter = getattr(qdt, attr, None)
            if callable(converter):
                py_dt = converter()
                break
        if py_dt is None:
            # Fallback to epoch conversion if direct Python conversion is unavailable.
            py_dt = datetime.fromtimestamp(qdt.toSecsSinceEpoch(), tz=timezone.utc)
        if py_dt.tzinfo is None:
            py_dt = py_dt.replace(tzinfo=timezone.utc)
        else:
            py_dt = py_dt.astimezone(timezone.utc)
        return py_dt

    def _populate_results_table(self, result) -> None:
        """Render the per-pass statistics inside the table widget."""
        passes = sorted(result.passes, key=lambda item: item.aos)
        self.results_table.setRowCount(len(passes))
        for row, item in enumerate(passes):
            display_index = row + 1
            self.results_table.setItem(row, 0, QTableWidgetItem(str(display_index)))
            station_name = item.station_name or "Ground Station"
            self.results_table.setItem(row, 1, QTableWidgetItem(station_name))
            self.results_table.setItem(
                row, 2, QTableWidgetItem(item.aos.strftime("%d-%b-%Y %H:%M:%S"))
            )
            self.results_table.setItem(
                row, 3, QTableWidgetItem(item.los.strftime("%d-%b-%Y %H:%M:%S"))
            )
            self.results_table.setItem(
                row, 4, QTableWidgetItem(f"{item.duration_minutes:.2f}")
            )
            self.results_table.setItem(
                row, 5, QTableWidgetItem(f"{item.max_elevation_deg:.1f}")
            )
            max_slew = float(getattr(item, "max_sc_slew_rate_deg_s", float("nan")))
            max_slew_text = f"{max_slew:.3f}" if math.isfinite(max_slew) else "N/A"
            self.results_table.setItem(row, 6, QTableWidgetItem(max_slew_text))
            volume_gbit = self._get_per_pass_volume_gbit(item)
            if volume_gbit is None:
                volume_text = "—"
            else:
                fmt = getattr(self, "_format_data_quantity", None)
                volume_text = fmt(volume_gbit) if callable(fmt) else f"{volume_gbit:.4f}"
            self.results_table.setItem(row, 7, QTableWidgetItem(volume_text))
            close_elev = self._get_per_pass_link_close_elevation(item)
            close_elev_text = f"{close_elev:.1f}" if close_elev is not None else "—"
            self.results_table.setItem(row, 8, QTableWidgetItem(close_elev_text))

    def _get_per_pass_volume_gbit(self, pass_stat) -> float | None:
        """Return integrated downlink data volume (Gbit) for a single pass, or None if unavailable."""
        access_series = getattr(self, "_latest_access_series", None)
        station_rate_lookup = getattr(self, "_latest_station_rate_series", None)
        current_config = getattr(self, "_current_config", None)
        if not access_series or not station_rate_lookup or current_config is None:
            return None
        time_seconds = np.asarray(access_series.get("time_seconds", []), dtype=float)
        if time_seconds.size < 2:
            return None
        station_name = getattr(pass_stat, "station_name", None)
        if not station_name or station_name not in station_rate_lookup:
            return None
        rates_array = np.asarray(station_rate_lookup[station_name], dtype=float)
        if rates_array.size != time_seconds.size:
            return None
        start_time = getattr(getattr(current_config, "scenario", None), "start_time", None)
        if start_time is None:
            return None
        start_axis = float(time_seconds[0])
        end_axis = float(time_seconds[-1])
        aos_seconds = (pass_stat.aos - start_time).total_seconds()
        los_seconds = (pass_stat.los - start_time).total_seconds()
        pass_start = max(start_axis, float(aos_seconds))
        pass_end = min(end_axis, float(los_seconds))
        if pass_end <= pass_start:
            return None
        integrate = getattr(self, "_integrate_data_volume_interval", None)
        if not callable(integrate):
            return None
        volume_gbit = integrate(time_seconds, rates_array, pass_start, pass_end)
        return float(volume_gbit) if volume_gbit > 0 else None

    def _get_per_pass_link_close_elevation(self, pass_stat) -> float | None:
        """Minimum elevation (deg) where the link is active in the pass window, or None."""
        access_series = getattr(self, "_latest_access_series", None)
        station_rate_lookup = getattr(self, "_latest_station_rate_series", None)
        current_config = getattr(self, "_current_config", None)
        if not access_series or not station_rate_lookup or current_config is None:
            return None
        time_seconds = np.asarray(access_series.get("time_seconds", []), dtype=float)
        if time_seconds.size < 2:
            return None
        station_name = getattr(pass_stat, "station_name", None)
        if not station_name or station_name not in station_rate_lookup:
            return None
        rates_array = np.asarray(station_rate_lookup[station_name], dtype=float)
        if rates_array.size != time_seconds.size:
            return None
        station_series: dict = access_series.get("station_series", {})
        elevations = np.asarray(station_series.get(station_name, []), dtype=float)
        if elevations.size != time_seconds.size:
            return None
        start_time = getattr(getattr(current_config, "scenario", None), "start_time", None)
        if start_time is None:
            return None
        start_axis = float(time_seconds[0])
        end_axis = float(time_seconds[-1])
        aos_seconds = (pass_stat.aos - start_time).total_seconds()
        los_seconds = (pass_stat.los - start_time).total_seconds()
        pass_start = max(start_axis, float(aos_seconds))
        pass_end = min(end_axis, float(los_seconds))
        if pass_end <= pass_start:
            return None
        in_pass = (time_seconds >= pass_start) & (time_seconds <= pass_end)
        link_on = in_pass & (rates_array > 0.0) & np.isfinite(elevations)
        if not np.any(link_on):
            return None
        return float(np.min(elevations[link_on]))

    def _export_results_table(self, format_type: str) -> None:
        """Export the results table to NPZ or XLSX format."""
        if self.results_table.rowCount() == 0:
            QMessageBox.warning(
                self, "No Data", "No results to export. Please run an analysis first."
            )
            return

        # Generate timestamp for filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        outputs_dir = Path(__file__).resolve().parents[2] / "outputs"
        outputs_dir.mkdir(exist_ok=True)

        # Collect headers
        headers = []
        for col in range(self.results_table.columnCount()):
            header_item = self.results_table.horizontalHeaderItem(col)
            headers.append(header_item.text() if header_item else f"Column {col}")

        # Collect data rows
        data = []
        for row in range(self.results_table.rowCount()):
            row_data = []
            for col in range(self.results_table.columnCount()):
                item = self.results_table.item(row, col)
                row_data.append(item.text() if item else "")
            data.append(row_data)

        try:
            if format_type == "npz":
                filename = outputs_dir / f"contact_statistics_{timestamp}.npz"
                payload = {
                    "schema_version": np.array([1], dtype=int),
                    "headers": np.array(headers, dtype=object),
                    "rows": np.array(data, dtype=object),
                }
                np.savez_compressed(filename, **payload)
                QMessageBox.information(
                    self, "Export Successful", f"Data exported to:\n{filename}"
                )
            elif format_type == "xlsx":
                try:
                    import openpyxl
                    from openpyxl.styles import Font

                    filename = outputs_dir / f"contact_statistics_{timestamp}.xlsx"
                    wb = openpyxl.Workbook()
                    ws = wb.active
                    ws.title = "Contact Statistics"

                    # Write headers with bold font
                    for col_idx, header in enumerate(headers, start=1):
                        cell = ws.cell(row=1, column=col_idx, value=header)
                        cell.font = Font(bold=True)

                    # Write data
                    for row_idx, row_data in enumerate(data, start=2):
                        for col_idx, value in enumerate(row_data, start=1):
                            ws.cell(row=row_idx, column=col_idx, value=value)

                    # Auto-adjust column widths
                    for col in ws.columns:
                        max_length = 0
                        column = col[0].column_letter
                        for cell in col:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        adjusted_width = min(max_length + 2, 50)
                        ws.column_dimensions[column].width = adjusted_width

                    wb.save(filename)
                    QMessageBox.information(
                        self, "Export Successful", f"Data exported to:\n{filename}"
                    )
                except ImportError:
                    QMessageBox.critical(
                        self,
                        "Missing Dependency",
                        "openpyxl is required for XLSX export.\n"
                        "Install it with: pip install openpyxl",
                    )
                    raise
        except Exception as e:
            QMessageBox.critical(
                self, "Export Failed", f"Failed to export data:\n{str(e)}"
            )
            raise

    def _update_summary_label(self, result) -> None:
        """Show the aggregated statistics to the user."""
        # If ground-station analysis was disabled, show a dedicated message.
        if (
            getattr(self, "_current_config", None) is not None
            and getattr(self._current_config, "options", None) is not None
            and not self._current_config.options.compute_ground_station_passes
        ):
            self.summary_label.setText(
                "Ground-station pass analysis was disabled for this run."
            )
            if self.analysis_context_label is not None:
                self.analysis_context_label.setText("Ground-station analysis disabled.")
            return

        summary = result.summary
        base_text = (
            f"Passes: {summary.total_passes} | Total Access: {summary.total_access_minutes:.1f} min | "
            f"Coverage: {summary.coverage_percent:.2f}% | Avg: {summary.avg_duration_minutes:.2f} min | "
            f"Min: {summary.min_duration_minutes:.2f} min | Max: {summary.max_duration_minutes:.2f} min"
        )
        mean_ic_report = getattr(result, "mean_ic_report", None)
        if isinstance(mean_ic_report, str) and mean_ic_report.strip():
            base_text += f" | {mean_ic_report}"
        station_summaries = getattr(result, "station_summaries", [])
        if len(station_summaries) > 1:
            breakdown = ", ".join(
                f"{entry.station_name}: {entry.total_access_minutes:.1f} min"
                for entry in station_summaries
            )
            base_text += f" | Stations: {breakdown}"
        elif station_summaries:
            entry = station_summaries[0]
            base_text += f" | Station: {entry.station_name}"
        if self._use_link_close_for_contact_windows():
            base_text += " | Windows: dynamic link close"
        self.summary_label.setText(base_text)
        if self.analysis_context_label is not None:
            if station_summaries:
                names_list = [entry.station_name for entry in station_summaries]
                display_names = names_list[:5]
                if len(names_list) > 5:
                    display_names.append("…")
                names = ", ".join(display_names)
                self.analysis_context_label.setText(
                    f"Active stations ({len(station_summaries)}): {names}"
                )
            else:
                self.analysis_context_label.setText("Active stations: manual entry")

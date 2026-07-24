"""Orbit summary tab mixin for visualizing basic orbit metrics over time."""

from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List

import numpy as np
import pyqtgraph as pg
from PySide6.QtWidgets import (
    QComboBox,
    QFileDialog,
    QHBoxLayout,
    QLabel,
    QMessageBox,
    QPushButton,
    QVBoxLayout,
    QWidget,
)

from cosmic_toolbox.analyses.orbit_summary import (
    compute_orbit_averaged_orbit_summary as _compute_orbit_averaged,
    compute_series_stats as _compute_stats,
)
from cosmic_toolbox.models import AnalysisResult
from cosmic_toolbox.tools.orbit_summary_npz import save_orbit_summary_npz

WGS84_EQUATORIAL_RADIUS_KM = 6378.137


class OrbitSummaryTabMixin:
    """Provides a tab with orbit summary plots (altitude and basic elements)."""

    def _build_orbit_summary_tab(self) -> QWidget:
        """Create the Orbit Summary tab with secular and instantaneous element plots."""
        # Avoid heavy antialiasing for dense time series to keep things responsive.
        pg.setConfigOptions(antialias=False)

        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.setContentsMargins(4, 4, 4, 4)

        # --- Ephemeris export ---
        export_row = QHBoxLayout()
        export_row.addWidget(QLabel("Ephemeris export:"))
        self.ephemeris_frame_combo = QComboBox()
        self.ephemeris_frame_combo.addItems(["ECI (EME2000)", "ECEF (ITRF)"])
        export_row.addWidget(self.ephemeris_frame_combo)
        self.ephemeris_export_npz_button = QPushButton("Export NPZ")
        self.ephemeris_export_npz_button.clicked.connect(  # type: ignore[arg-type]
            lambda: self._export_ephemeris_series("npz")
        )
        export_row.addWidget(self.ephemeris_export_npz_button)
        self.ephemeris_export_xlsx_button = QPushButton("Export XLSX")
        self.ephemeris_export_xlsx_button.clicked.connect(  # type: ignore[arg-type]
            lambda: self._export_ephemeris_series("xlsx")
        )
        export_row.addWidget(self.ephemeris_export_xlsx_button)
        self.orbit_summary_export_npz_button = QPushButton("Export Orbit Summary NPZ")
        self.orbit_summary_export_npz_button.clicked.connect(  # type: ignore[arg-type]
            self._export_orbit_summary_npz
        )
        export_row.addWidget(self.orbit_summary_export_npz_button)
        export_row.addStretch(1)
        layout.addLayout(export_row)
        self._set_ephemeris_export_enabled(False)
        self._set_orbit_summary_export_enabled(False)

        # --- Short-period averaged (Brouwer–Lyddane-style) elements ---
        self.orbit_secular_combo = QComboBox()
        self.orbit_secular_combo.addItems(
            [
                "Geodetic altitude (orbit-averaged)",
                "Semi-major axis and apsides (Brouwer–Lyddane mean)",
                "Eccentricity (Brouwer–Lyddane mean)",
                "Inclination (Brouwer–Lyddane mean)",
                "Argument of perigee (Brouwer–Lyddane mean)",
                "Orbital period (Brouwer–Lyddane mean)",
                "Angle of attack (orbit-averaged)",
            ]
        )
        self.orbit_secular_combo.currentIndexChanged.connect(  # type: ignore[arg-type]
            self._handle_secular_metric_changed
        )
        layout.addWidget(self.orbit_secular_combo)

        self.orbit_secular_plot = pg.PlotWidget(title="Geodetic Altitude (orbit-averaged)")
        self.orbit_secular_plot.setLabel("bottom", "Time since start", units="h")
        self.orbit_secular_plot.setLabel("left", "Geodetic altitude", units="km")
        self.orbit_secular_plot.showGrid(x=True, y=True, alpha=0.3)
        self._orbit_secular_legend = self.orbit_secular_plot.addLegend()
        layout.addWidget(self.orbit_secular_plot, stretch=1)

        # --- Instantaneous (osculating) elements ---
        self.orbit_instant_combo = QComboBox()
        self.orbit_instant_combo.addItems(
            [
                "Geodetic altitude (instantaneous)",
                "Semi-major axis and apsides (instantaneous)",
                "Eccentricity (instantaneous)",
                "Inclination (instantaneous)",
                "Argument of perigee (instantaneous)",
                "True anomaly (instantaneous)",
                "Orbital period (instantaneous)",
                "Angle of attack (instantaneous)",
            ]
        )
        self.orbit_instant_combo.currentIndexChanged.connect(  # type: ignore[arg-type]
            self._handle_instant_metric_changed
        )
        layout.addWidget(self.orbit_instant_combo)

        self.orbit_instant_plot = pg.PlotWidget(title="Geodetic Altitude (instantaneous)")
        self.orbit_instant_plot.setLabel("bottom", "Time since start", units="h")
        self.orbit_instant_plot.setLabel("left", "Geodetic altitude", units="km")
        self.orbit_instant_plot.showGrid(x=True, y=True, alpha=0.3)
        self._orbit_instant_legend = self.orbit_instant_plot.addLegend()
        layout.addWidget(self.orbit_instant_plot, stretch=1)

        # Cache for latest time series (instantaneous + secular).
        self._orbit_summary_data: Dict[str, np.ndarray] | None = None

        # Initial empty state.
        self._update_orbit_summary(None)
        return tab

    def _set_ephemeris_export_enabled(self, enabled: bool) -> None:
        """Enable/disable the Orbit Summary ephemeris export buttons."""
        for attr in ("ephemeris_export_npz_button", "ephemeris_export_xlsx_button"):
            button = getattr(self, attr, None)
            if button is not None:
                button.setEnabled(bool(enabled))

    def _set_orbit_summary_export_enabled(self, enabled: bool) -> None:
        button = getattr(self, "orbit_summary_export_npz_button", None)
        if button is not None:
            button.setEnabled(bool(enabled))

    def _export_ephemeris_series(self, format_type: str) -> None:
        """Export the ephemeris (ECI/ECEF r,v) to NPZ or XLSX."""
        result = getattr(self, "_last_result", None)
        eph = result.ephemeris if result is not None else None
        if eph is None or eph.timestamps_unix is None or eph.timestamps_unix.size == 0:
            QMessageBox.warning(
                self,
                "No Data",
                "No ephemeris data to export. Please run an analysis first.",
            )
            return

        frame_mode = "ECI" if self.ephemeris_frame_combo.currentIndex() == 0 else "ECEF"

        timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
        outputs_dir = Path(__file__).resolve().parents[3] / "outputs"
        outputs_dir.mkdir(exist_ok=True)
        filename_base = outputs_dir / f"ephemeris_{frame_mode}_{timestamp}"

        headers = [
            "timestamp_utc",
            "seconds_from_start",
            "x_km",
            "y_km",
            "z_km",
            "vx_km_s",
            "vy_km_s",
            "vz_km_s",
        ]

        ts_arr = eph.timestamps_unix
        t0_unix = float(ts_arr[0])
        pos_arr = eph.eci_pos_km if frame_mode == "ECI" else eph.ecef_pos_km
        vel_arr = eph.eci_vel_km_s if frame_mode == "ECI" else eph.ecef_vel_km_s
        if pos_arr is None or vel_arr is None:
            QMessageBox.warning(self, "No Data", "Ephemeris does not contain the requested frame data.")
            return

        rows: list[list[object]] = []
        for i in range(len(ts_arr)):
            t_unix = float(ts_arr[i])
            dt_s = t_unix - t0_unix
            ts_dt = datetime.fromtimestamp(t_unix, tz=timezone.utc)
            rows.append(
                [
                    ts_dt.strftime("%d-%b-%Y %H:%M:%S"),
                    dt_s,
                    float(pos_arr[i, 0]),
                    float(pos_arr[i, 1]),
                    float(pos_arr[i, 2]),
                    float(vel_arr[i, 0]),
                    float(vel_arr[i, 1]),
                    float(vel_arr[i, 2]),
                ]
            )

        try:
            if format_type == "npz":
                filename = Path(str(filename_base) + ".npz")
                columns = list(zip(*rows)) if rows else []
                payload: dict[str, np.ndarray] = {
                    "schema_version": np.array([1], dtype=int),
                    "frame_mode": np.array([frame_mode], dtype=object),
                    "headers": np.array(headers, dtype=object),
                }
                if columns:
                    payload.update(
                        {
                            "timestamp_utc": np.array(columns[0], dtype=object),
                            "seconds_from_start": np.asarray(columns[1], dtype=float),
                            "x_km": np.asarray(columns[2], dtype=float),
                            "y_km": np.asarray(columns[3], dtype=float),
                            "z_km": np.asarray(columns[4], dtype=float),
                            "vx_km_s": np.asarray(columns[5], dtype=float),
                            "vy_km_s": np.asarray(columns[6], dtype=float),
                            "vz_km_s": np.asarray(columns[7], dtype=float),
                        }
                    )
                np.savez_compressed(filename, **payload)
                QMessageBox.information(
                    self, "Export Successful", f"Ephemeris exported to:\n{filename}"
                )
                return

            if format_type == "xlsx":
                try:
                    import openpyxl
                    from openpyxl.styles import Font

                    filename = Path(str(filename_base) + ".xlsx")
                    wb = openpyxl.Workbook()
                    ws = wb.active
                    ws.title = "Ephemeris"

                    for col_idx, header in enumerate(headers, start=1):
                        cell = ws.cell(row=1, column=col_idx, value=header)
                        cell.font = Font(bold=True)

                    for row_idx, row_data in enumerate(rows, start=2):
                        for col_idx, value in enumerate(row_data, start=1):
                            ws.cell(row=row_idx, column=col_idx, value=value)

                    for col in ws.columns:
                        max_length = 0
                        column = col[0].column_letter
                        for cell in col:
                            if cell.value is not None:
                                max_length = max(max_length, len(str(cell.value)))
                        ws.column_dimensions[column].width = min(max_length + 2, 50)

                    wb.save(filename)
                    QMessageBox.information(
                        self, "Export Successful", f"Ephemeris exported to:\n{filename}"
                    )
                    return
                except ImportError:
                    QMessageBox.critical(
                        self,
                        "Missing Dependency",
                        "openpyxl is required for XLSX export.\n"
                        "Install it with: pip install openpyxl",
                    )
                    raise

            QMessageBox.warning(
                self,
                "Export Failed",
                f"Unknown export format: {format_type}",
            )
        except Exception as exc:  # pragma: no cover - GUI error handling
            QMessageBox.critical(self, "Export Failed", f"Failed to export ephemeris:\n{exc}")
            raise

    def _export_orbit_summary_npz(self) -> None:
        """Export all Orbit Summary metrics (instantaneous + orbit-averaged) as compressed NPZ."""
        result = getattr(self, "_last_result", None)
        if result is None or not getattr(result, "timeline_seconds", np.empty(0)).size:
            QMessageBox.warning(
                self,
                "No Data",
                "No orbit summary data to export. Please run an analysis first.",
            )
            return

        # Resolve output path (default outputs/ with timestamp).
        timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
        outputs_dir = Path(__file__).resolve().parents[3] / "outputs"
        outputs_dir.mkdir(exist_ok=True)
        default_path = outputs_dir / f"orbit_summary_{timestamp}.npz"

        filename, _ = QFileDialog.getSaveFileName(
            self,
            "Export Orbit Summary NPZ",
            str(default_path),
            "NPZ files (*.npz);;All files (*.*)",
        )
        if not filename:
            return

        out_path = Path(filename).expanduser().resolve()
        if out_path.suffix.lower() != ".npz":
            out_path = out_path.with_suffix(".npz")

        times_s = np.asarray(getattr(result, "timeline_seconds", []), dtype=float)
        times_h = times_s / 3600.0

        inst_json = self._collect_instantaneous_orbit_summary(result=result, n=int(times_s.size))
        averaged_json, smoothing_meta = self._compute_orbit_averaged_orbit_summary(
            times_s=times_s, inst=inst_json
        )

        def _to_float_array(v: Any) -> np.ndarray:
            # v is typically a list[float|None] from _json_list
            if isinstance(v, np.ndarray):
                return v.astype(float, copy=False)
            if isinstance(v, list):
                out = np.empty((len(v),), dtype=float)
                for i, x in enumerate(v):
                    out[i] = float("nan") if x is None else float(x)
                return out
            return np.asarray(v, dtype=float)

        inst_arr: dict[str, np.ndarray] = {
            k: _to_float_array(v) for k, v in inst_json.items() if isinstance(v, list)
        }
        avg_arr: dict[str, np.ndarray] = {
            k: _to_float_array(v) for k, v in averaged_json.items() if isinstance(v, list)
        }

        meta: dict[str, Any] = {
            "schema_version": 1,
            "generated_at_utc": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
            "smoothing": smoothing_meta,
        }

        # Include scenario metadata if the current config is available.
        cfg = getattr(self, "_current_config", None)
        if cfg is not None:
            start = getattr(getattr(cfg, "scenario", None), "start_time", None)
            end = getattr(getattr(cfg, "scenario", None), "end_time", None)
            if start is not None and end is not None:
                meta["scenario"] = {
                    "start_time_utc": start.astimezone(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
                    "end_time_utc": end.astimezone(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
                    "duration_s": float((end - start).total_seconds()),
                }

        try:
            save_orbit_summary_npz(
                out_path,
                time_seconds=times_s,
                time_hours=times_h,
                instantaneous=inst_arr,
                orbit_averaged=avg_arr,
                meta=meta,
            )
            QMessageBox.information(
                self, "Export Successful", f"Orbit summary exported to:\n{out_path}"
            )
        except Exception as exc:  # pragma: no cover - GUI error handling
            QMessageBox.critical(self, "Export Failed", f"Failed to export orbit summary:\n{exc}")
            raise

    def _collect_instantaneous_orbit_summary(self, *, result: AnalysisResult, n: int) -> dict[str, Any]:
        """Collect all instantaneous Orbit Summary metrics aligned to timeline_seconds."""
        def _arr(name: str) -> np.ndarray:
            return np.asarray(getattr(result, name, []), dtype=float)

        def _aligned(name: str) -> list[float | None]:
            a = _arr(name)
            if a.size != n:
                a = np.full((n,), np.nan, dtype=float)
            return self._json_list(a)

        return {
            "orbital_altitude_km": _aligned("orbital_altitude_km"),
            "semi_major_axis_km": _aligned("semi_major_axis_km"),
            "perigee_altitude_km": _aligned("perigee_altitude_km"),
            "apogee_altitude_km": _aligned("apogee_altitude_km"),
            "eccentricity": _aligned("eccentricity"),
            "inclination_deg": _aligned("inclination_deg"),
            "argument_of_perigee_deg": _aligned("argument_of_perigee_deg"),
            "true_anomaly_deg": _aligned("true_anomaly_deg"),
            "orbital_period_s": _aligned("orbital_period_series_s"),
            "angle_of_attack_deg": _aligned("angle_of_attack_deg"),
        }

    def _compute_orbit_averaged_orbit_summary(
        self, *, times_s: np.ndarray, inst: dict[str, Any]
    ) -> tuple[dict[str, Any], dict[str, Any]]:
        """Wrapper around :func:`cosmic_toolbox.analyses.orbit_summary.compute_orbit_averaged_orbit_summary`."""

        return _compute_orbit_averaged(times_s=np.asarray(times_s, dtype=float), inst=inst)

    def _compute_series_stats(self, series: dict[str, Any]) -> dict[str, Any]:
        """Wrapper around :func:`cosmic_toolbox.analyses.orbit_summary.compute_series_stats`."""

        return _compute_stats(series)

    def _json_list(self, arr: np.ndarray) -> list[float | None]:
        a = np.asarray(arr, dtype=float)
        out: list[float | None] = []
        for v in a.tolist():
            fv = float(v)
            out.append(fv if np.isfinite(fv) else None)
        return out

    def _sanitize_for_json(self, obj: Any) -> Any:
        """Recursively convert numpy/scalars and non-finite floats into JSON-safe types."""
        if isinstance(obj, dict):
            return {str(k): self._sanitize_for_json(v) for k, v in obj.items()}
        if isinstance(obj, list):
            return [self._sanitize_for_json(v) for v in obj]
        if isinstance(obj, tuple):
            return [self._sanitize_for_json(v) for v in obj]
        if isinstance(obj, np.ndarray):
            return self._sanitize_for_json(obj.tolist())
        if isinstance(obj, (np.floating, float)):
            v = float(obj)
            return v if np.isfinite(v) else None
        if isinstance(obj, (np.integer, int)):
            return int(obj)
        return obj

    def _downsample_series(
        self, times: np.ndarray, *series: np.ndarray, max_points: int = 2000
    ) -> tuple[np.ndarray, List[np.ndarray]]:
        """Return downsampled time and series arrays to keep plotting responsive."""
        if times.size == 0:
            return times, [s for s in series]
        if times.size <= max_points:
            return times, [s for s in series]
        indices = np.linspace(0, times.size - 1, max_points, dtype=int)
        indices = np.unique(indices)
        return (
            times[indices],
            [s[indices] if s.size == times.size else s for s in series],
        )

    def _downsample_peak_series(
        self, times: np.ndarray, values: np.ndarray, max_points: int = 2000
    ) -> tuple[np.ndarray, np.ndarray]:
        """Downsample a spiky time series while preserving narrow peaks.

        This keeps the point budget similar to `_downsample_series`, but chooses
        representative min/max samples within bins so short-lived excursions
        (e.g., AoA during attitude switching) are less likely to disappear.
        """
        if times.size == 0 or values.size == 0:
            return times, values
        if times.size <= max_points or values.size != times.size:
            return times, values

        n = int(times.size)
        # Aim for ~2 points per bin (min+max), with some slack for endpoints.
        bins = max(1, max_points // 2)
        bin_size = int(np.ceil(n / float(bins)))

        idxs: list[int] = [0, n - 1]
        for start in range(0, n, bin_size):
            end = min(n, start + bin_size)
            sl = values[start:end]
            finite = np.isfinite(sl)
            if not finite.any():
                # No finite data in this bin; keep a representative point.
                idxs.append(start)
                continue
            # Compute min/max on finite subset.
            sl_f = np.where(finite, sl, np.nan)
            i_min = int(np.nanargmin(sl_f))
            i_max = int(np.nanargmax(sl_f))
            idxs.append(start + i_min)
            idxs.append(start + i_max)

        idx_arr = np.array(sorted(set(idxs)), dtype=int)
        # Enforce max_points strictly (rarely needed).
        if idx_arr.size > max_points:
            sel = np.linspace(0, idx_arr.size - 1, max_points, dtype=int)
            sel = np.unique(sel)
            idx_arr = idx_arr[sel]

        return times[idx_arr], values[idx_arr]

    def _update_orbit_summary(self, result: AnalysisResult | None) -> None:
        """Refresh the orbit summary plots from the latest analysis result."""
        if not getattr(self, "orbit_secular_plot", None) or not getattr(
            self, "orbit_instant_plot", None
        ):
            return

        eph = result.ephemeris if result is not None else None
        if result is None or eph is None or eph.ecef_pos_km is None or eph.ecef_pos_km.shape[0] == 0:
            # Clear plots and show placeholders.
            for plot, legend, title in [
                (
                    self.orbit_secular_plot,
                    getattr(self, "_orbit_secular_legend", None),
                    "Orbit-averaged orbit summary\n(Run a simulation to view)",
                ),
                (
                    self.orbit_instant_plot,
                    getattr(self, "_orbit_instant_legend", None),
                    "Instantaneous orbit summary\n(Run a simulation to view)",
                ),
            ]:
                plot.clear()
                plot.setTitle(title)
                plot.setLabel("bottom", "Time since start", units="h")
                plot.setLabel("left", "", units="")
                if legend is not None:
                    legend.clear()

            self._orbit_summary_data = None
            if getattr(self, "orbit_secular_combo", None) is not None:
                self.orbit_secular_combo.setEnabled(False)
            if getattr(self, "orbit_instant_combo", None) is not None:
                self.orbit_instant_combo.setEnabled(False)
            self._set_ephemeris_export_enabled(False)
            self._set_orbit_summary_export_enabled(False)
            return

        if getattr(self, "orbit_secular_combo", None) is not None:
            self.orbit_secular_combo.setEnabled(True)
        if getattr(self, "orbit_instant_combo", None) is not None:
            self.orbit_instant_combo.setEnabled(True)

        self._set_ephemeris_export_enabled(eph is not None and eph.timestamps_unix is not None)
        self._set_orbit_summary_export_enabled(bool(getattr(result, "timeline_seconds", np.empty(0)).size))

        # Use timeline seconds as the common time base.
        times_hours = np.asarray(result.timeline_seconds, dtype=float) / 3600.0

        alt = np.asarray(getattr(result, "orbital_altitude_km", []), dtype=float)
        sma = np.asarray(getattr(result, "semi_major_axis_km", []), dtype=float)
        perigee = np.asarray(getattr(result, "perigee_altitude_km", []), dtype=float)
        apogee = np.asarray(getattr(result, "apogee_altitude_km", []), dtype=float)
        ecc = np.asarray(getattr(result, "eccentricity", []), dtype=float)
        inc = np.asarray(getattr(result, "inclination_deg", []), dtype=float)
        argp = np.asarray(getattr(result, "argument_of_perigee_deg", []), dtype=float)
        # Normalize instantaneous argument of perigee and true anomaly to [0, 360) degrees for plotting.
        if argp.size:
            argp = np.mod(argp, 360.0)
        true_anom = np.asarray(getattr(result, "true_anomaly_deg", []), dtype=float)
        if true_anom.size:
            true_anom = np.mod(true_anom, 360.0)
        period = np.asarray(getattr(result, "orbital_period_series_s", []), dtype=float)
        aoa = np.asarray(getattr(result, "angle_of_attack_deg", []), dtype=float)

        # Ensure AoA is safe to plot even if missing from older results.
        if aoa.size != times_hours.size:
            aoa = np.full(times_hours.shape, np.nan, dtype=float)
        if inc.size != times_hours.size:
            inc = np.full(times_hours.shape, np.nan, dtype=float)

        # Fallback for older results without per-sample metrics.
        if alt.size == 0 or alt.size != times_hours.size:
            if eph is not None and eph.ecef_pos_km is not None and eph.ecef_pos_km.shape[0] == times_hours.size:
                radii = np.linalg.norm(eph.ecef_pos_km, axis=1)
                alt = radii - WGS84_EQUATORIAL_RADIUS_KM
            else:
                alt = np.full(times_hours.shape, np.nan, dtype=float)

        # Cache full-resolution data for per-metric rendering.
        self._orbit_summary_data = {
            "times_hours": times_hours,
            # Instantaneous
            "inst_altitude_km": alt,
            "inst_semi_major_axis_km": sma,
            "inst_perigee_altitude_km": perigee,
            "inst_apogee_altitude_km": apogee,
            "inst_eccentricity": ecc,
            "inst_inclination_deg": inc,
            "inst_argument_of_perigee_deg": argp,
            "inst_true_anomaly_deg": true_anom,
            "inst_period_seconds": period,
            "inst_angle_of_attack_deg": aoa,
        }

        # Estimate a representative orbital period (in seconds) from the
        # instantaneous Keplerian period series. This lets us build a
        # smoothing window that roughly spans one orbit, which is a simple
        # numerical approximation to Brouwer–Lyddane-style mean elements
        # (short-period averaged quantities).
        times_sec = times_hours * 3600.0
        valid_periods = period[np.isfinite(period) & (period > 0.0)]
        if valid_periods.size and times_sec.size > 1:
            one_orbit_s = float(np.median(valid_periods))
            dt_s = float(np.median(np.diff(times_sec)))
            if not np.isfinite(dt_s) or dt_s <= 0.0:
                dt_s = float(times_sec[-1] - times_sec[0]) / max(times_sec.size - 1, 1)
            window = int(round(one_orbit_s / dt_s)) if dt_s > 0.0 else 0
        else:
            one_orbit_s = 0.0
            window = 0

        # Clamp window to a sensible range.
        window = max(5, min(window, max(5, times_hours.size))) if window > 0 else 0

        # Compute orbit-period-based smoothed "Brouwer–Lyddane mean" versions
        # of each quantity. We explicitly drop samples that don't have a full
        # window of support to avoid edge artefacts near the start and end of
        # the simulation.
        if window > 1:
            kernel = np.ones(window, dtype=float) / float(window)

            def _smooth(values: np.ndarray) -> np.ndarray:
                if values.size == 0:
                    return values
                smoothed = np.convolve(values, kernel, mode="same")
                # Mask out the first/last half-window where the convolution
                # relies on partially defined windows, which otherwise creates
                # artificial behaviour close to the simulation boundaries.
                half = window // 2
                if values.size > window and half > 0:
                    smoothed[:half] = np.nan
                    smoothed[-half:] = np.nan
                return smoothed

            alt_sec = _smooth(alt)
            sma_sec = _smooth(sma)
            per_sec = _smooth(perigee)
            apo_sec = _smooth(apogee)
            ecc_sec = _smooth(ecc)
            inc_sec = _smooth(inc)
            # For angular quantities, use a circular mean based on the complex
            # exponential to avoid artificial jumps near 0/360 deg. We smooth
            # cos(ω) and sin(ω) separately and then recover the mean angle.
            if argp.size:
                argp_rad = np.deg2rad(argp)
                cos_w = np.cos(argp_rad)
                sin_w = np.sin(argp_rad)
                cos_w_sec = _smooth(cos_w)
                sin_w_sec = _smooth(sin_w)
                argp_sec = np.rad2deg(np.arctan2(sin_w_sec, cos_w_sec))
                argp_sec = np.mod(argp_sec, 360.0)
            else:
                argp_sec = np.array([], dtype=float)
            period_sec = _smooth(period)
            aoa_sec = _smooth(aoa)
        else:
            alt_sec = alt
            sma_sec = sma
            per_sec = perigee
            apo_sec = apogee
            ecc_sec = ecc
            inc_sec = inc
            argp_sec = argp
            period_sec = period
            aoa_sec = aoa

        # For Brouwer–Lyddane-mean curves, omit the first and last orbit instead of relying on
        # partially defined smoothing windows at the boundaries.
        if period.size and times_hours.size and one_orbit_s > 0.0:
            t_start = float(times_sec[0])
            t_end = float(times_sec[-1])
            inner_mask = (times_sec - t_start >= one_orbit_s) & (
                t_end - times_sec >= one_orbit_s
            )
            # Apply mask: mark edge samples as NaN so they are visually omitted.
            for arr in (
                alt_sec,
                sma_sec,
                per_sec,
                apo_sec,
                ecc_sec,
                inc_sec,
                argp_sec,
                period_sec,
                aoa_sec,
            ):
                if arr.size == inner_mask.size:
                    arr[~inner_mask] = np.nan

        self._orbit_summary_data.update(
            {
                "sec_altitude_km": alt_sec,
                "sec_semi_major_axis_km": sma_sec,
                "sec_perigee_altitude_km": per_sec,
                "sec_apogee_altitude_km": apo_sec,
                "sec_eccentricity": ecc_sec,
                "sec_inclination_deg": inc_sec,
                "sec_argument_of_perigee_deg": argp_sec,
                "sec_period_seconds": period_sec,
                "sec_angle_of_attack_deg": aoa_sec,
            }
        )

        # Render the currently selected metrics for both plots.
        sec_index = (
            self.orbit_secular_combo.currentIndex()
            if getattr(self, "orbit_secular_combo", None) is not None
            else 0
        )
        inst_index = (
            self.orbit_instant_combo.currentIndex()
            if getattr(self, "orbit_instant_combo", None) is not None
            else 0
        )
        self._render_secular_metric(sec_index)
        self._render_instant_metric(inst_index)

    def _handle_secular_metric_changed(self, index: int) -> None:
        """React to dropdown changes for secular elements."""
        self._render_secular_metric(index)

    def _handle_instant_metric_changed(self, index: int) -> None:
        """React to dropdown changes for instantaneous elements."""
        self._render_instant_metric(index)

    def _render_secular_metric(self, index: int) -> None:
        """Render the short-period averaged (Brouwer–Lyddane-style) metric selected in the combo box."""
        if not self._orbit_summary_data or not getattr(self, "orbit_secular_plot", None):
            return

        times_hours = self._orbit_summary_data["times_hours"]
        alt = self._orbit_summary_data.get("sec_altitude_km", np.array([], dtype=float))
        sma = self._orbit_summary_data.get(
            "sec_semi_major_axis_km", np.array([], dtype=float)
        )
        perigee = self._orbit_summary_data.get(
            "sec_perigee_altitude_km", np.array([], dtype=float)
        )
        apogee = self._orbit_summary_data.get(
            "sec_apogee_altitude_km", np.array([], dtype=float)
        )
        ecc = self._orbit_summary_data.get("sec_eccentricity", np.array([], dtype=float))
        inc = self._orbit_summary_data.get("sec_inclination_deg", np.array([], dtype=float))
        argp = self._orbit_summary_data.get(
            "sec_argument_of_perigee_deg", np.array([], dtype=float)
        )
        period = self._orbit_summary_data.get("sec_period_seconds", np.array([], dtype=float))
        aoa = self._orbit_summary_data.get(
            "sec_angle_of_attack_deg", np.array([], dtype=float)
        )

        plot = self.orbit_secular_plot
        legend = getattr(self, "_orbit_secular_legend", None)
        plot.clear()
        if legend is not None:
            legend.clear()

        # Geodetic altitude only (orbit-averaged).
        if index == 0:
            plot.setTitle("Geodetic Altitude (orbit-averaged)")
            plot.setLabel("bottom", "Time since start", units="h")
            plot.setLabel("left", "Geodetic altitude", units="km")
            times_ds, (alt_ds,) = self._downsample_series(times_hours, alt)
            plot.plot(
                times_ds,
                alt_ds,
                pen=pg.mkPen("#76c7ff", width=2),
                name="Geodetic altitude (orbit-averaged)",
            )
            plot.getViewBox().autoRange()
            return

        # Semi-major axis and apsides from orbital elements (Brouwer–Lyddane mean).
        if index == 1:
            plot.setTitle("Semi-major Axis and Apsides (Brouwer–Lyddane mean)")
            plot.setLabel("bottom", "Time since start", units="h")
            plot.setLabel("left", "Altitude (from elements)", units="km")
            # Convert element-derived radii to an altitude-like quantity so this
            # plot aligns with the Mean-IC altitude target and the geodetic altitude plots.
            sma_plot = sma - WGS84_EQUATORIAL_RADIUS_KM if sma.size else sma
            per_plot = perigee - WGS84_EQUATORIAL_RADIUS_KM if perigee.size else perigee
            apo_plot = apogee - WGS84_EQUATORIAL_RADIUS_KM if apogee.size else apogee
            times_ds, (sma_ds, per_ds, apo_ds) = self._downsample_series(
                times_hours, sma_plot, per_plot, apo_plot
            )
            if sma_ds.size == times_ds.size:
                plot.plot(
                    times_ds,
                    sma_ds,
                    pen=pg.mkPen("#2196f3", width=2),
                    name="Semi-major axis altitude",
                )
            if per_ds.size == times_ds.size:
                plot.plot(
                    times_ds,
                    per_ds,
                    pen=pg.mkPen(
                        "#ff9800",
                        width=1.5,
                        style=pg.QtCore.Qt.PenStyle.DashLine,  # type: ignore[attr-defined]
                    ),  # type: ignore[arg-type]
                    name="Perigee altitude (from elements)",
                )
            if apo_ds.size == times_ds.size:
                plot.plot(
                    times_ds,
                    apo_ds,
                    pen=pg.mkPen(
                        "#4caf50",
                        width=1.5,
                        style=pg.QtCore.Qt.PenStyle.DashLine,  # type: ignore[attr-defined]
                    ),  # type: ignore[arg-type]
                    name="Apogee altitude (from elements)",
                )
            plot.getViewBox().autoRange()
            return

        # Eccentricity (Brouwer–Lyddane mean).
        if index == 2:
            plot.setTitle("Eccentricity (Brouwer–Lyddane mean)")
            plot.setLabel("bottom", "Time since start", units="h")
            plot.setLabel("left", "e", units="")
            times_ds, (ecc_ds,) = self._downsample_series(times_hours, ecc)
            plot.plot(
                times_ds,
                ecc_ds,
                pen=pg.mkPen("#f06292", width=2),
                name="Eccentricity (Brouwer–Lyddane mean)",
            )
            plot.getViewBox().autoRange()
            return

        # Inclination (Brouwer–Lyddane mean).
        if index == 3:
            plot.setTitle("Inclination (Brouwer–Lyddane mean)")
            plot.setLabel("bottom", "Time since start", units="h")
            plot.setLabel("left", "i", units="deg")
            times_ds, (inc_ds,) = self._downsample_series(times_hours, inc)
            plot.plot(
                times_ds,
                inc_ds,
                pen=pg.mkPen("#7cb342", width=2),
                name="Inclination (Brouwer–Lyddane mean)",
            )
            plot.getViewBox().autoRange()
            return

        # Argument of perigee (Brouwer–Lyddane mean).
        if index == 4:
            plot.setTitle("Argument of Perigee (Brouwer–Lyddane mean)")
            plot.setLabel("bottom", "Time since start", units="h")
            plot.setLabel("left", "ω", units="deg")
            times_ds, (argp_ds,) = self._downsample_series(times_hours, argp)
            plot.plot(
                times_ds,
                argp_ds,
                pen=pg.mkPen("#ba68c8", width=2),
                name="Argument of perigee (Brouwer–Lyddane mean)",
            )
            plot.getViewBox().autoRange()
            return

        # Orbital period (Brouwer–Lyddane mean).
        if index == 5:
            plot.setTitle("Orbital Period (Brouwer–Lyddane mean)")
            plot.setLabel("bottom", "Time since start", units="h")
            plot.setLabel("left", "Period", units="s")
            times_ds, (period_ds,) = self._downsample_series(times_hours, period)
            plot.plot(
                times_ds,
                period_ds,
                pen=pg.mkPen("#ffb74d", width=2),
                name="Orbital period (Brouwer–Lyddane mean)",
            )
            plot.getViewBox().autoRange()
            return

        # Angle of attack (orbit-averaged).
        if index == 6:
            plot.setTitle("Angle of Attack (orbit-averaged)")
            plot.setLabel("bottom", "Time since start", units="h")
            plot.setLabel("left", "AoA", units="deg")
            times_ds, (aoa_ds,) = self._downsample_series(times_hours, aoa)
            plot.plot(
                times_ds,
                aoa_ds,
                pen=pg.mkPen("#8e24aa", width=2),
                name="Angle of attack (orbit-averaged)",
            )
            plot.getViewBox().autoRange()
            return

        # Fallback: show angle of attack.
        plot.setTitle("Angle of Attack (orbit-averaged)")
        plot.setLabel("bottom", "Time since start", units="h")
        plot.setLabel("left", "AoA", units="deg")
        times_ds, (aoa_ds,) = self._downsample_series(times_hours, aoa)
        plot.plot(
            times_ds,
            aoa_ds,
            pen=pg.mkPen("#8e24aa", width=2),
            name="Angle of attack (orbit-averaged)",
        )
        plot.getViewBox().autoRange()

    def _render_instant_metric(self, index: int) -> None:
        """Render the instantaneous (osculating) metric selected in the combo box."""
        if not self._orbit_summary_data or not getattr(self, "orbit_instant_plot", None):
            return

        times_hours = self._orbit_summary_data["times_hours"]
        alt = self._orbit_summary_data.get("inst_altitude_km", np.array([], dtype=float))
        sma = self._orbit_summary_data.get(
            "inst_semi_major_axis_km", np.array([], dtype=float)
        )
        perigee = self._orbit_summary_data.get(
            "inst_perigee_altitude_km", np.array([], dtype=float)
        )
        apogee = self._orbit_summary_data.get(
            "inst_apogee_altitude_km", np.array([], dtype=float)
        )
        ecc = self._orbit_summary_data.get("inst_eccentricity", np.array([], dtype=float))
        inc = self._orbit_summary_data.get("inst_inclination_deg", np.array([], dtype=float))
        argp = self._orbit_summary_data.get(
            "inst_argument_of_perigee_deg", np.array([], dtype=float)
        )
        true_anom = self._orbit_summary_data.get(
            "inst_true_anomaly_deg", np.array([], dtype=float)
        )
        period = self._orbit_summary_data.get("inst_period_seconds", np.array([], dtype=float))
        aoa = self._orbit_summary_data.get(
            "inst_angle_of_attack_deg", np.array([], dtype=float)
        )

        plot = self.orbit_instant_plot
        legend = getattr(self, "_orbit_instant_legend", None)
        plot.clear()
        if legend is not None:
            legend.clear()

        # Geodetic altitude only (instantaneous).
        if index == 0:
            plot.setTitle("Geodetic Altitude (instantaneous)")
            plot.setLabel("bottom", "Time since start", units="h")
            plot.setLabel("left", "Geodetic altitude", units="km")
            times_ds, (alt_ds,) = self._downsample_series(
                times_hours, alt, max_points=2000
            )
            plot.plot(
                times_ds,
                alt_ds,
                pen=pg.mkPen("#76c7ff", width=2),
                name="Geodetic altitude",
            )
            plot.getViewBox().autoRange()
            return

        # Semi-major axis and apsides from orbital elements (instantaneous).
        if index == 1:
            plot.setTitle("Semi-major Axis and Apsides (instantaneous)")
            plot.setLabel("bottom", "Time since start", units="h")
            plot.setLabel("left", "Altitude (from elements)", units="km")
            sma_plot = sma - WGS84_EQUATORIAL_RADIUS_KM if sma.size else sma
            per_plot = perigee - WGS84_EQUATORIAL_RADIUS_KM if perigee.size else perigee
            apo_plot = apogee - WGS84_EQUATORIAL_RADIUS_KM if apogee.size else apogee
            times_ds, (sma_ds, per_ds, apo_ds) = self._downsample_series(
                times_hours, sma_plot, per_plot, apo_plot, max_points=2000
            )
            if sma_ds.size == times_ds.size:
                plot.plot(
                    times_ds,
                    sma_ds,
                    pen=pg.mkPen("#2196f3", width=2),
                    name="Semi-major axis altitude",
                )
            if per_ds.size == times_ds.size:
                plot.plot(
                    times_ds,
                    per_ds,
                    pen=pg.mkPen(
                        "#ff9800",
                        width=1.5,
                        style=pg.QtCore.Qt.PenStyle.DashLine,  # type: ignore[attr-defined]
                    ),
                    name="Perigee altitude (from elements)",
                )
            if apo_ds.size == times_ds.size:
                plot.plot(
                    times_ds,
                    apo_ds,
                    pen=pg.mkPen(
                        "#4caf50",
                        width=1.5,
                        style=pg.QtCore.Qt.PenStyle.DashLine,  # type: ignore[attr-defined]
                    ),
                    name="Apogee altitude (from elements)",
                )
            plot.getViewBox().autoRange()
            return

        # Eccentricity (instantaneous).
        if index == 2:
            plot.setTitle("Eccentricity (instantaneous)")
            plot.setLabel("bottom", "Time since start", units="h")
            plot.setLabel("left", "e", units="")
            times_ds, (ecc_ds,) = self._downsample_series(times_hours, ecc, max_points=2000)
            plot.plot(
                times_ds,
                ecc_ds,
                pen=pg.mkPen("#f06292", width=2),
                name="Eccentricity",
            )
            plot.getViewBox().autoRange()
            return

        # Inclination (instantaneous).
        if index == 3:
            plot.setTitle("Inclination (instantaneous)")
            plot.setLabel("bottom", "Time since start", units="h")
            plot.setLabel("left", "i", units="deg")
            times_ds, (inc_ds,) = self._downsample_series(times_hours, inc, max_points=2000)
            plot.plot(
                times_ds,
                inc_ds,
                pen=pg.mkPen("#7cb342", width=2),
                name="Inclination",
            )
            plot.getViewBox().autoRange()
            return

        # Argument of perigee (instantaneous).
        if index == 4:
            plot.setTitle("Argument of Perigee (instantaneous)")
            plot.setLabel("bottom", "Time since start", units="h")
            plot.setLabel("left", "ω", units="deg")
            times_ds, (argp_ds,) = self._downsample_series(times_hours, argp, max_points=2000)
            plot.plot(
                times_ds,
                argp_ds,
                pen=pg.mkPen("#ba68c8", width=2),
                name="Argument of perigee",
            )
            plot.getViewBox().autoRange()
            return

        # True anomaly (instantaneous).
        if index == 5:
            plot.setTitle("True Anomaly (instantaneous)")
            plot.setLabel("bottom", "Time since start", units="h")
            plot.setLabel("left", "ν", units="deg")
            times_ds, (ta_ds,) = self._downsample_series(
                times_hours, true_anom, max_points=2000
            )
            plot.plot(
                times_ds,
                ta_ds,
                pen=pg.mkPen("#26a69a", width=2),
                name="True anomaly",
            )
            plot.getViewBox().autoRange()
            return

        # Orbital period (instantaneous).
        if index == 6:
            plot.setTitle("Orbital Period (instantaneous)")
            plot.setLabel("bottom", "Time since start", units="h")
            plot.setLabel("left", "Period", units="s")
            times_ds, (period_ds,) = self._downsample_series(
                times_hours, period, max_points=2000
            )
            plot.plot(
                times_ds,
                period_ds,
                pen=pg.mkPen("#ffb74d", width=2),
                name="Orbital period",
            )
            plot.getViewBox().autoRange()
            return

        # Angle of attack (instantaneous).
        if index == 7:
            plot.setTitle("Angle of Attack (instantaneous)")
            plot.setLabel("bottom", "Time since start", units="h")
            plot.setLabel("left", "AoA", units="deg")
            times_ds, aoa_ds = self._downsample_peak_series(
                times_hours, aoa, max_points=2000
            )
            plot.plot(
                times_ds,
                aoa_ds,
                pen=pg.mkPen("#8e24aa", width=2),
                name="Angle of attack",
            )
            plot.getViewBox().autoRange()
            return

        # Fallback: show angle of attack.
        plot.setTitle("Angle of Attack (instantaneous)")
        plot.setLabel("bottom", "Time since start", units="h")
        plot.setLabel("left", "AoA", units="deg")
        times_ds, aoa_ds = self._downsample_peak_series(
            times_hours, aoa, max_points=2000
        )
        plot.plot(
            times_ds,
            aoa_ds,
            pen=pg.mkPen("#8e24aa", width=2),
            name="Angle of attack",
        )
        plot.getViewBox().autoRange()
        return

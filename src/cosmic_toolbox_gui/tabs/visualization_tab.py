"""Visualization tab mixin."""

from __future__ import annotations

import math
from datetime import datetime, timedelta, timezone
from pathlib import Path
import re
from typing import NamedTuple

import numpy as np
import pyqtgraph as pg
from PySide6.QtCore import Qt
from PySide6.QtGui import QImage
from PySide6.QtWidgets import (
    QApplication,
    QAbstractScrollArea,
    QCheckBox,
    QComboBox,
    QFileDialog,
    QHBoxLayout,
    QLabel,
    QListWidget,
    QListWidgetItem,
    QMessageBox,
    QPushButton,
    QSizePolicy,
    QSplitter,
    QSlider,
    QTabWidget,
    QVBoxLayout,
    QWidget,
)

from cosmic_toolbox.models import GroundStationConfig, PassStatistic
from cosmic_toolbox import pfd_math
from cosmic_toolbox.services import antenna_pattern
from cosmic_toolbox_gui.globe_math import gmst_rad_utc, rotate_vector_z
from cosmic_toolbox.analyses.visualization import ecef_to_globe_coords  # re-exported
from cosmic_toolbox_gui.opengl import GlobeWidget

STATION_VISUAL_OFFSET_KM = 25.0
SPEED_OF_LIGHT_MPS = 299_792_458.0


class _TrackPoint(NamedTuple):
    """Lightweight per-sample container for visualization animation."""

    x_km: float
    y_km: float
    z_km: float
    timestamp: datetime


class VisualizationTabMixin:
    """Builds and controls the visualization tab and animations."""

    def _get_comms_pointing_mode(self) -> str:
        cfg = getattr(self, "_current_config", None)
        propagation = getattr(cfg, "propagation", None) if cfg is not None else None
        mode = str(getattr(propagation, "comms_pointing_mode", "prograde_pointing") or "")
        mode = mode.strip().lower()
        if mode not in {"prograde_pointing", "free_to_roll", "constrained_aoa"}:
            raise ValueError(f"Unsupported comms pointing mode: {mode!r}")
        return mode

    def _get_comms_pointing_aoa_limit_deg(self) -> float:
        cfg = getattr(self, "_current_config", None)
        propagation = getattr(cfg, "propagation", None) if cfg is not None else None
        aoa_limit = float(getattr(propagation, "comms_pointing_aoa_limit_deg", 0.0))
        if not np.isfinite(aoa_limit) or aoa_limit < 0.0 or aoa_limit > 180.0:
            raise ValueError(f"Invalid comms AoA limit: {aoa_limit!r}")
        return aoa_limit

    def _is_comms_pointing_active(self, timestamp: datetime) -> bool:
        """Return True if comms steering should be active at timestamp."""
        if self._get_comms_pointing_mode() == "prograde_pointing":
            return False
        cfg = getattr(self, "_current_config", None)
        propagation = getattr(cfg, "propagation", None) if cfg is not None else None
        station_name = None
        if getattr(self, "_visual_selected_pass", None) is not None:
            station_name = getattr(self._visual_selected_pass, "station_name", None)
        if not station_name:
            return False
        result = getattr(self, "_last_result", None)
        if result is None:
            return False
        series = getattr(result, "station_elevation_series", {}).get(station_name)
        if series is None:
            return False
        idx = self._nearest_timeline_index(timestamp)
        if idx is None or idx < 0 or idx >= len(series):
            return False
        contact_el = float(getattr(propagation, "contact_elevation_deg", 10.0))
        el = float(series[idx])
        return bool(np.isfinite(el) and el >= contact_el)

    @staticmethod
    def _rotation_matrix_z(angle_rad: float) -> np.ndarray:
        c = math.cos(angle_rad)
        s = math.sin(angle_rad)
        return np.array(
            [
                [c, -s, 0.0],
                [s, c, 0.0],
                [0.0, 0.0, 1.0],
            ],
            dtype=float,
        )

    @staticmethod
    def _quat_from_rot_matrix(R: np.ndarray) -> np.ndarray:
        """Convert 3x3 rotation matrix -> quaternion (w,x,y,z)."""
        m = np.asarray(R, dtype=float).reshape(3, 3)
        tr = float(np.trace(m))
        if tr > 0.0:
            S = math.sqrt(tr + 1.0) * 2.0
            w = 0.25 * S
            x = (m[2, 1] - m[1, 2]) / S
            y = (m[0, 2] - m[2, 0]) / S
            z = (m[1, 0] - m[0, 1]) / S
        elif m[0, 0] > m[1, 1] and m[0, 0] > m[2, 2]:
            S = math.sqrt(1.0 + m[0, 0] - m[1, 1] - m[2, 2]) * 2.0
            w = (m[2, 1] - m[1, 2]) / S
            x = 0.25 * S
            y = (m[0, 1] + m[1, 0]) / S
            z = (m[0, 2] + m[2, 0]) / S
        elif m[1, 1] > m[2, 2]:
            S = math.sqrt(1.0 + m[1, 1] - m[0, 0] - m[2, 2]) * 2.0
            w = (m[0, 2] - m[2, 0]) / S
            x = (m[0, 1] + m[1, 0]) / S
            y = 0.25 * S
            z = (m[1, 2] + m[2, 1]) / S
        else:
            S = math.sqrt(1.0 + m[2, 2] - m[0, 0] - m[1, 1]) * 2.0
            w = (m[1, 0] - m[0, 1]) / S
            x = (m[0, 2] + m[2, 0]) / S
            y = (m[1, 2] + m[2, 1]) / S
            z = 0.25 * S
        q = np.array([w, x, y, z], dtype=float)
        n = float(np.linalg.norm(q))
        if not np.isfinite(n) or n < 1e-12:
            return np.array([1.0, 0.0, 0.0, 0.0], dtype=float)
        return q / n

    @staticmethod
    def _rot_matrix_from_quat(q: np.ndarray) -> np.ndarray:
        """Quaternion (w,x,y,z) -> 3x3 rotation matrix."""
        w, x, y, z = [float(v) for v in np.asarray(q, dtype=float).reshape(4)]
        n = w * w + x * x + y * y + z * z
        if not np.isfinite(n) or n < 1e-12:
            return np.identity(3, dtype=float)
        s = 2.0 / n
        wx, wy, wz = s * w * x, s * w * y, s * w * z
        xx, xy, xz = s * x * x, s * x * y, s * x * z
        yy, yz, zz = s * y * y, s * y * z, s * z * z
        return np.array(
            [
                [1.0 - (yy + zz), xy - wz, xz + wy],
                [xy + wz, 1.0 - (xx + zz), yz - wx],
                [xz - wy, yz + wx, 1.0 - (xx + yy)],
            ],
            dtype=float,
        )

    @staticmethod
    def _slerp(q0: np.ndarray, q1: np.ndarray, t: float) -> np.ndarray:
        """Spherical linear interpolation between quaternions."""
        q0 = np.asarray(q0, dtype=float).reshape(4)
        q1 = np.asarray(q1, dtype=float).reshape(4)
        dot = float(np.dot(q0, q1))
        # Ensure shortest path.
        if dot < 0.0:
            q1 = -q1
            dot = -dot
        dot = float(np.clip(dot, -1.0, 1.0))
        if dot > 0.9995:
            # Nearly identical - use lerp.
            q = q0 + float(t) * (q1 - q0)
            n = float(np.linalg.norm(q))
            return q / max(n, 1e-12)
        theta_0 = math.acos(dot)
        sin_theta_0 = math.sin(theta_0)
        theta = theta_0 * float(t)
        sin_theta = math.sin(theta)
        s0 = math.cos(theta) - dot * sin_theta / max(sin_theta_0, 1e-12)
        s1 = sin_theta / max(sin_theta_0, 1e-12)
        q = (s0 * q0) + (s1 * q1)
        n = float(np.linalg.norm(q))
        return q / max(n, 1e-12)

    def _nearest_timeline_index(self, timestamp: datetime) -> int | None:
        """Return the nearest analysis sample index for a timestamp."""
        result = getattr(self, "_last_result", None)
        if result is None:
            return None
        timeline = getattr(result, "timeline_seconds", None)
        if timeline is None:
            return None
        t_s = float(self._seconds_since_reference(timestamp))
        timeline_arr = np.asarray(timeline, dtype=float)
        if timeline_arr.size == 0:
            return None
        idx = int(np.searchsorted(timeline_arr, t_s))
        if idx <= 0:
            return 0
        if idx >= timeline_arr.size:
            return int(timeline_arr.size - 1)
        left = idx - 1
        right = idx
        return left if abs(timeline_arr[left] - t_s) <= abs(timeline_arr[right] - t_s) else right

    def _lookup_body_axes_from_attitude(self, timestamp: datetime) -> np.ndarray | None:
        """Return body axes (3x3, columns X/Y/Z) in the current render frame."""
        result = getattr(self, "_last_result", None)
        if result is None:
            return None
        eph = getattr(result, "ephemeris", None)
        timeline = getattr(result, "timeline_seconds", None)
        if eph is None or eph.body_x_ecef is None or timeline is None:
            return None

        timeline_arr = np.asarray(timeline, dtype=float)
        if timeline_arr.size == 0:
            return None
        if eph.body_x_ecef.shape[0] != int(timeline_arr.size):
            return None
        t_s = float(self._seconds_since_reference(timestamp))

        idx = int(np.searchsorted(timeline_arr, t_s))
        if idx <= 0:
            i0 = i1 = 0
            frac = 0.0
        elif idx >= timeline_arr.size:
            i0 = i1 = int(timeline_arr.size - 1)
            frac = 0.0
        else:
            i0 = idx - 1
            i1 = idx
            dt = float(timeline_arr[i1] - timeline_arr[i0])
            frac = 0.0 if dt <= 1e-9 else float((t_s - timeline_arr[i0]) / dt)
            frac = float(np.clip(frac, 0.0, 1.0))

        def _axes_ecef(i: int) -> np.ndarray | None:
            bx = eph.body_x_ecef[i].copy()
            by = eph.body_y_ecef[i].copy()
            bz = eph.body_z_ecef[i].copy()
            if not (
                np.all(np.isfinite(bx))
                and np.all(np.isfinite(by))
                and np.all(np.isfinite(bz))
            ):
                return None
            return np.stack([bx, by, bz], axis=1)

        R0 = _axes_ecef(i0)
        if R0 is None:
            return None
        if i1 == i0:
            R_ecef = R0
        else:
            R1 = _axes_ecef(i1)
            if R1 is None:
                R_ecef = R0
            else:
                q0 = self._quat_from_rot_matrix(R0)
                q1 = self._quat_from_rot_matrix(R1)
                q = self._slerp(q0, q1, frac)
                R_ecef = self._rot_matrix_from_quat(q)

        # Fixed ECEF -> globe (+90° about +Z) mapping: (x,y,z) -> (-y,x,z)
        G = np.array([[0.0, -1.0, 0.0], [1.0, 0.0, 0.0], [0.0, 0.0, 1.0]], dtype=float)
        R_globe = G @ R_ecef

        if self._visual_frame_mode == "ECI":
            angle = gmst_rad_utc(timestamp)
            Z = self._rotation_matrix_z(angle)
            R_cur = Z @ R_globe
        else:
            R_cur = R_globe

        return np.asarray(R_cur, dtype=np.float32)

    def _get_station_position_unoffset(
        self, timestamp: datetime
    ) -> tuple[float, float, float] | None:
        """Station position in the current frame without visual offset."""
        if self._visual_station_ecef is None:
            return None
        globe_vec = np.array(ecef_to_globe_coords(*self._visual_station_ecef), dtype=float)
        return self._convert_vector_to_current_frame(tuple(globe_vec.tolist()), timestamp)

    def _is_contact_attitude_active(self, timestamp: datetime) -> bool:
        """Return True if contact-based attitude switching should be active at timestamp."""
        cfg = getattr(self, "_current_config", None)
        propagation = getattr(cfg, "propagation", None) if cfg is not None else None
        attitude_mode = str(getattr(propagation, "attitude_mode", "prograde") or "prograde").lower()
        enable_switching = bool(getattr(propagation, "enable_contact_attitude_switching", True))
        if not enable_switching or attitude_mode == "nadir":
            return False
        station_name = None
        if getattr(self, "_visual_selected_pass", None) is not None:
            station_name = getattr(self._visual_selected_pass, "station_name", None)
        if not station_name:
            return False
        result = getattr(self, "_last_result", None)
        if result is None:
            return False
        series = getattr(result, "station_elevation_series", {}).get(station_name)
        if not series:
            return False
        idx = self._nearest_timeline_index(timestamp)
        if idx is None or idx < 0 or idx >= len(series):
            return False
        contact_el = float(getattr(propagation, "contact_elevation_deg", 10.0))
        el = float(series[idx])
        return bool(np.isfinite(el) and el >= contact_el)

    @staticmethod
    def _orthonormal_axes_from_xz(x: np.ndarray, z: np.ndarray) -> np.ndarray | None:
        """Build orthonormal [X,Y,Z] columns given approximate x and z."""
        x = np.asarray(x, dtype=float).reshape(3)
        z = np.asarray(z, dtype=float).reshape(3)
        zx = float(np.linalg.norm(z))
        xx = float(np.linalg.norm(x))
        if zx < 1e-6 or xx < 1e-6:
            return None
        z = z / zx
        # Make x orthogonal to z
        x = x - float(np.dot(x, z)) * z
        xx = float(np.linalg.norm(x))
        if xx < 1e-6:
            return None
        x = x / xx
        y = np.cross(z, x)
        yy = float(np.linalg.norm(y))
        if yy < 1e-6:
            return None
        y = y / yy
        # Recompute z to ensure perfect orthogonality
        z = np.cross(x, y)
        zz = float(np.linalg.norm(z))
        if zz < 1e-6:
            return None
        z = z / zz
        return np.stack([x, y, z], axis=1).astype(np.float32)

    @classmethod
    def _constrained_aoa_axes_from_x_los(
        cls,
        x_nominal: np.ndarray,
        los: np.ndarray,
        *,
        max_aoa_deg: float,
    ) -> np.ndarray | None:
        """Build body axes with steering limited by |angle(body +X, prograde)|."""
        if not np.isfinite(max_aoa_deg) or max_aoa_deg < 0.0 or max_aoa_deg > 180.0:
            raise ValueError(f"Invalid max_aoa_deg: {max_aoa_deg!r}")
        x0 = np.asarray(x_nominal, dtype=float).reshape(3)
        s = np.asarray(los, dtype=float).reshape(3)
        x0_norm = float(np.linalg.norm(x0))
        s_norm = float(np.linalg.norm(s))
        if x0_norm < 1e-6 or s_norm < 1e-6:
            return None
        x0 = x0 / x0_norm
        s = s / s_norm
        dot_x = float(np.clip(np.dot(s, x0), -1.0, 1.0))
        perp = s - dot_x * x0
        perp_norm = float(np.linalg.norm(perp))
        if perp_norm > 1e-12:
            perp_dir = perp / perp_norm
        else:
            perp_dir = np.array([0.0, 0.0, 1.0], dtype=float)
            perp_dir = perp_dir - float(np.dot(perp_dir, x0)) * x0
            perp_dir_norm = float(np.linalg.norm(perp_dir))
            if perp_dir_norm < 1e-12:
                perp_dir = np.array([0.0, 1.0, 0.0], dtype=float)
                perp_dir = perp_dir - float(np.dot(perp_dir, x0)) * x0
                perp_dir_norm = float(np.linalg.norm(perp_dir))
            if perp_dir_norm < 1e-12:
                return None
            perp_dir = perp_dir / perp_dir_norm
        aoa_cmd_rad = min(
            math.radians(float(max_aoa_deg)),
            math.atan2(abs(dot_x), perp_norm),
        )
        x_sign = 1.0 if dot_x >= 0.0 else -1.0
        commanded_x = (
            math.cos(aoa_cmd_rad) * x0 - x_sign * math.sin(aoa_cmd_rad) * perp_dir
        )
        z_cmd = s - float(np.dot(s, commanded_x)) * commanded_x
        z_norm = float(np.linalg.norm(z_cmd))
        if z_norm < 1e-12:
            return None
        return cls._orthonormal_axes_from_xz(commanded_x, z_cmd)

    @staticmethod
    def _roll_axes_about_body_x(base_axes: np.ndarray, roll_rad: float) -> np.ndarray:
        """Rotate body Y/Z axes about the existing body +X axis."""
        axes = np.asarray(base_axes, dtype=float)
        if axes.shape != (3, 3):
            raise ValueError("Base axes must have shape (3, 3) for free-to-roll visualization.")
        if not np.all(np.isfinite(axes)):
            raise ValueError("Base axes contain non-finite values for free-to-roll visualization.")
        angle = float(roll_rad)
        if not np.isfinite(angle):
            raise ValueError("Roll angle must be finite for free-to-roll visualization.")

        x_axis = axes[:, 0]
        y_axis = axes[:, 1]
        z_axis = axes[:, 2]
        cos_roll = math.cos(angle)
        sin_roll = math.sin(angle)
        y_rot = y_axis * cos_roll - z_axis * sin_roll
        z_rot = y_axis * sin_roll + z_axis * cos_roll

        x_norm = float(np.linalg.norm(x_axis))
        y_norm = float(np.linalg.norm(y_rot))
        z_norm = float(np.linalg.norm(z_rot))
        if x_norm < 1e-6 or y_norm < 1e-6 or z_norm < 1e-6:
            raise ValueError("Rolled body axes became degenerate in free-to-roll visualization.")

        x_axis = x_axis / x_norm
        y_rot = y_rot / y_norm
        z_rot = z_rot / z_norm
        return np.stack([x_axis, y_rot, z_rot], axis=1).astype(np.float32)

    @classmethod
    def _free_to_roll_axes_from_base_and_los(
        cls,
        base_axes: np.ndarray,
        los: np.ndarray,
    ) -> np.ndarray:
        """Apply a pure roll about body +X so LOS lies in the X/Z plane."""
        axes = np.asarray(base_axes, dtype=float)
        if axes.shape != (3, 3):
            raise ValueError("Base axes must have shape (3, 3) for free-to-roll visualization.")
        los_vec = np.asarray(los, dtype=float).reshape(3)
        los_norm = float(np.linalg.norm(los_vec))
        if los_norm < 1e-6:
            raise ValueError("LOS vector is degenerate for free-to-roll visualization.")
        los_hat = los_vec / los_norm
        los_body = np.array(
            [
                [
                    float(np.dot(los_hat, axes[:, 0])),
                    float(np.dot(los_hat, axes[:, 1])),
                    float(np.dot(los_hat, axes[:, 2])),
                ]
            ],
            dtype=float,
        )
        _rolled_body, roll_deg = antenna_pattern.apply_roll_toward_station_about_x(
            los_body
        )
        return cls._roll_axes_about_body_x(axes, math.radians(float(roll_deg[0])))

    def _compute_body_axes_from_pointing_profile(
        self,
        *,
        timestamp: datetime,
        sat_pos: np.ndarray,
        sat_vel: np.ndarray | None,
    ) -> np.ndarray | None:
        """Compute body axes matching the configured attitude + selected station pass.

        This mirrors the intent of `access_analysis.py`:
        - Default: prograde (LOF TNW-ish): +X along velocity.
        - If contact switching enabled and in contact: +Z points to station LOS,
          +X constrained in plane of LOS and velocity (projection of v onto plane ⟂ z).
        - If attitude_mode == "nadir": +Z points nadir (toward Earth).
        """
        cfg = getattr(self, "_current_config", None)
        propagation = getattr(cfg, "propagation", None) if cfg is not None else None
        attitude_mode = str(getattr(propagation, "attitude_mode", "prograde") or "prograde").lower()

        r = np.asarray(sat_pos, dtype=float).reshape(3)
        if sat_vel is None:
            return None
        v = np.asarray(sat_vel, dtype=float).reshape(3)
        if float(np.linalg.norm(r)) < 1e-6 or float(np.linalg.norm(v)) < 1e-6:
            return None

        if self._is_comms_pointing_active(timestamp):
            station = self._get_station_position_unoffset(timestamp)
            if station is not None:
                los = np.asarray(station, dtype=float) - r  # spacecraft -> station
                if float(np.linalg.norm(los)) > 1e-6:
                    comms_mode = self._get_comms_pointing_mode()
                    if comms_mode == "free_to_roll":
                        base_axes = self._orthonormal_axes_from_xz(v, -r)
                        if base_axes is None:
                            raise ValueError(
                                "Unable to construct nominal body axes for free-to-roll visualization."
                            )
                        return self._free_to_roll_axes_from_base_and_los(base_axes, los)
                    elif comms_mode == "constrained_aoa":
                        axes = self._constrained_aoa_axes_from_x_los(
                            v,
                            los,
                            max_aoa_deg=self._get_comms_pointing_aoa_limit_deg(),
                        )
                        if axes is not None:
                            return axes

        if attitude_mode == "nadir":
            z = -r  # nadir
            x = v
            return self._orthonormal_axes_from_xz(x, z)

        # Default (non-contact): +Z "down" (nadir), +X "forward" (velocity-constrained).
        # This matches the analysis convention used in `access_analysis.py`.
        z = -r
        x = v
        return self._orthonormal_axes_from_xz(x, z)

    def _resolve_visualization_body_axes(
        self,
        *,
        timestamp: datetime,
        sat_pos: np.ndarray,
        sat_vel: np.ndarray | None,
    ) -> np.ndarray | None:
        """Resolve the body axes that should be rendered for the current frame."""
        if self._is_comms_pointing_active(timestamp):
            if self._get_comms_pointing_mode() == "free_to_roll":
                station = self._get_station_position_unoffset(timestamp)
                if station is None:
                    raise ValueError("Free-to-roll visualization requires a valid station position.")
                base_axes = self._lookup_body_axes_from_attitude(timestamp)
                if base_axes is None:
                    if sat_vel is None:
                        raise ValueError(
                            "Free-to-roll visualization requires either sampled attitude axes or "
                            "a valid satellite velocity vector."
                        )
                    base_axes = self._orthonormal_axes_from_xz(
                        np.asarray(sat_vel, dtype=float).reshape(3),
                        -np.asarray(sat_pos, dtype=float).reshape(3),
                    )
                if base_axes is None:
                    raise ValueError(
                        "Unable to construct base body axes for free-to-roll visualization."
                    )
                los = np.asarray(station, dtype=float) - np.asarray(sat_pos, dtype=float)
                return self._free_to_roll_axes_from_base_and_los(base_axes, los)
            if sat_vel is None:
                raise ValueError("Comms-pointed visualization requires a valid satellite velocity vector.")
            axes = self._compute_body_axes_from_pointing_profile(
                timestamp=timestamp,
                sat_pos=sat_pos,
                sat_vel=sat_vel,
            )
            if axes is None:
                raise ValueError("Unable to compute comms-pointed body axes for visualization.")
            return axes

        axes = self._lookup_body_axes_from_attitude(timestamp)
        if axes is not None:
            return axes
        if sat_vel is None:
            return None
        return self._compute_body_axes_from_pointing_profile(
            timestamp=timestamp,
            sat_pos=sat_pos,
            sat_vel=sat_vel,
        )

    @staticmethod
    def _compute_sensor_cone_outline(
        sat_pos: np.ndarray,
        sensor_axis: np.ndarray,
        *,
        half_angle_rad: float,
        earth_radius_km: float,
        samples: int = 120,
    ) -> np.ndarray | None:
        if samples < 8:
            raise ValueError("Sensor outline requires at least 8 samples.")
        axis = np.asarray(sensor_axis, dtype=float).reshape(3)
        axis_norm = float(np.linalg.norm(axis))
        if axis_norm < 1e-9:
            raise ValueError("Sensor axis is degenerate.")
        axis /= axis_norm
        sat = np.asarray(sat_pos, dtype=float).reshape(3)
        radius = float(earth_radius_km)
        if not (np.isfinite(radius) and radius > 0.0):
            raise ValueError(f"Invalid Earth radius: {earth_radius_km!r}")
        if not (np.isfinite(half_angle_rad) and 0.0 < half_angle_rad < math.pi / 2.0 + 1e-6):
            raise ValueError(f"Invalid sensor half-angle: {half_angle_rad!r}")

        up = np.array([0.0, 0.0, 1.0], dtype=float)
        if abs(float(np.dot(axis, up))) > 0.9:
            up = np.array([0.0, 1.0, 0.0], dtype=float)
        u = np.cross(axis, up)
        u_norm = float(np.linalg.norm(u))
        if u_norm < 1e-9:
            raise ValueError("Failed to build sensor outline basis.")
        u /= u_norm
        v = np.cross(axis, u)

        angles = np.linspace(0.0, 2.0 * math.pi, num=samples, endpoint=False)
        sin_half = float(math.sin(half_angle_rad))
        cos_half = float(math.cos(half_angle_rad))
        points: list[np.ndarray] = []
        disc_epsilon = -1e-6
        for phi in angles:
            perp = math.cos(phi) * u + math.sin(phi) * v
            direction = cos_half * axis + sin_half * perp
            d_norm = float(np.linalg.norm(direction))
            if d_norm < 1e-9:
                raise ValueError("Generated a degenerate sensor direction.")
            d = direction / d_norm
            b = 2.0 * float(np.dot(sat, d))
            c = float(np.dot(sat, sat) - radius * radius)
            disc = b * b - 4.0 * c
            if disc < disc_epsilon:
                continue
            if disc < 0.0:
                disc = 0.0
            sqrt_disc = math.sqrt(disc)
            t1 = (-b - sqrt_disc) / 2.0
            t2 = (-b + sqrt_disc) / 2.0
            t_candidates = [t for t in (t1, t2) if t > 0.0]
            if not t_candidates:
                continue
            t = min(t_candidates)
            points.append(sat + t * d)
        if len(points) < 2:
            return None
        max_step = 3.5 * radius * math.sin(math.pi / float(samples))
        segments: list[np.ndarray] = []
        for i in range(len(points) - 1):
            if float(np.linalg.norm(points[i + 1] - points[i])) <= max_step:
                segments.append(points[i])
                segments.append(points[i + 1])
        if float(np.linalg.norm(points[0] - points[-1])) <= max_step:
            segments.append(points[-1])
            segments.append(points[0])
        if len(segments) < 2:
            return None
        return np.asarray(segments, dtype=np.float32)

    @staticmethod
    def _compute_body_axes_from_pos_vel(
        position_vec: np.ndarray, velocity_vec: np.ndarray
    ) -> np.ndarray | None:
        """Compute a right-handed body triad from position and velocity vectors.

        Convention (LVLH-ish):
        - Body +X: along velocity (prograde)
        - Body +Z: nadir (-position)
        - Body +Y: completes right-handed triad

        Returns:
            3x3 matrix whose columns are [X, Y, Z] unit vectors in world coords.
        """
        r = np.asarray(position_vec, dtype=float).reshape(3)
        v = np.asarray(velocity_vec, dtype=float).reshape(3)
        r_norm = float(np.linalg.norm(r))
        v_norm = float(np.linalg.norm(v))
        if r_norm < 1e-6 or v_norm < 1e-6:
            return None
        z = -r / r_norm
        x = v / v_norm
        y = np.cross(z, x)
        y_norm = float(np.linalg.norm(y))
        if y_norm < 1e-6:
            # Fallback: pick a world axis that isn't parallel to x
            fallback = np.array([0.0, 0.0, 1.0], dtype=float)
            if float(np.linalg.norm(np.cross(fallback, x))) < 1e-6:
                fallback = np.array([0.0, 1.0, 0.0], dtype=float)
            y = np.cross(fallback, x)
            y_norm = float(np.linalg.norm(y))
            if y_norm < 1e-6:
                return None
        y /= y_norm
        # Re-orthogonalize x to ensure orthonormal basis
        x = np.cross(y, z)
        x_norm = float(np.linalg.norm(x))
        if x_norm < 1e-6:
            return None
        x /= x_norm
        z = np.cross(x, y)
        z_norm = float(np.linalg.norm(z))
        if z_norm < 1e-6:
            return None
        z /= z_norm
        return np.stack([x, y, z], axis=1).astype(np.float32)

    def _build_visualization_tab(self) -> QWidget:
        """Create the visualization tab with pass selection and globe view."""
        tab = QWidget()
        tab_layout = QVBoxLayout(tab)
        splitter = QSplitter(Qt.Orientation.Horizontal)
        tab_layout.addWidget(splitter)
        left_panel = QWidget()
        left_layout = QVBoxLayout(left_panel)
        instructions = QLabel(
            "Select a station tab to browse its passes. Use the sort control to reorder passes."
        )
        instructions.setWordWrap(True)
        left_layout.addWidget(instructions)
        sort_row = QHBoxLayout()
        sort_row.addWidget(QLabel("Pass sort:"))
        self.visual_pass_sort_combo = QComboBox()
        self.visual_pass_sort_combo.addItems(["Max elevation", "Date"])
        self.visual_pass_sort_combo.setCurrentIndex(0)
        self.visual_pass_sort_combo.currentIndexChanged.connect(
            self._handle_visual_pass_sort_changed
        )  # type: ignore[attr-defined]
        sort_row.addWidget(self.visual_pass_sort_combo, stretch=1)
        left_layout.addLayout(sort_row)
        frame_toggle_row = QHBoxLayout()
        frame_toggle_row.addWidget(QLabel("Reference frame:"))
        self.visual_frame_combo = QComboBox()
        self.visual_frame_combo.addItems(["ECI (inertial)", "ECEF (earth-fixed)"])
        self.visual_frame_combo.setCurrentIndex(0)
        self.visual_frame_combo.currentIndexChanged.connect(
            self._handle_visual_frame_changed
        )  # type: ignore[attr-defined]
        frame_toggle_row.addWidget(self.visual_frame_combo, stretch=1)
        left_layout.addLayout(frame_toggle_row)
        self.visual_station_tabs = QTabWidget()
        self.visual_station_tabs.setTabPosition(QTabWidget.TabPosition.North)
        self.visual_station_tabs.setDocumentMode(True)
        tab_bar = self.visual_station_tabs.tabBar()
        tab_bar.setElideMode(Qt.TextElideMode.ElideRight)
        tab_bar.setUsesScrollButtons(True)
        tab_bar.setExpanding(False)
        self.visual_station_tabs.setSizePolicy(
            QSizePolicy.Policy.Ignored, QSizePolicy.Policy.Expanding
        )
        self.visual_station_tabs.setMinimumWidth(0)
        left_layout.addWidget(self.visual_station_tabs, stretch=1)
        left_panel.setMinimumWidth(0)
        left_panel.setSizePolicy(
            QSizePolicy.Policy.Ignored, QSizePolicy.Policy.Expanding
        )
        splitter.addWidget(left_panel)
        right_panel = QWidget()
        right_panel.setMinimumWidth(0)
        right_panel.setSizePolicy(
            QSizePolicy.Policy.Ignored, QSizePolicy.Policy.Expanding
        )
        right_layout = QVBoxLayout(right_panel)
        viewer_tab = QWidget()
        viewer_layout = QVBoxLayout(viewer_tab)
        viewer_layout.setContentsMargins(0, 0, 0, 0)
        globe_panel = self._build_visual_globe_panel()
        viewer_layout.addWidget(globe_panel, stretch=1)
        controls_row = QHBoxLayout()
        self.visual_play_button = QPushButton("Play")
        self.visual_play_button.setEnabled(False)
        self.visual_play_button.clicked.connect(self._toggle_visualization_playback)  # type: ignore[attr-defined]
        controls_row.addWidget(self.visual_play_button)
        self.visual_antenna_lobe_checkbox = QCheckBox("Antenna lobe")
        self.visual_antenna_lobe_checkbox.setChecked(False)
        self.visual_antenna_lobe_checkbox.toggled.connect(
            self._handle_visual_antenna_lobe_toggled
        )  # type: ignore[attr-defined]
        controls_row.addWidget(self.visual_antenna_lobe_checkbox)
        self.visual_export_mp4_button = QPushButton("Export MP4")
        self.visual_export_mp4_button.setEnabled(False)
        self.visual_export_mp4_button.clicked.connect(
            self._export_visualization_video
        )  # type: ignore[attr-defined]
        controls_row.addWidget(self.visual_export_mp4_button)
        controls_row.addWidget(QLabel("Speed"))
        self.visual_speed_slider = QSlider(Qt.Orientation.Horizontal)
        self.visual_speed_slider.setRange(25, 300)
        self.visual_speed_slider.setValue(100)
        self.visual_speed_slider.setFixedWidth(120)
        self.visual_speed_slider.valueChanged.connect(
            self._handle_visual_speed_changed
        )  # type: ignore[attr-defined]
        controls_row.addWidget(self.visual_speed_slider)
        self.visual_time_slider = QSlider(Qt.Orientation.Horizontal)
        self.visual_time_slider.setRange(0, 0)
        self.visual_time_slider.setEnabled(False)
        self.visual_time_slider.sliderMoved.connect(
            self._handle_visualization_slider_moved
        )  # type: ignore[attr-defined]
        controls_row.addWidget(self.visual_time_slider, stretch=1)
        self.visual_pass_status_label = QLabel("Run an analysis to populate passes.")
        self.visual_pass_status_label.setWordWrap(True)
        controls_row.addWidget(self.visual_pass_status_label, stretch=1)
        viewer_layout.addLayout(controls_row)
        graphs_tab = self._build_visual_graph_tab()
        self.visual_view_tabs = QTabWidget()
        self.visual_view_tabs.setTabPosition(QTabWidget.TabPosition.North)
        self.visual_view_tabs.setUsesScrollButtons(True)
        self.visual_view_tabs.tabBar().setElideMode(Qt.TextElideMode.ElideRight)
        self.visual_view_tabs.tabBar().setExpanding(False)
        self.visual_view_tabs.addTab(viewer_tab, "Viewer")
        self.visual_view_tabs.addTab(graphs_tab, "Graphs")
        right_layout.addWidget(self.visual_view_tabs, stretch=1)
        splitter.addWidget(right_panel)
        splitter.setChildrenCollapsible(True)
        splitter.setStretchFactor(0, 1)
        splitter.setStretchFactor(1, 3)
        splitter.setSizes([360, 1040])
        self._refresh_visualization_pass_tabs(None)
        return tab

    def _handle_visual_pass_sort_changed(self, _idx: int) -> None:
        """Update pass list ordering based on selected sort."""
        self._refresh_visualization_pass_tabs(getattr(self, "_last_result", None))

    def _build_visual_globe_panel(self) -> QWidget:
        """Create the visualization globe panel showing ground tracks."""
        panel = QWidget()
        layout = QVBoxLayout(panel)
        self.visual_globe_widget = GlobeWidget()
        self.visual_globe_widget.setMinimumWidth(0)
        self.visual_globe_widget.set_earth_render_mode("wireframe")
        self.visual_globe_widget.reset_camera()
        self._update_visual_earth_rotation(None)
        layout.addWidget(self.visual_globe_widget, stretch=1)
        brightness_row = QHBoxLayout()
        brightness_row.addWidget(QLabel("Sun brightness"))
        self.visual_brightness_slider = QSlider(Qt.Orientation.Horizontal)
        self.visual_brightness_slider.setRange(20, 300)
        self.visual_brightness_slider.setValue(100)
        self.visual_brightness_slider.setSingleStep(5)
        self.visual_brightness_slider.valueChanged.connect(
            self._handle_visual_brightness_changed
        )  # type: ignore[attr-defined]
        brightness_row.addWidget(self.visual_brightness_slider, stretch=1)
        layout.addLayout(brightness_row)
        self._handle_visual_brightness_changed(self.visual_brightness_slider.value())
        return panel

    def _build_visual_graph_tab(self) -> QWidget:
        """Create the pass-metric plotting tab."""
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.setContentsMargins(4, 4, 4, 4)
        self.visual_graph_combo = QComboBox()
        self.visual_graph_combo.addItems(
            [
                "Doppler shift",
                "Doppler rate",
                "Normalized Doppler shift (Δf/f)",
                "Normalized Doppler rate (1/s)",
                "LOS antenna gain",
                "Azimuth angle",
                "Elevation angle",
                "Azimuth rate",
                "Elevation rate",
                "Azimuth acceleration",
                "Elevation acceleration",
                "PFD vs Elevation",
            ]
        )
        self.visual_graph_combo.setEnabled(False)
        self.visual_graph_combo.currentIndexChanged.connect(
            self._handle_visual_graph_metric_changed
        )  # type: ignore[attr-defined]
        graph_controls_row = QHBoxLayout()
        graph_controls_row.addWidget(self.visual_graph_combo, stretch=1)
        self.visual_export_npz_button = QPushButton("Export NPZ")
        self.visual_export_npz_button.setEnabled(False)
        self.visual_export_npz_button.clicked.connect(
            lambda: self._export_visualization_series("npz")
        )  # type: ignore[attr-defined]
        graph_controls_row.addWidget(self.visual_export_npz_button)
        self.visual_export_xlsx_button = QPushButton("Export XLSX")
        self.visual_export_xlsx_button.setEnabled(False)
        self.visual_export_xlsx_button.clicked.connect(
            lambda: self._export_visualization_series("xlsx")
        )  # type: ignore[attr-defined]
        graph_controls_row.addWidget(self.visual_export_xlsx_button)
        layout.addLayout(graph_controls_row)
        self.visual_graph_plot = pg.PlotWidget(title="Select a pass to populate graphs.")
        self.visual_graph_plot.setLabel("bottom", "Minutes from AOS", units="min")
        self.visual_graph_plot.showGrid(x=True, y=True, alpha=0.3)
        layout.addWidget(self.visual_graph_plot, stretch=1)
        self._visual_graph_data = None
        self._update_visual_graph_placeholder("Select a pass to populate graphs.")
        return tab

    def _update_visual_graph_placeholder(self, message: str) -> None:
        """Clear the graph panel and show a status message."""
        if getattr(self, "visual_graph_plot", None) is None:
            return
        self.visual_graph_plot.clear()
        self.visual_graph_plot.setTitle(message)
        self.visual_graph_plot.setLabel("bottom", "Minutes from AOS", units="min")
        self.visual_graph_plot.setLabel("left", "", units="")
        if getattr(self, "visual_graph_combo", None) is not None:
            self.visual_graph_combo.setEnabled(False)

    def _reset_visual_graph_panel(self, message: str | None = None) -> None:
        """Reset stored graph data and update placeholder text."""
        self._visual_graph_data = None
        text = message or "Select a pass to populate graphs."
        self._update_visual_graph_placeholder(text)
        combo = getattr(self, "visual_graph_combo", None)
        if combo is not None:
            combo.blockSignals(True)
            combo.setCurrentIndex(0)
            combo.blockSignals(False)

    def _build_visualization_placeholder_panel(self) -> QWidget:
        """Show a hint in the Mission tab pointing to the Visualization tab."""
        panel = QWidget()
        layout = QVBoxLayout(panel)
        label = QLabel(
            "Open the Visualization tab to explore animated passes on the globe."
        )
        label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        label.setWordWrap(True)
        layout.addWidget(label, alignment=Qt.AlignmentFlag.AlignCenter)
        return panel

    def _refresh_visualization_pass_tabs(self, result) -> None:
        """Populate the per-station pass tabs used for visualization playback."""
        if self.visual_station_tabs is None:
            return
        self.visual_station_tabs.blockSignals(True)
        self.visual_station_tabs.clear()
        self._visual_station_lists = {}
        if result is None or not result.passes:
            placeholder = QWidget()
            placeholder_layout = QVBoxLayout(placeholder)
            label = QLabel("Run the analysis to see pass animations.")
            label.setAlignment(Qt.AlignmentFlag.AlignCenter)
            label.setWordWrap(True)
            placeholder_layout.addWidget(label, alignment=Qt.AlignmentFlag.AlignCenter)
            self.visual_station_tabs.addTab(placeholder, "Stations")
            self.visual_station_tabs.blockSignals(False)
            self._clear_visualization_display("Run an analysis to populate passes.")
            self._update_visual_frame_label()
            return
        station_groups: dict[str, list[PassStatistic]] = {}
        for item in result.passes:
            name = item.station_name or "Ground Station"
            station_groups.setdefault(name, []).append(item)
        self._compute_visualization_reference_frames(result)
        station_records: list[tuple[str, float, list[PassStatistic]]] = []
        sort_mode = "max"
        combo = getattr(self, "visual_pass_sort_combo", None)
        if combo is not None and combo.currentText().strip().lower().startswith("date"):
            sort_mode = "date"
        for name, items in station_groups.items():
            if sort_mode == "date":
                sorted_items = sorted(items, key=lambda p: p.aos)
            else:
                sorted_items = sorted(
                    items, key=lambda p: p.max_elevation_deg, reverse=True
                )
            max_elev = sorted_items[0].max_elevation_deg if sorted_items else 0.0
            station_records.append((name, max_elev, sorted_items))
        if sort_mode == "date":
            station_records.sort(key=lambda entry: entry[0].lower())
        else:
            station_records.sort(key=lambda entry: entry[1], reverse=True)
        for station_name, _, passes in station_records:
            tab = QWidget()
            tab_layout = QVBoxLayout(tab)
            list_widget = QListWidget()
            list_widget.setSizeAdjustPolicy(
                QAbstractScrollArea.SizeAdjustPolicy.AdjustIgnored
            )
            list_widget.setHorizontalScrollBarPolicy(
                Qt.ScrollBarPolicy.ScrollBarAsNeeded
            )
            list_widget.setTextElideMode(Qt.TextElideMode.ElideRight)
            list_widget.setSizePolicy(
                QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding
            )
            list_widget.setProperty("station_name", station_name)
            list_widget.itemSelectionChanged.connect(self._handle_visual_pass_selection)  # type: ignore[attr-defined]
            for idx, pass_stat in enumerate(passes, start=1):
                text = (
                    f"{idx:02d}. {pass_stat.aos:%d-%b %H:%M:%S} UTC  |  "
                    f"Max {pass_stat.max_elevation_deg:.1f}°  |  "
                    f"{pass_stat.duration_minutes:.1f} min"
                )
                item = QListWidgetItem(text)
                item.setData(Qt.ItemDataRole.UserRole, pass_stat)
                list_widget.addItem(item)
            tab_layout.addWidget(list_widget)
            self.visual_station_tabs.addTab(tab, station_name)
            self._visual_station_lists[station_name] = list_widget
        self.visual_station_tabs.blockSignals(False)
        if station_records:
            first_station = station_records[0][0]
            first_list = self._visual_station_lists.get(first_station)
            if first_list and first_list.count() > 0:
                first_list.setCurrentRow(0)
        self._update_visual_frame_label()

    def _compute_visualization_reference_frames(self, result) -> None:
        """Initialize reference epoch information for visualization transforms."""
        if self._current_config is not None:
            self._visual_reference_epoch = self._current_config.scenario.start_time
        elif result.passes:
            self._visual_reference_epoch = result.passes[0].aos
        else:
            eph = getattr(result, "ephemeris", None)
            if eph is not None and eph.timestamps_unix is not None and eph.timestamps_unix.size > 0:
                self._visual_reference_epoch = datetime.fromtimestamp(
                    float(eph.timestamps_unix[0]), tz=timezone.utc
                )

    def _handle_visual_pass_selection(self) -> None:
        """React to pass selection within a station tab."""
        widget = self.sender()
        if not isinstance(widget, QListWidget):
            return
        item = widget.currentItem()
        if item is None:
            return
        pass_stat = item.data(Qt.ItemDataRole.UserRole)
        if isinstance(pass_stat, PassStatistic):
            self._load_visualization_pass(pass_stat)

    def _handle_visual_frame_changed(self, index: int) -> None:
        """Switch between ECI and ECEF rendering modes."""
        mode = "ECI" if index == 0 else "ECEF"
        if mode == self._visual_frame_mode:
            return
        self._visual_frame_mode = mode
        if self._visual_animation_timer.isActive():
            self._visual_animation_timer.stop()
            if self.visual_play_button:
                self.visual_play_button.setText("Play")
        self._rebuild_visualization_scene()
        self._update_visual_frame_label()

    def _handle_visual_brightness_changed(self, value: int) -> None:
        """Update sun brightness multiplier from the UI slider."""
        if self.visual_globe_widget is None:
            return
        intensity = max(0.05, value / 100.0)
        self.visual_globe_widget.set_sun_brightness(intensity)

    def _rebuild_visualization_scene(self, *, force: bool = False) -> None:
        """Reapply the visualization assets for the current frame mode."""
        _ = force  # force is kept for backwards compatibility with older calls
        if self.visual_globe_widget is None:
            return
        self._update_visual_earth_rotation(None)
        if self._visual_pass_track is not None:
            self._update_visualization_pass_geometry(self._visual_pass_track)
            self._update_visualization_frame()

    def _update_visual_frame_label(self) -> None:
        """Show the active frame mode in the status label."""
        if not self.visual_pass_status_label:
            return
        prefix = "ECI" if self._visual_frame_mode == "ECI" else "ECEF"
        text = self.visual_pass_status_label.text()
        if " | " in text:
            text = text.split(" | ", 1)[1]
        self.visual_pass_status_label.setText(f"{prefix} | {text}")

    def _load_visualization_pass(self, pass_stat: PassStatistic) -> None:
        """Prepare the globe animation for the selected pass."""
        if self._visual_animation_timer.isActive():
            self._visual_animation_timer.stop()
        track = self._extract_pass_track(pass_stat)
        if not track:
            self._clear_visualization_display("No track samples for the selected pass.")
            return
        self._visual_selected_pass = pass_stat
        self._visual_pass_track = track
        self._visual_animation_index = 0
        self._visual_anim_fraction = 0.0
        self._visual_contact_window = (pass_stat.aos, pass_stat.los)
        self._visual_station_ecef = self._resolve_station_coordinates(pass_stat)
        self._update_visualization_pass_geometry(track)
        slider_max = max(len(track) - 1, 0)
        if self.visual_time_slider:
            self.visual_time_slider.setEnabled(slider_max > 0)
            self.visual_time_slider.setRange(0, slider_max)
            self.visual_time_slider.setPageStep(max(1, slider_max // 20 or 1))
            self.visual_time_slider.setValue(0)
        if track:
            self.visual_globe_widget.set_sun_datetime(track[0].timestamp)
        if self.visual_play_button:
            self.visual_play_button.setEnabled(True)
            self.visual_play_button.setText("Play")
        self._set_visual_video_export_enabled(True)
        if self.visual_pass_status_label:
            station_name = pass_stat.station_name or "Ground Station"
            self.visual_pass_status_label.setText(
                f"{self._visual_frame_mode} | {station_name}: {pass_stat.aos:%d-%b %H:%M:%S} → "
                f"{pass_stat.los:%H:%M:%S} UTC (max {pass_stat.max_elevation_deg:.1f}°)"
            )
        self._update_visualization_frame()
        self._prepare_visualization_graph_data(pass_stat)

    def _prepare_visualization_graph_data(self, pass_stat: PassStatistic) -> None:
        """Extract per-pass look-angle series for graphing."""
        plot_ready = getattr(self, "visual_graph_plot", None) is not None
        if not plot_ready:
            return
        series = self._extract_pass_graph_series(pass_stat)
        if series is None:
            self._reset_visual_graph_panel("No samples available for this pass.")
            self._set_visual_export_buttons_enabled(False)
            return
        self._visual_graph_data = series
        self._set_visual_export_buttons_enabled(True)
        combo = getattr(self, "visual_graph_combo", None)
        if combo is not None:
            combo.setEnabled(True)
        current_index = combo.currentIndex() if combo is not None else 0
        self._render_visual_graph_metric(current_index)

    def _extract_pass_track(self, pass_stat: PassStatistic) -> list[_TrackPoint]:
        """Return the subset of ground-track points covering the selected pass."""
        eph = getattr(self._last_result, "ephemeris", None) if self._last_result is not None else None
        if eph is None or eph.timestamps_unix is None or eph.ecef_pos_km is None:
            return []
        pad_s = 120.0
        t_start = pass_stat.aos.timestamp() - pad_s
        t_end = pass_stat.los.timestamp() + pad_s
        mask = (eph.timestamps_unix >= t_start) & (eph.timestamps_unix <= t_end)
        if not np.any(mask):
            mask = np.ones(eph.timestamps_unix.shape[0], dtype=bool)
        indices = np.where(mask)[0]
        ecef = eph.ecef_pos_km
        ts = eph.timestamps_unix
        return [
            _TrackPoint(
                x_km=float(ecef[i, 0]),
                y_km=float(ecef[i, 1]),
                z_km=float(ecef[i, 2]),
                timestamp=datetime.fromtimestamp(float(ts[i]), tz=timezone.utc),
            )
            for i in indices
        ]

    def _extract_pass_graph_series(
        self, pass_stat: PassStatistic
    ) -> dict[str, np.ndarray] | None:
        """Slice the per-station time series down to the requested pass."""
        result = getattr(self, "_last_result", None)
        tl = getattr(result, "timeline_seconds", None)
        if result is None or tl is None:
            return None
        timeline = np.asarray(result.timeline_seconds, dtype=float)
        if timeline.size == 0:
            return None
        station_series = getattr(result, "station_elevation_series", {})
        station_name = pass_stat.station_name or next(iter(station_series), None)
        if station_name is None:
            return None
        if station_name not in station_series:
            if station_series:
                station_name = next(iter(station_series))
            else:
                return None

        scenario_start = None
        config = getattr(self, "_current_config", None)
        if config is not None and getattr(config, "scenario", None) is not None:
            scenario_start = config.scenario.start_time
        if scenario_start is None:
            scenario_start = self._visual_reference_epoch
        if scenario_start is None:
            return None

        propagation = getattr(config, "propagation", None) if config else None
        sample_step = float(getattr(propagation, "sample_step_seconds", 10.0))

        def _series_from(source: dict[str, list[float]]) -> np.ndarray:
            values = source.get(station_name)
            return (
                np.asarray(values, dtype=float)
                if values is not None
                else np.array([], dtype=float)
            )

        elev = _series_from(station_series)
        azimuth = _series_from(getattr(result, "station_azimuth_series", {}))
        az_rate = _series_from(getattr(result, "station_az_rate_series", {}))
        el_rate = _series_from(getattr(result, "station_el_rate_series", {}))
        range_rate = _series_from(getattr(result, "station_range_rate_series", {}))
        range_accel = _series_from(getattr(result, "station_range_accel_series", {}))
        eph = getattr(result, "ephemeris", None)

        expected = timeline.size
        for array in (elev, azimuth, az_rate, el_rate, range_rate, range_accel):
            if array.size != expected or expected == 0:
                return None
        if eph is None or eph.ecef_pos_km is None or eph.ecef_pos_km.shape[0] != expected:
            return None
        station_cfg = self._resolve_visual_station_config(station_name)
        if station_cfg is None:
            return None
        station_ecef_m = 1000.0 * np.asarray(
            self._station_to_ecef_km(station_cfg),
            dtype=float,
        )
        lut = self._get_visual_antenna_lut()
        sat_ecef_m = eph.ecef_pos_km * 1000.0
        body_x_ecef = eph.body_x_ecef if eph.body_x_ecef is not None else np.full((expected, 3), np.nan)
        body_y_ecef = eph.body_y_ecef if eph.body_y_ecef is not None else np.full((expected, 3), np.nan)
        body_z_ecef = eph.body_z_ecef if eph.body_z_ecef is not None else np.full((expected, 3), np.nan)
        cfg = getattr(self, "_current_config", None)
        propagation = getattr(cfg, "propagation", None) if cfg is not None else None
        threshold_deg = float(getattr(propagation, "contact_elevation_deg", 10.0))
        steering_active_mask = np.isfinite(elev) & (elev >= threshold_deg)
        los_antenna_gain_dbi, _az_body_deg, _el_body_deg, roll_command_deg = (
            antenna_pattern.evaluate_station_gain_series(
                lut=lut,
                station=station_cfg,
                sat_ecef_m=sat_ecef_m,
                body_x_ecef=body_x_ecef,
                body_y_ecef=body_y_ecef,
                body_z_ecef=body_z_ecef,
                pointing_mode=self._get_comms_pointing_mode(),
                max_aoa_deg=self._get_comms_pointing_aoa_limit_deg(),
                steering_active_mask=steering_active_mask,
            )
        )
        slant_range_m = np.linalg.norm(sat_ecef_m - station_ecef_m.reshape(1, 3), axis=1)

        aos_sec = (pass_stat.aos - scenario_start).total_seconds()
        los_sec = (pass_stat.los - scenario_start).total_seconds()
        pad_seconds = max(sample_step, 5.0)
        mask = (timeline >= aos_sec - pad_seconds) & (timeline <= los_sec + pad_seconds)
        if not np.any(mask):
            return None

        return {
            "times_minutes": (timeline[mask] - aos_sec) / 60.0,
            "azimuth_deg": azimuth[mask],
            "elevation_deg": elev[mask],
            "los_antenna_gain_dbi": los_antenna_gain_dbi[mask],
            "slant_range_m": slant_range_m[mask],
            "roll_command_deg": roll_command_deg[mask],
            "az_rate_deg_s": az_rate[mask],
            "el_rate_deg_s": el_rate[mask],
            "range_rate_mps": range_rate[mask],
            "range_accel_mps2": range_accel[mask],
            "station_name": station_name,
        }

    def _resolve_visual_station_config(
        self, station_name: str
    ) -> GroundStationConfig | None:
        """Resolve a station config by name for graph/LUT evaluation."""
        if station_name:
            station = self._active_station_lookup.get(station_name)
            if station is not None:
                return station
            station = next(
                (cfg for cfg in self._station_presets if cfg.name == station_name),
                None,
            )
            if station is not None:
                return station
        return self._station_presets[0] if self._station_presets else None

    def _get_visual_antenna_lut(self) -> antenna_pattern.SphericalGainLut:
        """Load and cache the active antenna LUT for pass visualization."""
        path_text = str(getattr(self, "_antenna_lut_path", "") or "").strip()
        lut_path = (
            Path(path_text).expanduser().resolve()
            if path_text
            else antenna_pattern.default_synthesized_lut_path().expanduser().resolve()
        )
        cache = getattr(self, "_visual_antenna_lut_cache", None)
        cache_key = getattr(self, "_visual_antenna_lut_cache_key", None)
        if cache_key == str(lut_path) and isinstance(cache, antenna_pattern.SphericalGainLut):
            return cache
        lut = antenna_pattern.load_spherical_gain_lut(lut_path)
        self._visual_antenna_lut_cache = lut
        self._visual_antenna_lut_cache_key = str(lut_path)
        return lut

    @staticmethod
    def _antenna_gain_to_rgba(gain_dbi: np.ndarray, max_gain_dbi: float) -> np.ndarray:
        """Map gain linearly from blue (0 dB) to red (max gain) using HSV hue."""
        gain = np.asarray(gain_dbi, dtype=float)
        if gain.ndim != 1:
            raise ValueError("Antenna gain color mapping expects a 1D gain array.")
        max_gain = float(max_gain_dbi)
        if not np.isfinite(max_gain) or max_gain <= 0.0:
            raise ValueError(f"Invalid maximum antenna gain for color mapping: {max_gain_dbi!r}")
        norm = np.clip(gain, 0.0, max_gain) / max_gain
        hue = (1.0 - norm) * (240.0 / 360.0)
        h6 = hue * 6.0
        sector = np.floor(h6).astype(int)
        frac = h6 - sector
        p = np.zeros_like(norm, dtype=np.float32)
        q = 1.0 - frac
        t = frac
        rgb = np.zeros((gain.shape[0], 3), dtype=np.float32)
        s0 = (sector % 6) == 0
        s1 = (sector % 6) == 1
        s2 = (sector % 6) == 2
        s3 = (sector % 6) == 3
        s4 = (sector % 6) == 4
        s5 = (sector % 6) == 5
        rgb[s0] = np.column_stack([np.ones(np.count_nonzero(s0), dtype=np.float32), t[s0], p[s0]])
        rgb[s1] = np.column_stack([q[s1], np.ones(np.count_nonzero(s1), dtype=np.float32), p[s1]])
        rgb[s2] = np.column_stack([p[s2], np.ones(np.count_nonzero(s2), dtype=np.float32), t[s2]])
        rgb[s3] = np.column_stack([p[s3], q[s3], np.ones(np.count_nonzero(s3), dtype=np.float32)])
        rgb[s4] = np.column_stack([t[s4], p[s4], np.ones(np.count_nonzero(s4), dtype=np.float32)])
        rgb[s5] = np.column_stack([np.ones(np.count_nonzero(s5), dtype=np.float32), p[s5], q[s5]])
        alpha = np.full((gain.shape[0], 1), 0.45, dtype=np.float32)
        return np.hstack([rgb, alpha])

    def _get_visual_antenna_lobe_template(
        self,
    ) -> tuple[np.ndarray, np.ndarray]:
        """Return cached body-local antenna-lobe vertices and colors."""
        widget = getattr(self, "visual_globe_widget", None)
        if widget is None:
            raise ValueError("Visualization globe widget is not available.")
        lut = self._get_visual_antenna_lut()
        scale_km = float(getattr(widget, "_satellite_scale_km", 200.0) or 200.0)
        local_extent = float(getattr(widget, "_satellite_local_extent", 0.5) or 0.5)
        cache_key = (
            str(getattr(self, "_visual_antenna_lut_cache_key", "")),
            round(scale_km, 6),
            round(local_extent, 6),
            lut.gain_dbi_grid.shape,
        )
        cached_key = getattr(self, "_visual_antenna_lobe_template_cache_key", None)
        cached_template = getattr(self, "_visual_antenna_lobe_template_cache", None)
        if (
            cache_key == cached_key
            and isinstance(cached_template, tuple)
            and len(cached_template) == 2
        ):
            return cached_template

        root_offset_km = max(scale_km * local_extent * 0.55 / 2.0, 60.0)
        lobe_length_km = max(scale_km * local_extent * 4.5 / 2.0, 500.0)
        max_gain_dbi = float(np.max(lut.gain_dbi_grid))
        if not np.isfinite(max_gain_dbi) or max_gain_dbi <= 0.0:
            raise ValueError(f"Invalid maximum antenna gain in LUT: {max_gain_dbi!r}")

        az_samples = np.linspace(0.0, 360.0, 49, dtype=float)
        el_samples = np.linspace(0.0, 90.0, 50, dtype=float)
        az_mesh, el_mesh = np.meshgrid(az_samples, el_samples, indexing="ij")
        gain_mesh = lut.gain_dbi(az_mesh.reshape(-1), el_mesh.reshape(-1)).reshape(az_mesh.shape)
        gain_clamped = np.clip(gain_mesh, 0.0, max_gain_dbi)
        radius_scale = gain_clamped / max_gain_dbi

        az_rad = np.deg2rad(az_mesh)
        el_rad = np.deg2rad(el_mesh)
        cos_el = np.cos(el_rad)
        local_dirs = np.stack(
            [
                cos_el * np.cos(az_rad),
                cos_el * np.sin(az_rad),
                np.sin(el_rad),
            ],
            axis=-1,
        )
        local_root = np.array([0.0, 0.0, root_offset_km], dtype=np.float32)
        local_points = local_root + (
            lobe_length_km * radius_scale[..., None] * local_dirs
        ).astype(np.float32)
        local_colors = self._antenna_gain_to_rgba(
            gain_clamped.reshape(-1),
            max_gain_dbi,
        ).reshape(az_mesh.shape + (4,))

        p00 = local_points[:-1, :-1]
        p10 = local_points[1:, :-1]
        p01 = local_points[:-1, 1:]
        p11 = local_points[1:, 1:]
        c00 = local_colors[:-1, :-1]
        c10 = local_colors[1:, :-1]
        c01 = local_colors[:-1, 1:]
        c11 = local_colors[1:, 1:]

        vertices = np.stack([p00, p10, p11, p00, p11, p01], axis=2).reshape(-1, 3)
        colors = np.stack([c00, c10, c11, c00, c11, c01], axis=2).reshape(-1, 4)

        template = (
            np.asarray(vertices, dtype=np.float32),
            np.asarray(colors, dtype=np.float32),
        )
        self._visual_antenna_lobe_template_cache_key = cache_key
        self._visual_antenna_lobe_template_cache = template
        return template

    def _build_visual_antenna_lobe_coords(
        self,
        *,
        sat_coords: np.ndarray,
        body_axes: np.ndarray,
    ) -> tuple[np.ndarray, np.ndarray]:
        """Build a filled gain-colored lobe attached to the spacecraft +Z face."""
        axes = np.asarray(body_axes, dtype=float)
        if axes.shape != (3, 3):
            raise ValueError("Antenna lobe visualization requires a 3x3 body-axis matrix.")
        sat = np.asarray(sat_coords, dtype=float).reshape(3)
        local_vertices, colors = self._get_visual_antenna_lobe_template()
        world_vertices = local_vertices @ axes.T + sat
        return np.asarray(world_vertices, dtype=np.float32), colors

    def _handle_visual_antenna_lobe_toggled(self, checked: bool) -> None:
        """Show or hide the spacecraft antenna lobe overlay."""
        widget = getattr(self, "visual_globe_widget", None)
        if widget is None:
            return
        if not checked:
            widget.update_sensor_cone_outline(None)
            return
        if self._visual_pass_track:
            self._update_visualization_frame()

    def _handle_visual_graph_metric_changed(self, index: int) -> None:
        """Update the pass-metric plot when the selected metric changes."""
        self._render_visual_graph_metric(index)

    def _render_visual_graph_metric(self, index: int) -> None:
        """Render the selected pass metric."""
        plot = getattr(self, "visual_graph_plot", None)
        data = getattr(self, "_visual_graph_data", None)
        if plot is None:
            return
        if not data:
            self._update_visual_graph_placeholder("Select a pass to populate graphs.")
            return
        times = data.get("times_minutes")
        if times is None or len(times) == 0:
            self._update_visual_graph_placeholder("No samples available for this pass.")
            return
        times = np.asarray(times, dtype=float)

        station_name = data.get("station_name", "Station")
        if index == 0:
            values = self._compute_doppler_shift(data.get("range_rate_mps"))
            y_label = "Doppler shift"
            units = "Hz"
            title = "Doppler Shift"
        elif index == 1:
            values = self._compute_doppler_rate(data.get("range_accel_mps2"))
            y_label = "Doppler rate"
            units = "Hz/s"
            title = "Doppler Rate"
        elif index == 2:
            values = self._compute_normalized_doppler_shift(data.get("range_rate_mps"))
            y_label = "Normalized Doppler shift"
            units = "Δf/f"
            title = "Normalized Doppler Shift"
        elif index == 3:
            values = self._compute_normalized_doppler_rate(data.get("range_accel_mps2"))
            y_label = "Normalized Doppler rate"
            units = "1/s"
            title = "Normalized Doppler Rate"
        elif index == 4:
            values = np.asarray(data.get("los_antenna_gain_dbi"))
            y_label = "LOS antenna gain"
            units = "dBi"
            title = "LOS Antenna Gain"
        elif index == 5:
            values = np.asarray(data.get("azimuth_deg"))
            y_label = "Azimuth"
            units = "deg"
            title = "Azimuth Angle"
        elif index == 6:
            values = np.asarray(data.get("elevation_deg"))
            y_label = "Elevation"
            units = "deg"
            title = "Elevation Angle"
        elif index == 7:
            values = np.asarray(data.get("az_rate_deg_s"))
            y_label = "Azimuth rate"
            units = "deg/s"
            title = "Required Azimuth Rate"
        elif index == 8:
            values = np.asarray(data.get("el_rate_deg_s"))
            y_label = "Elevation rate"
            units = "deg/s"
            title = "Required Elevation Rate"
        elif index == 9:
            rate = np.asarray(data.get("az_rate_deg_s"))
            values = self._compute_angular_acceleration(rate, times)
            y_label = "Azimuth acceleration"
            units = "deg/s²"
            title = "Required Azimuth Acceleration"
        elif index == 10:
            rate = np.asarray(data.get("el_rate_deg_s"))
            values = self._compute_angular_acceleration(rate, times)
            y_label = "Elevation acceleration"
            units = "deg/s²"
            title = "Required Elevation Acceleration"
        elif index == 11:
            x_values, values, limit_x, limit_y, units = self._compute_visual_pfd_series(
                data
            )
            x_values, values = self._downsample_graph_series(x_values, values)
            y_label = "PFD"
            title = "PFD vs Elevation"
        else:
            values = np.asarray(data.get("elevation_deg"))
            y_label = "Elevation"
            units = "deg"
            title = "Elevation Angle"

        if index != 11 and values.size != times.size:
            self._update_visual_graph_placeholder("Incomplete samples for this metric.")
            return

        plot.clear()
        plot.setTitle(f"{title} — {station_name}")
        plot.setLabel("left", y_label, units=units)
        if index == 11:
            plot.setLabel("bottom", "Elevation", units="deg")
            plot.plot(limit_x, limit_y, pen=pg.mkPen("#ff5252", width=2))
            plot.plot(x_values, values, pen=pg.mkPen("#4f8cff", width=2))
        else:
            times_ds, values_ds = self._downsample_graph_series(times, values)
            plot.setLabel("bottom", "Minutes from AOS", units="min")
            plot.plot(times_ds, values_ds, pen=pg.mkPen("#76c7ff", width=2))
        plot.getViewBox().autoRange()

    def _compute_angular_acceleration(
        self, rate_deg_s: np.ndarray | None, times_minutes: np.ndarray
    ) -> np.ndarray:
        """Numerically differentiate an angular-rate series to obtain acceleration.

        Args:
            rate_deg_s: Angular rate in deg/s.
            times_minutes: Time axis in minutes (same length as rate_deg_s).

        Returns:
            Angular acceleration in deg/s² (same length as inputs).
        """
        if rate_deg_s is None:
            return np.array([])
        rate = np.asarray(rate_deg_s, dtype=float)
        if rate.size == 0:
            return np.array([])
        if rate.size == 1:
            return np.zeros_like(rate)
        t_s = np.asarray(times_minutes, dtype=float) * 60.0
        if t_s.size != rate.size:
            raise ValueError(
                "Cannot compute angular acceleration: time axis length does not match rate series "
                f"(t_s.size={int(t_s.size)}, rate.size={int(rate.size)})"
            )
        if not np.all(np.isfinite(t_s)):
            raise ValueError("Cannot compute angular acceleration: time axis contains non-finite values")
        if np.nanmax(t_s) == np.nanmin(t_s):
            raise ValueError("Cannot compute angular acceleration: time axis is degenerate")
        # Suppress divide-by-zero from duplicate timestamps; replace non-finite
        # results with zero (acceleration is undefined when dt == 0).
        with np.errstate(divide="ignore", invalid="ignore"):
            result = np.gradient(rate, t_s)
        return np.where(np.isfinite(result), result, 0.0)

    def _compute_doppler_shift(self, range_rate: np.ndarray | None) -> np.ndarray:
        """Convert the line-of-sight range rate into Doppler shift."""
        if range_rate is None:
            return np.array([])
        freq_input = getattr(self, "lb_frequency_input", None)
        freq_ghz = float(freq_input.value()) if freq_input is not None else 8.2
        freq_hz = max(freq_ghz, 0.0) * 1e9
        if freq_hz <= 0.0:
            return np.zeros_like(range_rate)
        scale = freq_hz / SPEED_OF_LIGHT_MPS
        return -np.asarray(range_rate) * scale

    def _compute_doppler_rate(self, range_accel: np.ndarray | None) -> np.ndarray:
        """Convert line-of-sight acceleration into Doppler rate."""
        if range_accel is None:
            return np.array([])
        freq_input = getattr(self, "lb_frequency_input", None)
        freq_ghz = float(freq_input.value()) if freq_input is not None else 8.2
        freq_hz = max(freq_ghz, 0.0) * 1e9
        if freq_hz <= 0.0:
            return np.zeros_like(range_accel)
        scale = freq_hz / SPEED_OF_LIGHT_MPS
        return -np.asarray(range_accel) * scale

    def _compute_normalized_doppler_shift(
        self, range_rate: np.ndarray | None
    ) -> np.ndarray:
        """Convert LOS range rate into non-dimensional Doppler shift, Δf/f = -v_r/c."""
        if range_rate is None:
            return np.array([])
        return -np.asarray(range_rate, dtype=float) / float(SPEED_OF_LIGHT_MPS)

    def _compute_normalized_doppler_rate(
        self, range_accel: np.ndarray | None
    ) -> np.ndarray:
        """Convert LOS acceleration into non-dimensional Doppler rate, d(Δf/f)/dt = -a_r/c."""
        if range_accel is None:
            return np.array([])
        return -np.asarray(range_accel, dtype=float) / float(SPEED_OF_LIGHT_MPS)

    def _downsample_graph_series(
        self, times: np.ndarray, values: np.ndarray, max_points: int = 2000
    ) -> tuple[np.ndarray, np.ndarray]:
        """Reduce plotted sample count to keep PyQtGraph responsive."""
        if times.size <= max_points:
            return times, values
        indices = np.linspace(0, times.size - 1, max_points, dtype=int)
        indices = np.unique(indices)
        return times[indices], values[indices]

    def _handle_visual_frequency_changed(self, _value: float) -> None:
        """Re-render metrics that depend on link-budget controls."""
        combo = getattr(self, "visual_graph_combo", None)
        if combo is None:
            return
        current = combo.currentIndex()
        if current not in (0, 1, 11):
            return
        self._render_visual_graph_metric(current)

    def _attach_visualization_frequency_listener(self) -> None:
        """Connect link-budget controls so graph metrics stay in sync."""
        if getattr(self, "_visual_freq_listener_connected", False):
            return
        controls = [
            getattr(self, "lb_frequency_input", None),
            getattr(self, "lb_tx_power_input", None),
            getattr(self, "lb_tx_losses_input", None),
            getattr(self, "lb_tx_backoff_input", None),
            getattr(self, "lb_symbol_rate_input", None),
            getattr(self, "lb_rolloff_input", None),
        ]
        if any(control is None for control in controls):
            raise ValueError("Link-budget controls must exist before attaching visualization listeners.")
        for control in controls:
            control.valueChanged.connect(self._handle_visual_frequency_changed)  # type: ignore[attr-defined]
        self._visual_freq_listener_connected = True

    def _compute_visual_pfd_series(
        self, data: dict[str, np.ndarray]
    ) -> tuple[np.ndarray, np.ndarray, np.ndarray, np.ndarray, str]:
        """Return actual and limit PFD curves for the selected pass."""
        elevations = np.asarray(data.get("elevation_deg", np.array([])), dtype=float)
        gains_dbi = np.asarray(data.get("los_antenna_gain_dbi", np.array([])), dtype=float)
        slant_range_m = np.asarray(data.get("slant_range_m", np.array([])), dtype=float)
        if elevations.size == 0:
            raise ValueError("No elevation samples available for PFD plotting.")
        if gains_dbi.size != elevations.size or slant_range_m.size != elevations.size:
            raise ValueError("PFD plotting requires gain and slant-range series for every elevation sample.")

        frequency_input = getattr(self, "lb_frequency_input", None)
        tx_power_input = getattr(self, "lb_tx_power_input", None)
        tx_losses_input = getattr(self, "lb_tx_losses_input", None)
        tx_backoff_input = getattr(self, "lb_tx_backoff_input", None)
        symbol_rate_input = getattr(self, "lb_symbol_rate_input", None)
        rolloff_input = getattr(self, "lb_rolloff_input", None)
        controls = [
            frequency_input,
            tx_power_input,
            tx_losses_input,
            tx_backoff_input,
            symbol_rate_input,
            rolloff_input,
        ]
        if any(control is None for control in controls):
            raise ValueError("Link-budget controls are required for PFD plotting.")

        valid_mask = (
            np.isfinite(elevations)
            & np.isfinite(gains_dbi)
            & np.isfinite(slant_range_m)
            & (elevations >= 0.0)
            & (elevations <= 90.0)
            & (slant_range_m > 0.0)
        )
        if not np.any(valid_mask):
            raise ValueError("No valid in-domain elevation samples are available for PFD plotting.")

        occupied_bandwidth_hz = pfd_math.occupied_bandwidth_hz(
            float(symbol_rate_input.value()) * 1e6,
            float(rolloff_input.value()),
        )
        frequency_mhz = float(frequency_input.value()) * 1000.0
        mask_spec = pfd_math.itu_surface_pfd_mask_spec(frequency_mhz)
        directional_eirp = pfd_math.directional_eirp_dBW(
            tx_power_dBw=float(tx_power_input.value()),
            tx_losses_dB=float(tx_losses_input.value()),
            tx_backoff_dB=float(tx_backoff_input.value()),
            antenna_gain_dBi=gains_dbi[valid_mask],
        )
        pfd_curve = pfd_math.pfd_at_reference_bandwidth_dBW_per_m2(
            directional_eirp_dBW=directional_eirp,
            slant_range_m=slant_range_m[valid_mask],
            occupied_bandwidth_hz=occupied_bandwidth_hz,
            reference_bandwidth_hz=mask_spec.reference_bandwidth_hz,
        )
        limit_x = np.linspace(0.0, 90.0, 901, dtype=float)
        limit_y = pfd_math.itu_surface_pfd_limit_dBW_per_m2(
            limit_x,
            frequency_MHz=frequency_mhz,
        )
        unit_text = (
            "dBW/m^2/4 kHz"
            if np.isclose(mask_spec.reference_bandwidth_hz, 4000.0)
            else "dBW/m^2/1 MHz"
        )
        return (
            elevations[valid_mask],
            pfd_curve,
            limit_x,
            limit_y,
            unit_text,
        )

    def _resolve_station_coordinates(
        self, pass_stat: PassStatistic
    ) -> tuple[float, float, float] | None:
        """Resolve the ECEF coordinates for the pass station."""
        station: GroundStationConfig | None = None
        station_name = pass_stat.station_name or ""
        if station_name:
            station = self._active_station_lookup.get(station_name)
            if station is None:
                station = next(
                    (s for s in self._station_presets if s.name == station_name),
                    None,
                )
        if station is None and self._station_presets:
            station = self._station_presets[0]
        if station is None:
            return None
        return self._station_to_ecef_km(station)

    def _station_to_ecef_km(
        self, station: GroundStationConfig
    ) -> tuple[float, float, float]:
        """Convert a ground-station lat/lon/alt to pure ECEF coordinates in kilometers."""
        lat = math.radians(station.latitude_deg)
        lon = math.radians(station.longitude_deg)
        alt_km = station.altitude_m / 1000.0
        a = 6378.137
        e2 = 6.69437999014e-3
        sin_lat = math.sin(lat)
        N = a / math.sqrt(1 - e2 * sin_lat * sin_lat)
        x = (N + alt_km) * math.cos(lat) * math.cos(lon)
        y = (N + alt_km) * math.cos(lat) * math.sin(lon)
        z = (N * (1 - e2) + alt_km) * sin_lat
        return (x, y, z)

    def _seconds_since_reference(self, timestamp: datetime) -> float:
        """Return elapsed seconds from the visualization reference epoch."""
        ref = self._visual_reference_epoch
        if ref is None:
            return 0.0
        return (timestamp - ref).total_seconds()

    def _convert_vector_to_current_frame(
        self, vector: tuple[float, float, float], timestamp: datetime
    ) -> tuple[float, float, float]:
        """Convert a globe-space vector into the currently selected frame."""
        if self._visual_frame_mode != "ECI":
            return vector
        # Convert ECEF(globe) -> ECI(globe) using absolute GMST.
        # Our world +Z matches Earth's rotation axis, so a simple Z-rotation is sufficient.
        angle = gmst_rad_utc(timestamp)
        return rotate_vector_z(vector, angle)

    def _get_visualization_track_coords(
        self, track: list[_TrackPoint]
    ) -> np.ndarray:
        """Return track coordinates transformed into the current frame."""
        coords = np.array(
            [ecef_to_globe_coords(pt.x_km, pt.y_km, pt.z_km) for pt in track],
            dtype=float,
        )
        if coords.size == 0 or self._visual_frame_mode != "ECI":
            return coords
        # Convert ECEF(globe) -> ECI(globe) using absolute GMST per-sample.
        angles = np.array([gmst_rad_utc(pt.timestamp) for pt in track], dtype=float)
        cos_ang = np.cos(angles)
        sin_ang = np.sin(angles)
        x = coords[:, 0].copy()
        y = coords[:, 1].copy()
        coords[:, 0] = cos_ang * x - sin_ang * y
        coords[:, 1] = sin_ang * x + cos_ang * y
        return coords

    def _update_visualization_pass_geometry(
        self, track: list[_TrackPoint]
    ) -> None:
        """Render the selected pass path and initialize the satellite marker."""
        if self.visual_globe_widget is None or not track:
            return
        coords = self._get_visualization_track_coords(track)
        if coords.size == 0:
            return
        first_timestamp = track[0].timestamp if track else None
        self._update_visual_earth_rotation(first_timestamp)
        if first_timestamp:
            self.visual_globe_widget.set_sun_datetime(first_timestamp)
        self.visual_globe_widget.update_track(coords)
        axes = None
        if first_timestamp is not None:
            vel = coords[1] - coords[0] if len(coords) >= 2 else None
            axes = self._resolve_visualization_body_axes(
                timestamp=first_timestamp,
                sat_pos=coords[0],
                sat_vel=vel,
            )
        self.visual_globe_widget.update_satellite_position(tuple(coords[0]), body_axes=axes)
        self._update_visualization_sensor_outline(first_timestamp, coords[0], axes)
        self.visual_globe_widget.update_link_segment(None, None)
        self.visual_globe_widget.update_direction_arrow(None, None)

    def _clear_visualization_display(self, message: str | None = None) -> None:
        """Reset visualization playback state and remove temporary actors."""
        if self._visual_animation_timer.isActive():
            self._visual_animation_timer.stop()
        if self.visual_globe_widget is not None:
            self.visual_globe_widget.update_track(None)
            self.visual_globe_widget.update_satellite_position(None)
            self.visual_globe_widget.update_link_segment(None, None)
            self.visual_globe_widget.update_direction_arrow(None, None)
            self.visual_globe_widget.update_sensor_cone_outline(None)
            self.visual_globe_widget.set_sun_datetime(None)
            self.visual_globe_widget.update_direction_arrow(None, None)
            self.visual_globe_widget.set_focus_point(None)
        self._visual_pass_track = None
        self._visual_selected_pass = None
        self._visual_station_ecef = None
        self._visual_contact_window = None
        self._visual_animation_index = 0
        self._visual_anim_fraction = 0.0
        if self.visual_time_slider:
            self.visual_time_slider.setEnabled(False)
            self.visual_time_slider.setRange(0, 0)
            self.visual_time_slider.setValue(0)
        if self.visual_play_button:
            self.visual_play_button.setEnabled(False)
            self.visual_play_button.setText("Play")
        self._set_visual_export_buttons_enabled(False)
        self._set_visual_video_export_enabled(False)
        if message and self.visual_pass_status_label:
            self.visual_pass_status_label.setText(message)
        self._reset_visual_graph_panel(message)

    def _set_visual_export_buttons_enabled(self, enabled: bool) -> None:
        """Enable/disable the Viewer export buttons based on pass availability."""
        for attr in ("visual_export_npz_button", "visual_export_xlsx_button"):
            button = getattr(self, attr, None)
            if button is not None:
                button.setEnabled(bool(enabled))

    def _set_visual_video_export_enabled(self, enabled: bool) -> None:
        """Enable or disable the pass video export button."""
        button = getattr(self, "visual_export_mp4_button", None)
        if button is not None:
            button.setEnabled(bool(enabled))

    def _export_visualization_series(self, format_type: str) -> None:
        """Export the selected pass series (Doppler/az/el + rates/accelerations) to NPZ or XLSX."""
        pass_stat = getattr(self, "_visual_selected_pass", None)
        data = getattr(self, "_visual_graph_data", None)
        if pass_stat is None or not data:
            QMessageBox.warning(
                self,
                "No Data",
                "No pass series to export. Select a pass in the Visualization tab first.",
            )
            return

        times_minutes = np.asarray(data.get("times_minutes", np.array([])), dtype=float)
        if times_minutes.size == 0:
            QMessageBox.warning(
                self,
                "No Data",
                "No samples available for the selected pass.",
            )
            return

        azimuth_deg = np.asarray(data.get("azimuth_deg", np.array([])), dtype=float)
        elevation_deg = np.asarray(data.get("elevation_deg", np.array([])), dtype=float)
        az_rate_deg_s = np.asarray(data.get("az_rate_deg_s", np.array([])), dtype=float)
        el_rate_deg_s = np.asarray(data.get("el_rate_deg_s", np.array([])), dtype=float)
        range_rate_mps = np.asarray(data.get("range_rate_mps", np.array([])), dtype=float)
        range_accel_mps2 = np.asarray(
            data.get("range_accel_mps2", np.array([])), dtype=float
        )

        doppler_shift_hz = self._compute_doppler_shift(range_rate_mps)
        doppler_rate_hz_s = self._compute_doppler_rate(range_accel_mps2)
        doppler_shift_norm = self._compute_normalized_doppler_shift(range_rate_mps)
        doppler_rate_norm_s = self._compute_normalized_doppler_rate(range_accel_mps2)
        az_accel_deg_s2 = self._compute_angular_acceleration(az_rate_deg_s, times_minutes)
        el_accel_deg_s2 = self._compute_angular_acceleration(el_rate_deg_s, times_minutes)

        expected = times_minutes.size
        for series in (
            azimuth_deg,
            elevation_deg,
            az_rate_deg_s,
            el_rate_deg_s,
            az_accel_deg_s2,
            el_accel_deg_s2,
            doppler_shift_hz,
            doppler_rate_hz_s,
            doppler_shift_norm,
            doppler_rate_norm_s,
        ):
            if series.size != expected:
                QMessageBox.critical(
                    self,
                    "Export Failed",
                    "Series length mismatch while preparing export data.",
                )
                return

        aos = pass_stat.aos
        station_name = pass_stat.station_name or data.get("station_name") or "Station"
        safe_station = re.sub(r"[^A-Za-z0-9_-]+", "_", str(station_name)).strip("_")
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        pass_stamp = aos.strftime("%Y%m%d_%H%M%S")

        outputs_dir = Path(__file__).resolve().parents[3] / "outputs"
        outputs_dir.mkdir(exist_ok=True)

        headers = [
            "station",
            "pass_aos_utc",
            "pass_los_utc",
            "timestamp_utc",
            "minutes_from_aos",
            "doppler_shift_hz",
            "doppler_rate_hz_s",
            "doppler_shift_norm",
            "doppler_rate_norm_s",
            "azimuth_deg",
            "azimuth_rate_deg_s",
            "azimuth_accel_deg_s2",
            "elevation_deg",
            "elevation_rate_deg_s",
            "elevation_accel_deg_s2",
        ]

        rows: list[list[object]] = []
        for idx in range(expected):
            ts = aos + timedelta(minutes=float(times_minutes[idx]))
            rows.append(
                [
                    station_name,
                    pass_stat.aos.strftime("%d-%b-%Y %H:%M:%S"),
                    pass_stat.los.strftime("%d-%b-%Y %H:%M:%S"),
                    ts.strftime("%d-%b-%Y %H:%M:%S"),
                    float(times_minutes[idx]),
                    float(doppler_shift_hz[idx]),
                    float(doppler_rate_hz_s[idx]),
                    float(doppler_shift_norm[idx]),
                    float(doppler_rate_norm_s[idx]),
                    float(azimuth_deg[idx]),
                    float(az_rate_deg_s[idx]),
                    float(az_accel_deg_s2[idx]),
                    float(elevation_deg[idx]),
                    float(el_rate_deg_s[idx]),
                    float(el_accel_deg_s2[idx]),
                ]
            )

        try:
            if format_type == "npz":
                filename = (
                    outputs_dir / f"pass_series_{safe_station}_{pass_stamp}_{timestamp}.npz"
                )
                payload = {
                    "schema_version": np.array([1], dtype=int),
                    "station_name": np.array([station_name], dtype=object),
                    "pass_aos_utc": np.array([pass_stat.aos.strftime("%d-%b-%Y %H:%M:%S")], dtype=object),
                    "pass_los_utc": np.array([pass_stat.los.strftime("%d-%b-%Y %H:%M:%S")], dtype=object),
                    "headers": np.array(headers, dtype=object),
                    "timestamp_utc": np.array([row[3] for row in rows], dtype=object),
                    "minutes_from_aos": np.asarray([row[4] for row in rows], dtype=float),
                    "doppler_shift_hz": np.asarray([row[5] for row in rows], dtype=float),
                    "doppler_rate_hz_s": np.asarray([row[6] for row in rows], dtype=float),
                    "doppler_shift_norm": np.asarray([row[7] for row in rows], dtype=float),
                    "doppler_rate_norm_s": np.asarray([row[8] for row in rows], dtype=float),
                    "azimuth_deg": np.asarray([row[9] for row in rows], dtype=float),
                    "azimuth_rate_deg_s": np.asarray([row[10] for row in rows], dtype=float),
                    "azimuth_accel_deg_s2": np.asarray([row[11] for row in rows], dtype=float),
                    "elevation_deg": np.asarray([row[12] for row in rows], dtype=float),
                    "elevation_rate_deg_s": np.asarray([row[13] for row in rows], dtype=float),
                    "elevation_accel_deg_s2": np.asarray([row[14] for row in rows], dtype=float),
                }
                np.savez_compressed(filename, **payload)
                QMessageBox.information(
                    self, "Export Successful", f"Data exported to:\n{filename}"
                )
                return

            if format_type == "xlsx":
                try:
                    import openpyxl
                    from openpyxl.styles import Font

                    filename = (
                        outputs_dir
                        / f"pass_series_{safe_station}_{pass_stamp}_{timestamp}.xlsx"
                    )
                    wb = openpyxl.Workbook()
                    ws = wb.active
                    ws.title = "Pass Series"

                    for col_idx, header in enumerate(headers, start=1):
                        cell = ws.cell(row=1, column=col_idx, value=header)
                        cell.font = Font(bold=True)

                    for row_idx, row_data in enumerate(rows, start=2):
                        for col_idx, value in enumerate(row_data, start=1):
                            ws.cell(row=row_idx, column=col_idx, value=value)

                    for col in ws.columns:
                        max_length = 0
                        column = col[0].column_letter
                        for cell in col:
                            if cell.value is not None:
                                max_length = max(max_length, len(str(cell.value)))
                        ws.column_dimensions[column].width = min(max_length + 2, 50)

                    wb.save(filename)
                    QMessageBox.information(
                        self, "Export Successful", f"Data exported to:\n{filename}"
                    )
                    return
                except ImportError:
                    QMessageBox.critical(
                        self,
                        "Missing Dependency",
                        "openpyxl is required for XLSX export.\n"
                        "Install it with: pip install openpyxl",
                    )
                    raise

            QMessageBox.warning(
                self,
                "Export Failed",
                f"Unknown export format: {format_type}",
            )
        except Exception as exc:  # pragma: no cover - GUI error handling
            QMessageBox.critical(self, "Export Failed", f"Failed to export data:\n{exc}")
            raise

    @staticmethod
    def _qimage_to_rgba_array(image: QImage) -> np.ndarray:
        """Convert a Qt image into a tightly packed RGBA uint8 array."""
        if image.isNull():
            raise ValueError("Framebuffer capture returned an empty image.")
        converted = image.convertToFormat(QImage.Format.Format_RGBA8888)
        width = int(converted.width())
        height = int(converted.height())
        bytes_per_line = int(converted.bytesPerLine())
        raw = np.frombuffer(converted.constBits(), dtype=np.uint8).copy()
        rgba = raw.reshape((height, bytes_per_line // 4, 4))[:, :width, :]
        return np.ascontiguousarray(rgba)

    @staticmethod
    def _ensure_even_video_frame(frame: np.ndarray) -> np.ndarray:
        """Pad captured frames so H.264/yuv420p receives even dimensions."""
        array = np.asarray(frame, dtype=np.uint8)
        if array.ndim != 3 or array.shape[2] not in {3, 4}:
            raise ValueError("Video export expects frames with shape (H, W, 3|4).")
        pad_h = int(array.shape[0] % 2)
        pad_w = int(array.shape[1] % 2)
        if pad_h == 0 and pad_w == 0:
            return np.ascontiguousarray(array)
        return np.pad(
            array,
            ((0, pad_h), (0, pad_w), (0, 0)),
            mode="edge",
        )

    def _export_visualization_video(self) -> None:
        """Render the selected pass to an MP4 using the current viewer state."""
        pass_stat = getattr(self, "_visual_selected_pass", None)
        track = getattr(self, "_visual_pass_track", None)
        widget = getattr(self, "visual_globe_widget", None)
        if pass_stat is None or not track or widget is None:
            QMessageBox.warning(
                self,
                "No Pass Selected",
                "Select a pass in the Visualization tab before exporting a video.",
            )
            return
        if widget.width() <= 0 or widget.height() <= 0:
            raise ValueError("Visualization widget has no drawable size for video export.")

        try:
            import imageio.v2 as imageio
        except ImportError:
            QMessageBox.critical(
                self,
                "Missing Dependency",
                "imageio and imageio-ffmpeg are required for MP4 export.\n"
                "Install them with: pip install imageio imageio-ffmpeg",
            )
            return

        station_name = pass_stat.station_name or "GroundStation"
        safe_station = re.sub(r"[^A-Za-z0-9_-]+", "_", str(station_name)).strip("_")
        pass_stamp = pass_stat.aos.strftime("%Y%m%d_%H%M%S")
        outputs_dir = Path(__file__).resolve().parents[3] / "outputs"
        outputs_dir.mkdir(exist_ok=True)
        default_path = outputs_dir / f"pass_render_{safe_station}_{pass_stamp}.mp4"
        selected_path, _selected_filter = QFileDialog.getSaveFileName(
            self,
            "Export Pass Video",
            str(default_path),
            "MP4 Video (*.mp4)",
        )
        if not selected_path:
            return
        output_path = Path(selected_path)
        if output_path.suffix.lower() != ".mp4":
            output_path = output_path.with_suffix(".mp4")

        saved_index = int(self._visual_animation_index)
        saved_fraction = float(self._visual_anim_fraction)
        was_animating = bool(self._visual_animation_timer.isActive())
        saved_button_text = self.visual_play_button.text() if self.visual_play_button else "Play"
        if was_animating:
            self._visual_animation_timer.stop()
        if self.visual_play_button:
            self.visual_play_button.setText("Play")

        total_frames = len(track)
        fps = 30
        frames_written = 0
        try:
            with imageio.get_writer(
                str(output_path),
                fps=fps,
                codec="libx264",
                macro_block_size=1,
            ) as writer:
                for frame_idx in range(total_frames):
                    self._visual_animation_index = frame_idx
                    self._visual_anim_fraction = 0.0
                    self._update_visualization_frame(render=False)
                    QApplication.processEvents()
                    rgba = self._qimage_to_rgba_array(widget.grabFramebuffer())
                    rgba = self._ensure_even_video_frame(rgba)
                    writer.append_data(rgba[:, :, :3])
                    frames_written += 1
                    if self.visual_pass_status_label is not None:
                        self.visual_pass_status_label.setText(
                            f"Exporting video {frame_idx + 1}/{total_frames}: {output_path.name}"
                        )
                        QApplication.processEvents()
            if frames_written <= 0:
                raise RuntimeError("No video frames were written.")
            if not output_path.exists():
                raise RuntimeError("FFmpeg did not create the output file.")
            if output_path.stat().st_size <= 0:
                output_path.unlink(missing_ok=True)
                raise RuntimeError("FFmpeg created a zero-byte output file.")
        except Exception as exc:  # pragma: no cover - GUI error handling
            QMessageBox.critical(self, "Export Failed", f"Failed to export video:\n{exc}")
            return
        finally:
            self._visual_animation_index = saved_index
            self._visual_anim_fraction = saved_fraction
            self._update_visualization_frame()
            if was_animating:
                self._visual_animation_timer.start()
            if self.visual_play_button:
                self.visual_play_button.setText(saved_button_text)

        QMessageBox.information(
            self,
            "Export Successful",
            f"Pass video exported to:\n{output_path}",
        )

    def _toggle_visualization_playback(self) -> None:
        """Play or pause the current pass animation."""
        if not self._visual_pass_track or self.visual_play_button is None:
            return
        if self._visual_animation_timer.isActive():
            self._visual_animation_timer.stop()
            self.visual_play_button.setText("Play")
            return
        if self._visual_animation_index >= len(self._visual_pass_track) - 1:
            self._visual_animation_index = 0
            self._visual_anim_fraction = 0.0
            self._update_visualization_frame(render=False)
        else:
            self._visual_anim_fraction = 0.0
        self._visual_animation_timer.start()
        self.visual_play_button.setText("Pause")

    def _advance_visualization_animation(self) -> None:
        """Advance the satellite along the pass during playback."""
        if not self._visual_pass_track:
            self._stop_visualization_animation()
            return
        if self._visual_animation_index >= len(self._visual_pass_track) - 1:
            self._stop_visualization_animation(final=True)
            return
        step = self._visual_base_step * self._visual_speed_multiplier
        self._visual_anim_fraction += step
        while (
            self._visual_anim_fraction >= 1.0
            and self._visual_animation_index < len(self._visual_pass_track) - 1
        ):
            self._visual_anim_fraction -= 1.0
            self._visual_animation_index += 1
        if self._visual_animation_index >= len(self._visual_pass_track) - 1:
            self._visual_animation_index = len(self._visual_pass_track) - 1
            self._visual_anim_fraction = 0.0
            self._update_visualization_frame()
            self._stop_visualization_animation(final=True)
            return
        self._update_visualization_frame()

    def _stop_visualization_animation(self, final: bool = False) -> None:
        """Stop playback and update control labels."""
        if self._visual_animation_timer.isActive():
            self._visual_animation_timer.stop()
        if self.visual_play_button:
            self.visual_play_button.setText("Replay" if final else "Play")

    def _handle_visualization_slider_moved(self, value: int) -> None:
        """Scrub through the pass timeline."""
        if not self._visual_pass_track:
            return
        clamped = max(0, min(value, len(self._visual_pass_track) - 1))
        self._visual_animation_index = clamped
        self._visual_anim_fraction = 0.0
        if self._visual_animation_timer.isActive():
            self._visual_animation_timer.stop()
            if self.visual_play_button:
                self.visual_play_button.setText("Play")
        self._update_visualization_frame()

    def _handle_visual_speed_changed(self, value: int) -> None:
        """Adjust playback speed multiplier."""
        self._visual_speed_multiplier = max(0.1, value / 100.0)

    def _update_visualization_frame(self, render: bool = True) -> None:
        """Render the satellite position, slider, and ground link for the current frame."""
        if (
            not self._visual_pass_track
            or self.visual_globe_widget is None
            or self._visual_animation_index >= len(self._visual_pass_track)
        ):
            return
        point = self._visual_pass_track[self._visual_animation_index]
        base_vec = np.array(
            self._convert_vector_to_current_frame(
                ecef_to_globe_coords(point.x_km, point.y_km, point.z_km),
                point.timestamp,
            ),
            dtype=float,
        )
        timestamp = point.timestamp
        coords_vec = base_vec
        # Approximate velocity direction for body-frame triad
        vel_vec = None
        if (
            self._visual_anim_fraction > 1e-4
            and self._visual_animation_index < len(self._visual_pass_track) - 1
        ):
            next_point = self._visual_pass_track[self._visual_animation_index + 1]
            next_vec = np.array(
                self._convert_vector_to_current_frame(
                    ecef_to_globe_coords(next_point.x_km, next_point.y_km, next_point.z_km),
                    next_point.timestamp,
                ),
                dtype=float,
            )
            vel_vec = next_vec - base_vec
            alpha = self._visual_anim_fraction
            coords_vec = (1.0 - alpha) * base_vec + alpha * next_vec
            dt = next_point.timestamp - point.timestamp
            timestamp = point.timestamp + alpha * dt
        else:
            # Use forward or backward difference at sample points
            if self._visual_animation_index < len(self._visual_pass_track) - 1:
                next_point = self._visual_pass_track[self._visual_animation_index + 1]
                next_vec = np.array(
                    self._convert_vector_to_current_frame(
                        ecef_to_globe_coords(next_point.x_km, next_point.y_km, next_point.z_km),
                        next_point.timestamp,
                    ),
                    dtype=float,
                )
                vel_vec = next_vec - base_vec
            elif self._visual_animation_index > 0:
                prev_point = self._visual_pass_track[self._visual_animation_index - 1]
                prev_vec = np.array(
                    self._convert_vector_to_current_frame(
                        ecef_to_globe_coords(prev_point.x_km, prev_point.y_km, prev_point.z_km),
                        prev_point.timestamp,
                    ),
                    dtype=float,
                )
                vel_vec = base_vec - prev_vec
        coords = tuple(coords_vec.tolist())
        self._update_visual_earth_rotation(timestamp, request_repaint=False)
        self._update_visualization_focus_point(timestamp, request_repaint=False)
        if self.visual_globe_widget:
            self.visual_globe_widget.set_sun_datetime(timestamp, request_repaint=False)
        axes = self._resolve_visualization_body_axes(
            timestamp=timestamp,
            sat_pos=coords_vec,
            sat_vel=vel_vec,
        )
        self.visual_globe_widget.update_satellite_position(
            coords,
            body_axes=axes,
            request_repaint=False,
        )
        self._update_visualization_sensor_outline(
            timestamp,
            coords_vec,
            axes,
            request_repaint=False,
        )
        if self.visual_time_slider:
            self.visual_time_slider.blockSignals(True)
            self.visual_time_slider.setValue(self._visual_animation_index)
            self.visual_time_slider.blockSignals(False)
        if self.visual_pass_status_label and self._visual_selected_pass is not None:
            station_name = self._visual_selected_pass.station_name or "Ground Station"
            self.visual_pass_status_label.setText(
                f"{self._visual_frame_mode} | {station_name}: {timestamp:%d-%b %H:%M:%S} UTC"
            )
        self._update_visualization_link_actor(
            timestamp,
            coords,
            request_repaint=False,
        )
        if render:
            self.visual_globe_widget.update()

    def _get_station_position(
        self, timestamp: datetime
    ) -> tuple[float, float, float] | None:
        """Return the station position vector in the current frame, converted to globe coordinates."""
        if self._visual_station_ecef is None:
            return None
        # First align with the globe texture coordinate system (ECEF → globe)
        globe_vec = np.array(ecef_to_globe_coords(*self._visual_station_ecef), dtype=float)
        # Nudge outward slightly so the marker/link sit above the surface for visibility
        length = np.linalg.norm(globe_vec)
        if length > 1e-6 and STATION_VISUAL_OFFSET_KM > 0.0:
            scale = (length + STATION_VISUAL_OFFSET_KM) / length
            globe_vec *= scale
        # Then rotate into the current reference frame (ECI adds the time-varying spin)
        return self._convert_vector_to_current_frame(tuple(globe_vec.tolist()), timestamp)

    def _update_visualization_focus_point(
        self, timestamp: datetime | None, *, request_repaint: bool = True
    ) -> None:
        """Keep the camera focus aligned with the active ground station."""
        widget = getattr(self, "visual_globe_widget", None)
        if widget is None:
            return
        if timestamp is None:
            widget.set_focus_point(None, request_repaint=request_repaint)
            return
        station_point = self._get_station_position(timestamp)
        widget.set_focus_point(
            station_point if station_point is not None else None,
            request_repaint=request_repaint,
        )

    def _update_visual_earth_rotation(
        self, timestamp: datetime | None, *, request_repaint: bool = True
    ) -> None:
        """Rotate the globe actors to match Earth orientation in the selected frame."""
        # When rendering in ECI mode, rotate Earth geometry from ECEF->ECI using GMST.
        # (In ECEF mode, Earth mesh remains fixed and lighting uses ECEF sun direction.)
        if timestamp is None:
            angle_deg = 0.0
        else:
            angle_deg = math.degrees(gmst_rad_utc(timestamp))
        widget = getattr(self, "visual_globe_widget", None)
        if widget is None:
            return
        widget.set_frame_rotation(
            self._visual_frame_mode,
            angle_deg,
            request_repaint=request_repaint,
        )

    def _update_visualization_link_actor(
        self,
        timestamp: datetime,
        satellite_coords: tuple[float, float, float],
        *,
        request_repaint: bool = True,
    ) -> None:
        """Show or hide the green contact link for the current frame."""
        widget = getattr(self, "visual_globe_widget", None)
        if widget is None:
            return
        if self._visual_contact_window is None or self._visual_station_ecef is None:
            widget.update_link_segment(None, None, request_repaint=request_repaint)
            return
        aos, los = self._visual_contact_window
        if not (aos <= timestamp <= los):
            widget.update_link_segment(None, None, request_repaint=request_repaint)
            return
        station_point = self._get_station_position(timestamp)
        if station_point is None:
            return
        widget.update_link_segment(
            station_point,
            satellite_coords,
            request_repaint=request_repaint,
        )

    def _update_visualization_sensor_outline(
        self,
        timestamp: datetime | None,
        sat_coords: np.ndarray,
        body_axes: np.ndarray | None,
        *,
        request_repaint: bool = True,
    ) -> None:
        widget = getattr(self, "visual_globe_widget", None)
        if widget is None:
            return
        checkbox = getattr(self, "visual_antenna_lobe_checkbox", None)
        if (
            timestamp is None
            or body_axes is None
            or checkbox is None
            or not checkbox.isChecked()
        ):
            widget.update_sensor_cone_outline(None, request_repaint=request_repaint)
            return
        coords, colors = self._build_visual_antenna_lobe_coords(
            sat_coords=sat_coords,
            body_axes=body_axes,
        )
        widget.update_sensor_cone_outline(
            coords,
            colors_rgba=colors,
            request_repaint=request_repaint,
        )


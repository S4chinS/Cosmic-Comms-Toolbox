"""Plotting helpers shared across the UI tabs."""

from __future__ import annotations

from datetime import timedelta

import numpy as np
import pyqtgraph as pg
from matplotlib.backends.backend_qtagg import FigureCanvasQTAgg
from matplotlib.figure import Figure
from PySide6.QtCore import Qt
from PySide6.QtGui import QColor
from PySide6.QtWidgets import (
    QFileDialog,
    QFormLayout,
    QGroupBox,
    QHBoxLayout,
    QLabel,
    QListWidget,
    QListWidgetItem,
    QMessageBox,
    QPushButton,
    QSplitter,
    QVBoxLayout,
    QWidget,
)

from cosmic_toolbox.models import PassStatistic, StationSummary
from cosmic_toolbox_gui.constants import STATION_COLOR_PALETTE


class PlotHelpersMixin:
    """Reusable routines for timeline, histogram, and pie-chart plotting."""

    def _create_contact_plot_tabs(self) -> list[tuple[str, QWidget]]:
        """Initialize contact plots and return tab definitions."""
        pg.setConfigOptions(antialias=True)
        self.timeline_plot = pg.PlotWidget(title="Access Timeline")
        self.timeline_plot.setLabel("bottom", "Time since start", units="h")
        self.timeline_plot.hideAxis("left")
        self.timeline_plot.showGrid(x=True, y=False, alpha=0.3)
        self.timeline_plot.setMouseEnabled(x=False, y=False)
        self.timeline_plot.setMenuEnabled(False)
        self.timeline_legend = self.timeline_plot.addLegend()

        # Hover readout for cursor time.
        self._timeline_hover_label = pg.LabelItem(justify="left")
        self.timeline_plot.plotItem.layout.addItem(self._timeline_hover_label, 0, 0)
        self._timeline_hover_label.setVisible(False)
        self._timeline_hover_line = pg.InfiniteLine(
            angle=90, movable=False, pen=pg.mkPen("#d0d0d0", width=1)
        )
        self._timeline_hover_line.setVisible(False)
        self.timeline_plot.addItem(self._timeline_hover_line)
        self.timeline_plot.scene().sigMouseMoved.connect(self._handle_timeline_hover)  # type: ignore[arg-type]

        pie_figure = Figure(figsize=(4, 3), tight_layout=True)
        pie_figure.patch.set_facecolor("#0b0b10")
        self.pie_canvas = FigureCanvasQTAgg(pie_figure)
        self.pie_axes = pie_figure.add_subplot(111)
        self._configure_pie_axes()
        timeline_tab = QWidget()
        timeline_layout = QVBoxLayout(timeline_tab)
        timeline_layout.setContentsMargins(4, 4, 4, 4)

        # Timeline tab now includes:
        # - Left (75%): access timeline plot
        # - Right (25%): station enable/disable + gap statistics
        timeline_splitter = QSplitter(Qt.Orientation.Horizontal)
        timeline_layout.addWidget(timeline_splitter, stretch=1)

        timeline_left = QWidget()
        timeline_left.setMinimumWidth(0)
        timeline_left_layout = QVBoxLayout(timeline_left)
        timeline_left_layout.setContentsMargins(0, 0, 0, 0)
        timeline_left_layout.addWidget(self.timeline_plot, stretch=1)
        timeline_splitter.addWidget(timeline_left)

        timeline_right = QWidget()
        timeline_right.setMinimumWidth(0)
        timeline_right_layout = QVBoxLayout(timeline_right)
        timeline_right_layout.setContentsMargins(0, 0, 0, 0)

        stations_group = QGroupBox("Stations (toggle to update)")
        stations_layout = QVBoxLayout(stations_group)
        timeline_button_row = QHBoxLayout()
        self.timeline_select_all_button = QPushButton("Select all")
        self.timeline_select_all_button.clicked.connect(
            lambda: self._set_station_list_checks(self.timeline_station_list, True)
        )  # type: ignore[arg-type]
        timeline_button_row.addWidget(self.timeline_select_all_button)
        self.timeline_clear_selection_button = QPushButton("Clear selection")
        self.timeline_clear_selection_button.clicked.connect(
            lambda: self._set_station_list_checks(self.timeline_station_list, False)
        )  # type: ignore[arg-type]
        timeline_button_row.addWidget(self.timeline_clear_selection_button)
        stations_layout.addLayout(timeline_button_row)
        self.timeline_station_list = QListWidget()
        self.timeline_station_list.itemChanged.connect(  # type: ignore[arg-type]
            self._handle_timeline_station_toggle
        )
        stations_layout.addWidget(self.timeline_station_list, stretch=1)
        timeline_right_layout.addWidget(stations_group, stretch=2)

        stats_group = QGroupBox("Gap statistics (selected stations)")
        stats_layout = QFormLayout(stats_group)
        self._timeline_gap_stat_labels: dict[str, QLabel] = {}
        for key in (
            "Selected stations",
            "Scenario duration",
            "Total access",
            "Coverage",
            "Num gaps",
            "Mean gap",
            "Median gap",
            "P95 gap",
            "Longest no access",
        ):
            label = QLabel("—")
            label.setTextInteractionFlags(Qt.TextInteractionFlag.TextSelectableByMouse)
            self._timeline_gap_stat_labels[key] = label
            stats_layout.addRow(f"{key}:", label)
        timeline_right_layout.addWidget(stats_group, stretch=1)

        timeline_splitter.addWidget(timeline_right)
        timeline_splitter.setChildrenCollapsible(True)
        timeline_splitter.setStretchFactor(0, 3)
        timeline_splitter.setStretchFactor(1, 1)
        timeline_splitter.setSizes([3, 1])

        pie_tab = QWidget()
        pie_layout = QVBoxLayout(pie_tab)
        pie_layout.setContentsMargins(4, 4, 4, 4)
        pie_layout.addWidget(self.pie_canvas)
        elevation_tab = QWidget()
        elevation_layout = QVBoxLayout(elevation_tab)
        elevation_layout.setContentsMargins(4, 4, 4, 4)
        elevation_splitter = QSplitter(Qt.Orientation.Horizontal)
        elevation_layout.addWidget(elevation_splitter, stretch=1)
        elevation_button_row = QHBoxLayout()
        elevation_button_row.addStretch(1)
        self.elevation_export_npz_button = QPushButton("Export NPZ")
        self.elevation_export_npz_button.clicked.connect(  # type: ignore[arg-type]
            lambda: self._export_elevation_distribution("npz")
        )
        elevation_button_row.addWidget(self.elevation_export_npz_button)
        self.elevation_export_xlsx_button = QPushButton("Export XLSX")
        self.elevation_export_xlsx_button.clicked.connect(  # type: ignore[arg-type]
            lambda: self._export_elevation_distribution("xlsx")
        )
        elevation_button_row.addWidget(self.elevation_export_xlsx_button)
        elevation_layout.addLayout(elevation_button_row)
        elevation_left = QWidget()
        elevation_left.setMinimumWidth(0)
        elevation_left_layout = QVBoxLayout(elevation_left)
        elevation_left_layout.setContentsMargins(0, 0, 0, 0)
        elevation_figure = Figure(figsize=(5, 5), tight_layout=True)
        elevation_figure.patch.set_facecolor("#0b0b10")
        self.elevation_canvas = FigureCanvasQTAgg(elevation_figure)
        self.elevation_pdf_axes = elevation_figure.add_subplot(211)
        self.elevation_cdf_axes = elevation_figure.add_subplot(212, sharex=self.elevation_pdf_axes)
        self._configure_elevation_axes()
        elevation_left_layout.addWidget(self.elevation_canvas, stretch=1)
        elevation_splitter.addWidget(elevation_left)
        elevation_right = QWidget()
        elevation_right.setMinimumWidth(0)
        elevation_right_layout = QVBoxLayout(elevation_right)
        elevation_right_layout.setContentsMargins(0, 0, 0, 0)
        elevation_group = QGroupBox("Stations (toggle to update)")
        elevation_group_layout = QVBoxLayout(elevation_group)
        elevation_button_row = QHBoxLayout()
        self.elevation_select_all_button = QPushButton("Select all")
        self.elevation_select_all_button.clicked.connect(
            lambda: self._set_station_list_checks(self.elevation_station_list, True)
        )  # type: ignore[arg-type]
        elevation_button_row.addWidget(self.elevation_select_all_button)
        self.elevation_clear_selection_button = QPushButton("Clear selection")
        self.elevation_clear_selection_button.clicked.connect(
            lambda: self._set_station_list_checks(self.elevation_station_list, False)
        )  # type: ignore[arg-type]
        elevation_button_row.addWidget(self.elevation_clear_selection_button)
        elevation_group_layout.addLayout(elevation_button_row)
        self.elevation_station_list = QListWidget()
        self.elevation_station_list.itemChanged.connect(  # type: ignore[arg-type]
            self._handle_elevation_station_toggle
        )
        elevation_group_layout.addWidget(self.elevation_station_list, stretch=1)
        elevation_right_layout.addWidget(elevation_group, stretch=1)
        elevation_splitter.addWidget(elevation_right)
        elevation_splitter.setChildrenCollapsible(True)
        elevation_splitter.setStretchFactor(0, 3)
        elevation_splitter.setStretchFactor(1, 1)
        elevation_splitter.setSizes([3, 1])
        self._draw_empty_plots()
        self._set_elevation_export_enabled(False)
        return [
            ("Timeline", timeline_tab),
            ("Access Share", pie_tab),
            ("Elevation Distribution", elevation_tab),
        ]

    def _handle_timeline_station_toggle(self, _item: QListWidgetItem) -> None:
        """Re-render the timeline + gap stats when station toggles change."""
        result = self._get_contact_statistics_result()
        if result is None:
            return
        self._update_timeline_view(result)

    def _get_enabled_timeline_stations(self) -> set[str]:
        """Return the set of station names enabled in the timeline station list."""
        widget = getattr(self, "timeline_station_list", None)
        if widget is None:
            return set()
        enabled: set[str] = set()
        for idx in range(widget.count()):
            item = widget.item(idx)
            if item is None:
                continue
            if item.checkState() == Qt.CheckState.Checked:
                enabled.add(item.text())
        return enabled

    def _set_timeline_station_list(self, station_names: list[str]) -> None:
        """Populate the timeline station selector, defaulting all to enabled."""
        widget = getattr(self, "timeline_station_list", None)
        if widget is None:
            return
        widget.blockSignals(True)
        widget.clear()
        for name in station_names:
            item = QListWidgetItem(name)
            item.setFlags(item.flags() | Qt.ItemFlag.ItemIsUserCheckable)
            item.setCheckState(Qt.CheckState.Checked)
            widget.addItem(item)
        widget.blockSignals(False)

    def _set_elevation_export_enabled(self, enabled: bool) -> None:
        for button_name in ("elevation_export_npz_button", "elevation_export_xlsx_button"):
            button = getattr(self, button_name, None)
            if button is not None:
                button.setEnabled(bool(enabled))

    def _handle_elevation_station_toggle(self, _item: QListWidgetItem) -> None:
        """Re-render the elevation PDF/CDF when station toggles change."""
        result = self._get_contact_statistics_result()
        if result is None:
            return
        self._update_elevation_distribution(result)

    def _get_enabled_elevation_stations(self) -> set[str]:
        """Return the set of station names enabled in the elevation station list."""
        widget = getattr(self, "elevation_station_list", None)
        if widget is None:
            return set()
        enabled: set[str] = set()
        for idx in range(widget.count()):
            item = widget.item(idx)
            if item is None:
                continue
            if item.checkState() == Qt.CheckState.Checked:
                enabled.add(item.text())
        return enabled

    def _set_elevation_station_list(self, station_names: list[str]) -> None:
        """Populate the elevation selector, defaulting all to enabled."""
        widget = getattr(self, "elevation_station_list", None)
        if widget is None:
            return
        widget.blockSignals(True)
        widget.clear()
        for name in station_names:
            item = QListWidgetItem(name)
            item.setFlags(item.flags() | Qt.ItemFlag.ItemIsUserCheckable)
            item.setCheckState(Qt.CheckState.Checked)
            widget.addItem(item)
        widget.blockSignals(False)

    def _set_station_list_checks(self, widget: QListWidget | None, checked: bool) -> None:
        """Select or clear all station rows in a checkbox list."""
        if widget is None:
            return
        widget.blockSignals(True)
        state = Qt.CheckState.Checked if checked else Qt.CheckState.Unchecked
        for idx in range(widget.count()):
            item = widget.item(idx)
            if item is not None:
                item.setCheckState(state)
        widget.blockSignals(False)
        result = self._get_contact_statistics_result()
        if result is None:
            return
        if widget is getattr(self, "timeline_station_list", None):
            self._update_timeline_view(result)
            return
        if widget is getattr(self, "elevation_station_list", None):
            self._update_elevation_distribution(result)

    def _configure_elevation_axes(self) -> None:
        """Initialize PDF/CDF axes styling for the elevation distribution."""
        if not getattr(self, "elevation_pdf_axes", None) or not getattr(self, "elevation_cdf_axes", None):
            return
        for ax in (self.elevation_pdf_axes, self.elevation_cdf_axes):
            ax.clear()
            ax.set_facecolor("#0b0b10")
            ax.tick_params(axis="both", colors="#dddddd")
            ax.grid(True, color="#3a3a3a", alpha=0.4, linestyle="--")
            for spine in ax.spines.values():
                spine.set_color("#444444")
        self.elevation_pdf_axes.set_title("Elevation PDF", color="white", pad=8)
        self.elevation_pdf_axes.set_ylabel("PDF", color="white")
        self.elevation_cdf_axes.set_title("Elevation CDF", color="white", pad=8)
        self.elevation_cdf_axes.set_xlabel("Elevation (deg)", color="white")
        self.elevation_cdf_axes.set_ylabel("CDF", color="white")

    def _clear_elevation_distribution(self, title: str) -> None:
        """Reset elevation plots with a placeholder title."""
        if not getattr(self, "elevation_pdf_axes", None) or not getattr(self, "elevation_canvas", None):
            return
        self._configure_elevation_axes()
        self.elevation_pdf_axes.set_title(title, color="white", pad=8)
        self.elevation_cdf_axes.set_title("", color="white", pad=8)
        self.elevation_canvas.draw_idle()

    def _collect_elevation_samples(
        self,
        *,
        result,
        station_name: str,
        scenario_start,
    ) -> np.ndarray:
        """Collect elevation samples during pass intervals for a station."""
        timeline = np.asarray(getattr(result, "timeline_seconds", []), dtype=float)
        if timeline.size == 0:
            return np.asarray([], dtype=float)
        station_series = getattr(result, "station_elevation_series", {}) or {}
        series = station_series.get(station_name)
        if series is None:
            return np.asarray([], dtype=float)
        elev = np.asarray(series, dtype=float)
        if elev.size != timeline.size:
            return np.asarray([], dtype=float)
        mask = np.zeros_like(timeline, dtype=bool)
        passes = getattr(result, "passes", []) or []
        for p in passes:
            name = p.station_name or "Ground Station"
            if name != station_name:
                continue
            aos = getattr(p, "aos", None)
            los = getattr(p, "los", None)
            if aos is None or los is None:
                continue
            start = (aos - scenario_start).total_seconds()
            end = (los - scenario_start).total_seconds()
            if not np.isfinite(start) or not np.isfinite(end) or end <= start:
                continue
            mask |= (timeline >= start) & (timeline <= end)
        if not np.any(mask):
            return np.asarray([], dtype=float)
        samples = elev[mask]
        samples = samples[np.isfinite(samples)]
        samples = samples[samples >= 0.0]
        return samples

    def _update_elevation_distribution(self, result) -> None:
        """Render PDF/CDF of elevation angles for selected stations."""
        if getattr(self, "_current_config", None) is None:
            return
        if not result or not getattr(result, "passes", None):
            self._clear_elevation_distribution("Elevation Distribution\n(Run analysis to view)")
            self._set_elevation_export_enabled(False)
            return
        scenario_start = self._current_config.scenario.start_time
        enabled = self._get_enabled_elevation_stations()
        if not enabled:
            self._clear_elevation_distribution("Elevation Distribution\n(No stations selected)")
            self._set_elevation_export_enabled(False)
            return
        samples_list = [
            self._collect_elevation_samples(
                result=result, station_name=name, scenario_start=scenario_start
            )
            for name in sorted(enabled)
        ]
        samples = np.concatenate([arr for arr in samples_list if arr.size]) if samples_list else np.asarray([], dtype=float)
        if samples.size == 0:
            self._clear_elevation_distribution("Elevation Distribution\n(No pass samples)")
            self._set_elevation_export_enabled(False)
            return
        q25, q75 = np.percentile(samples, [25, 75])
        iqr = float(q75 - q25)
        if not (np.isfinite(iqr) and iqr > 0.0):
            raise ValueError("Elevation PDF requires a non-zero IQR for adaptive bin width.")
        bin_width = 2.0 * iqr * (samples.size ** (-1.0 / 3.0))
        if not (np.isfinite(bin_width) and bin_width > 0.0):
            raise ValueError("Elevation PDF bin width must be finite and positive.")
        self._configure_elevation_axes()
        bins = np.arange(0.0, 90.0 + bin_width, bin_width)
        if bins.size < 2:
            raise ValueError("Elevation PDF requires at least two histogram bins.")
        counts, edges = np.histogram(samples, bins=bins, density=True)
        centers = 0.5 * (edges[:-1] + edges[1:])
        self.elevation_pdf_axes.plot(
            centers,
            counts,
            color="#4fc3f7",
            linewidth=2.0,
        )
        sorted_samples = np.sort(samples)
        cdf = np.linspace(0.0, 1.0, sorted_samples.size, endpoint=True)
        self.elevation_cdf_axes.plot(
            sorted_samples,
            cdf,
            color="#81c784",
            linewidth=2.0,
        )
        self.elevation_canvas.draw_idle()
        self._set_elevation_export_enabled(True)

    def _export_elevation_distribution(self, format_type: str) -> None:
        """Export elevation distribution samples + PDF/CDF to NPZ or XLSX."""
        result = self._get_contact_statistics_result()
        if result is None or not getattr(result, "passes", None):
            QMessageBox.warning(self, "Export Failed", "Run an analysis first.")
            return
        if getattr(self, "_current_config", None) is None:
            QMessageBox.warning(self, "Export Failed", "Missing scenario configuration.")
            return
        enabled = sorted(self._get_enabled_elevation_stations())
        if not enabled:
            QMessageBox.warning(self, "Export Failed", "Select at least one station.")
            return
        scenario_start = self._current_config.scenario.start_time
        per_station_samples: dict[str, np.ndarray] = {}
        for name in enabled:
            samples = self._collect_elevation_samples(
                result=result, station_name=name, scenario_start=scenario_start
            )
            per_station_samples[name] = samples
        combined = np.concatenate(
            [arr for arr in per_station_samples.values() if arr.size]
        )
        if combined.size == 0:
            QMessageBox.warning(self, "Export Failed", "No elevation samples available.")
            return
        q25, q75 = np.percentile(combined, [25, 75])
        iqr = float(q75 - q25)
        if not (np.isfinite(iqr) and iqr > 0.0):
            raise ValueError("Elevation export requires a non-zero IQR for adaptive bin width.")
        bin_width = 2.0 * iqr * (combined.size ** (-1.0 / 3.0))
        if not (np.isfinite(bin_width) and bin_width > 0.0):
            raise ValueError("Elevation export bin width must be finite and positive.")
        bins = np.arange(0.0, 90.0 + bin_width, bin_width)
        if bins.size < 2:
            raise ValueError("Elevation export requires at least two histogram bins.")
        pdf_counts, pdf_edges = np.histogram(combined, bins=bins, density=True)
        pdf_centers = 0.5 * (pdf_edges[:-1] + pdf_edges[1:])
        sorted_samples = np.sort(combined)
        cdf_values = np.linspace(0.0, 1.0, sorted_samples.size, endpoint=True)

        if format_type == "npz":
            default_path = getattr(self, "_last_output_dir", None) or "."
            filename, _ = QFileDialog.getSaveFileName(
                self,
                "Export Elevation Distribution NPZ",
                str(default_path),
                "NPZ files (*.npz);;All files (*.*)",
            )
            if not filename:
                return
            payload: dict[str, np.ndarray] = {
                "pdf_centers_deg": np.asarray(pdf_centers, dtype=float),
                "pdf_density": np.asarray(pdf_counts, dtype=float),
                "cdf_elevation_deg": np.asarray(sorted_samples, dtype=float),
                "cdf_values": np.asarray(cdf_values, dtype=float),
            }
            for name, samples in per_station_samples.items():
                key = "".join(
                    ch if (ch.isalnum() or ch in "_-") else "_" for ch in str(name)
                )
                if not key:
                    raise ValueError("Station name must contain at least one alphanumeric character.")
                payload[f"samples_{key}"] = np.asarray(samples, dtype=float)
            np.savez_compressed(filename, **payload)
            QMessageBox.information(
                self, "Export Successful", f"Elevation distribution exported to:\n{filename}"
            )
            return

        if format_type == "xlsx":
            default_path = getattr(self, "_last_output_dir", None) or "."
            filename, _ = QFileDialog.getSaveFileName(
                self,
                "Export Elevation Distribution XLSX",
                str(default_path),
                "XLSX files (*.xlsx);;All files (*.*)",
            )
            if not filename:
                return
            try:
                import openpyxl  # type: ignore[import-not-found]
            except Exception:
                QMessageBox.warning(
                    self,
                    "Missing Dependency",
                    "openpyxl is required for XLSX export.\n"
                    "Install it with: pip install openpyxl",
                )
                return
            wb = openpyxl.Workbook()
            summary_ws = wb.active
            summary_ws.title = "Elevation PDF"
            summary_ws.append(["Elevation (deg)", "PDF"])
            for x, y in zip(pdf_centers, pdf_counts):
                summary_ws.append([float(x), float(y)])
            cdf_ws = wb.create_sheet("Elevation CDF")
            cdf_ws.append(["Elevation (deg)", "CDF"])
            for x, y in zip(sorted_samples, cdf_values):
                cdf_ws.append([float(x), float(y)])
            samples_ws = wb.create_sheet("Samples")
            samples_ws.append(["Station", "Elevation (deg)"])
            for name, samples in per_station_samples.items():
                for val in samples:
                    samples_ws.append([name, float(val)])
            wb.save(filename)
            QMessageBox.information(
                self, "Export Successful", f"Elevation distribution exported to:\n{filename}"
            )
            return

        QMessageBox.warning(
            self,
            "Export Failed",
            f"Unknown export format: {format_type}",
        )

    def _format_duration_minutes(self, minutes: float) -> str:
        if not np.isfinite(minutes):
            return "—"
        if minutes < 60.0:
            return f"{minutes:.2f} min"
        hours = minutes / 60.0
        if hours < 48.0:
            return f"{hours:.2f} h"
        days = hours / 24.0
        return f"{days:.2f} d"

    def _compute_union_intervals_seconds(
        self,
        passes: list[PassStatistic],
        *,
        scenario_start,
        scenario_end,
    ) -> list[tuple[float, float]]:
        """Return merged access intervals (seconds from scenario start)."""
        if not passes:
            return []
        intervals: list[tuple[float, float]] = []
        for p in passes:
            aos = getattr(p, "aos", None)
            los = getattr(p, "los", None)
            if aos is None or los is None:
                continue
            start = (aos - scenario_start).total_seconds()
            end = (los - scenario_start).total_seconds()
            if not np.isfinite(start) or not np.isfinite(end):
                continue
            if end <= start:
                continue
            # Clip to scenario window
            start = max(0.0, float(start))
            end = min((scenario_end - scenario_start).total_seconds(), float(end))
            if end <= start:
                continue
            intervals.append((start, end))
        if not intervals:
            return []
        intervals.sort(key=lambda t: t[0])
        merged: list[tuple[float, float]] = []
        cur_s, cur_e = intervals[0]
        for s, e in intervals[1:]:
            if s <= cur_e:
                cur_e = max(cur_e, e)
            else:
                merged.append((cur_s, cur_e))
                cur_s, cur_e = s, e
        merged.append((cur_s, cur_e))
        return merged

    def _compute_gap_durations_seconds(
        self,
        merged_intervals: list[tuple[float, float]],
        *,
        scenario_duration_s: float,
    ) -> list[float]:
        """Return gaps (seconds) covering the whole scenario window."""
        if scenario_duration_s <= 0:
            return []
        gaps: list[float] = []
        if not merged_intervals:
            return [float(scenario_duration_s)]
        # Leading gap
        first_start = float(merged_intervals[0][0])
        if first_start > 0:
            gaps.append(first_start)
        # In-between
        for (s0, e0), (s1, _e1) in zip(merged_intervals, merged_intervals[1:]):
            gap = float(s1 - e0)
            if gap > 0:
                gaps.append(gap)
        # Trailing
        last_end = float(merged_intervals[-1][1])
        tail = float(scenario_duration_s - last_end)
        if tail > 0:
            gaps.append(tail)
        return gaps

    def _update_timeline_gap_stats(
        self,
        *,
        enabled_station_names: list[str],
        merged_intervals: list[tuple[float, float]],
        scenario_duration_s: float,
    ) -> None:
        labels = getattr(self, "_timeline_gap_stat_labels", None)
        if not isinstance(labels, dict):
            return

        def _set(key: str, value: str) -> None:
            label = labels.get(key)
            if label is not None:
                label.setText(value)

        if scenario_duration_s <= 0:
            for key in labels:
                _set(key, "—")
            return

        total_access_s = float(sum(e - s for s, e in merged_intervals))
        coverage = 100.0 * total_access_s / float(scenario_duration_s) if scenario_duration_s else 0.0
        gaps_s = self._compute_gap_durations_seconds(
            merged_intervals, scenario_duration_s=float(scenario_duration_s)
        )
        gaps_min = np.asarray(gaps_s, dtype=float) / 60.0 if gaps_s else np.asarray([], dtype=float)
        longest_gap_min = float(np.max(gaps_min)) if gaps_min.size else float(scenario_duration_s / 60.0)

        def _fmt_gap(x: float) -> str:
            return self._format_duration_minutes(float(x))

        _set("Selected stations", f"{len(enabled_station_names)}")
        _set("Scenario duration", self._format_duration_minutes(float(scenario_duration_s) / 60.0))
        _set("Total access", self._format_duration_minutes(total_access_s / 60.0))
        _set("Coverage", f"{coverage:.2f}%")
        _set("Num gaps", f"{int(gaps_min.size)}")
        if gaps_min.size:
            _set("Mean gap", _fmt_gap(float(np.mean(gaps_min))))
            _set("Median gap", _fmt_gap(float(np.median(gaps_min))))
            _set("P95 gap", _fmt_gap(float(np.percentile(gaps_min, 95))))
        else:
            _set("Mean gap", "—")
            _set("Median gap", "—")
            _set("P95 gap", "—")
        _set("Longest no access", _fmt_gap(longest_gap_min))

    def _update_timeline_view(self, result) -> None:
        """Render timeline plot and gap statistics for the enabled station subset."""
        if getattr(self, "_current_config", None) is None:
            return
        if not result or not getattr(result, "passes", None):
            return
        config = self._current_config
        scenario_start = config.scenario.start_time
        scenario_end = config.scenario.end_time
        scenario_duration_s = float((scenario_end - scenario_start).total_seconds())
        if scenario_duration_s <= 0:
            return

        enabled = self._get_enabled_timeline_stations()
        passes_all = sorted(result.passes, key=lambda item: item.aos)
        if enabled:
            filtered_passes = [
                p
                for p in passes_all
                if (p.station_name or "Ground Station") in enabled
            ]
        else:
            filtered_passes = []

        # Re-render the timeline plot for the subset.
        station_groups: dict[str, list[PassStatistic]] = {}
        for item in filtered_passes:
            station_name = item.station_name or "Ground Station"
            station_groups.setdefault(station_name, []).append(item)
        station_colors = getattr(self, "_station_color_map", None) or self._build_station_color_map(
            sorted(station_groups.keys())
        )
        self.timeline_plot.clear()
        if getattr(self, "timeline_legend", None):
            self.timeline_legend.clear()
        for station_name, group in station_groups.items():
            durations = [p.duration_minutes for p in group]
            aos_hours = [(p.aos - scenario_start).total_seconds() / 3600.0 for p in group]
            widths = [d / 60.0 for d in durations]
            color = station_colors.get(station_name)
            if color is None:
                continue
            brush = pg.mkBrush(color)
            pen = pg.mkPen(color.darker(125))
            timeline_item = pg.BarGraphItem(
                x=[a + w / 2 for a, w in zip(aos_hours, widths)],
                height=[1.0] * len(aos_hours),
                width=widths,
                brush=brush,
                pen=pen,
            )
            self.timeline_plot.addItem(timeline_item)
            if getattr(self, "timeline_legend", None):
                self.timeline_legend.addItem(timeline_item, station_name)
        self.timeline_plot.setYRange(0, 1.2)
        self.timeline_plot.setTitle(
            f"Access Timeline{self._contact_window_mode_suffix()}"
        )

        merged = self._compute_union_intervals_seconds(
            filtered_passes, scenario_start=scenario_start, scenario_end=scenario_end
        )
        self._update_timeline_gap_stats(
            enabled_station_names=sorted(enabled),
            merged_intervals=merged,
            scenario_duration_s=scenario_duration_s,
        )

    def _draw_empty_plots(self) -> None:
        """Display placeholder text when no data is available."""
        self.timeline_plot.clear()
        if getattr(self, "timeline_legend", None):
            self.timeline_legend.clear()
        self.timeline_plot.setTitle("Access Timeline\n(Run analysis to view)")
        self._clear_pie_chart("Access Share by Station\n(Run analysis to view)")
        self._clear_elevation_distribution("Elevation Distribution\n(Run analysis to view)")
        if getattr(self, "_timeline_hover_label", None) is not None:
            self._timeline_hover_label.setVisible(False)
        if getattr(self, "_timeline_hover_line", None) is not None:
            self._timeline_hover_line.setVisible(False)

    def _contact_window_mode_suffix(self) -> str:
        use_link_close_windows = getattr(
            self, "_use_link_close_for_contact_windows", None
        )
        if callable(use_link_close_windows) and use_link_close_windows():
            return "\n(Dynamic link close)"
        return ""

    def _configure_pie_axes(self) -> None:
        """Initialize the matplotlib pie chart axes with dark styling."""
        if not getattr(self, "pie_axes", None):
            return
        self.pie_axes.clear()
        self.pie_axes.set_facecolor("#0b0b10")
        self.pie_axes.figure.set_facecolor("#0b0b10")
        self.pie_axes.tick_params(
            axis="both",
            colors="#dddddd",
            labelbottom=False,
            labelleft=False,
            bottom=False,
            left=False,
        )
        for spine in self.pie_axes.spines.values():
            spine.set_visible(False)
        self.pie_axes.set_title("Access Share by Station", color="white", pad=12)

    def _clear_pie_chart(self, title: str) -> None:
        """Reset the pie chart canvas with the provided title."""
        if not getattr(self, "pie_axes", None) or not getattr(self, "pie_canvas", None):
            return
        self._configure_pie_axes()
        self.pie_axes.set_title(title, color="white", pad=12)
        self.pie_canvas.draw_idle()

    def _update_plots(self, result) -> None:
        """Render the timeline and histogram plots."""
        if not result.passes or getattr(self, "_current_config", None) is None:
            self._draw_empty_plots()
            return
        passes = sorted(result.passes, key=lambda item: item.aos)
        start_time = self._current_config.scenario.start_time
        station_groups: dict[str, list[PassStatistic]] = {}
        for item in passes:
            station_name = item.station_name or "Ground Station"
            station_groups.setdefault(station_name, []).append(item)
        station_summaries = getattr(result, "station_summaries", [])
        ordered_station_names = (
            [entry.station_name for entry in station_summaries]
            if station_summaries
            else list(station_groups.keys())
        )
        station_colors = self._build_station_color_map(ordered_station_names)
        # Update timeline station selector (defaults all enabled).
        self._set_timeline_station_list(ordered_station_names)
        self.timeline_plot.clear()
        if getattr(self, "timeline_legend", None):
            self.timeline_legend.clear()
        for station_name, group in station_groups.items():
            durations = [p.duration_minutes for p in group]
            aos_hours = [(p.aos - start_time).total_seconds() / 3600.0 for p in group]
            widths = [d / 60.0 for d in durations]
            color = station_colors.get(station_name)
            brush = pg.mkBrush(color)
            pen = pg.mkPen(color.darker(125))
            timeline_item = pg.BarGraphItem(
                x=[a + w / 2 for a, w in zip(aos_hours, widths)],
                height=[1.0] * len(aos_hours),
                width=widths,
                brush=brush,
                pen=pen,
            )
            self.timeline_plot.addItem(timeline_item)
            if getattr(self, "timeline_legend", None):
                self.timeline_legend.addItem(timeline_item, station_name)
        self.timeline_plot.setYRange(0, 1.2)
        summaries_for_pie = station_summaries or self._summaries_from_groups(
            station_groups
        )
        self._update_pie_chart(summaries_for_pie, station_colors)
        # Recompute timeline subset stats based on current station toggles.
        self._update_timeline_view(result)
        self._set_elevation_station_list(ordered_station_names)
        self._update_elevation_distribution(result)

    def _build_station_color_map(self, station_names: list[str]) -> dict[str, QColor]:
        """Assign consistent colors per station for plot rendering."""
        if not station_names:
            self._station_color_map = {}
            return {}
        palette = STATION_COLOR_PALETTE or ["#2E8B57"]
        color_map: dict[str, QColor] = {}
        for idx, name in enumerate(station_names):
            color_map[name] = QColor(palette[idx % len(palette)])
        self._station_color_map = color_map
        return color_map

    def _summaries_from_groups(
        self, station_groups: dict[str, list[PassStatistic]]
    ) -> list[StationSummary]:
        """Build StationSummary objects from grouped pass stats."""
        summaries: list[StationSummary] = []
        for name, items in station_groups.items():
            summaries.append(
                StationSummary(
                    station_name=name,
                    total_passes=len(items),
                    total_access_minutes=float(
                        sum(item.duration_minutes for item in items)
                    ),
                )
            )
        return summaries

    def _update_pie_chart(
        self,
        station_summaries: list[StationSummary],
        station_colors: dict[str, QColor],
    ) -> None:
        """Render the pie chart showing per-station contribution."""
        if not getattr(self, "pie_axes", None) or not getattr(self, "pie_canvas", None):
            return
        if not station_summaries:
            self._clear_pie_chart("Access Share by Station\n(No data)")
            return
        total_minutes = sum(entry.total_access_minutes for entry in station_summaries)
        if total_minutes <= 0:
            self._clear_pie_chart("Access Share by Station\n(No access time)")
            return
        self._configure_pie_axes()
        sizes = [entry.total_access_minutes for entry in station_summaries]
        colors = [
            station_colors.get(
                entry.station_name,
                QColor(STATION_COLOR_PALETTE[idx % len(STATION_COLOR_PALETTE)]),
            ).name()
            for idx, entry in enumerate(station_summaries)
        ]
        labels = [
            f"{entry.station_name}: {entry.total_access_minutes / total_minutes * 100:.1f}%"
            for entry in station_summaries
        ]
        self.pie_axes.pie(
            sizes,
            labels=labels,
            colors=colors,
            startangle=90,
            wedgeprops={"edgecolor": "#111", "linewidth": 1},
            textprops={"color": "white"},
        )
        self.pie_axes.axis("equal")
        self.pie_axes.set_title("Access Share by Station", color="white", pad=12)
        self.pie_canvas.draw_idle()

    def _handle_timeline_hover(self, position) -> None:
        """Show the timeline x-coordinate under the mouse cursor."""
        label = getattr(self, "_timeline_hover_label", None)
        line = getattr(self, "_timeline_hover_line", None)
        if label is None or line is None or not getattr(self, "timeline_plot", None):
            return
        if not self.timeline_plot.sceneBoundingRect().contains(position):
            label.setVisible(False)
            line.setVisible(False)
            return
        mouse_point = self.timeline_plot.plotItem.vb.mapSceneToView(position)
        x_hours = float(mouse_point.x())
        if not np.isfinite(x_hours):
            label.setVisible(False)
            line.setVisible(False)
            return
        # Build label text.
        text = f"{x_hours:.3f} h"
        config = getattr(self, "_current_config", None)
        if config is not None and getattr(config, "scenario", None) is not None:
            start_time = config.scenario.start_time
            timestamp = start_time + timedelta(hours=x_hours)
            # Display as UTC (scenario times are stored as UTC in this app).
            text = f"{x_hours:.3f} h | {timestamp:%d-%b %H:%M:%S} UTC"
        label.setText(text)
        label.setVisible(True)
        line.setPos(x_hours)
        line.setVisible(True)



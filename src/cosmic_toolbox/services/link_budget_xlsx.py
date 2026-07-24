"""Export an uplink/downlink link budget to a styled ``.xlsx`` workbook.

The workbook reproduces the layout, styling and merged-cell structure of the
reference "N-STAR" link-budget spreadsheet (three sheets: ``Uplink``,
``Downlink`` and a combined ``Link Budget`` sheet) but writes *static computed
values* rather than live Excel formulas.  Every quantity is computed in Python
by porting the spreadsheet's own closed-form relations, so the output is
self-consistent and fully parameterised by :class:`LinkBudgetExportConfig`.

This module is part of the headless analysis core: it imports only ``numpy``
and ``openpyxl`` (both hard dependencies) plus, lazily, the toolbox ITU loss
model.  No Qt or plotting imports are permitted here.

Typical use::

    from cosmic_toolbox.services.link_budget_xlsx import (
        LinkBudgetExportConfig,
        write_link_budget_xlsx,
    )

    cfg = LinkBudgetExportConfig(ground_station_name="Svalbard, NO")
    write_link_budget_xlsx(cfg, "link_budget.xlsx")
"""

from __future__ import annotations

import json
import math
from dataclasses import asdict, dataclass, fields
from pathlib import Path
from typing import Any, Dict, Mapping

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# ── Physical constants (match the reference spreadsheet) ────────────────────
EARTH_RADIUS_KM = 6378.137
BOLTZMANN_CONSTANT_J_K = 1.380649e-23
NOISE_FIGURE_REFERENCE_TEMP_K = 290.0
SPEED_OF_LIGHT_M_S = 299792458.0


# ════════════════════════════════════════════════════════════════════════════
# Configuration
# ════════════════════════════════════════════════════════════════════════════
@dataclass
class LinkBudgetExportConfig:
    """All inputs required to render an uplink/downlink link budget workbook.

    Defaults reproduce the reference N-STAR scenario, so an empty config still
    produces a valid, representative workbook.  Atmospheric loss values may be
    supplied directly (``*_atmospheric_loss_*_dB``); when left as ``None`` and
    station coordinates (``latitude_deg``/``longitude_deg``) are provided, they
    are computed with the toolbox ITU-R slant-path loss model.
    """

    # ── Document / scenario metadata ────────────────────────────────────────
    title: str = "Link Budget"
    ground_station_name: str = "Svalbard, NO"

    # ── Common geometry ─────────────────────────────────────────────────────
    satellite_altitude_km: float = 250.0
    slant_elevation_deg: float = 5.0
    ground_altitude_km: float = 0.484

    # Station coordinates — only used to compute ITU atmospheric losses when the
    # explicit loss values below are left as ``None``.
    latitude_deg: float | None = None
    longitude_deg: float | None = None
    unavailability_percent: float = 0.1

    system_margin_dB: float = 6.0

    # ── Uplink (Ground → Spacecraft) ────────────────────────────────────────
    ul_frequency_MHz: float = 2106.0
    ul_modulation_type: str = "BPSK"
    ul_modulation_order: int = 1
    ul_coding_label: str = "BCH Coding Rate (63,56)"
    ul_coding_rate: float = 56.0 / 63.0
    ul_rolloff: float = 0.5
    ul_symbol_rate_ksps: float = 512.0
    ul_gs_eirp_dBW: float = 44.0
    ul_pointing_loss_dB: float = 0.0
    ul_atmospheric_loss_slant_dB: float | None = 1.14
    ul_atmospheric_loss_nadir_dB: float | None = 0.07
    ul_atmospheric_temperature_K: float = 266.0
    ul_background_noise_temp_K: float = 255.0
    ul_sc_rx_antenna_gain_dBi: float = 6.5
    ul_polarization_loss_dB: float = 0.5
    ul_sc_mispointing_loss_dB: float = 0.5
    ul_sc_harness_loss_dB: float = 2.0
    ul_sc_receiver_noise_figure_dB: float = 3.0
    ul_required_ebn0_dB: float = 9.6

    # ── Downlink (Spacecraft → Ground) ──────────────────────────────────────
    dl_frequency_MHz: float = 2212.0
    dl_modulation_type: str = "QPSK"
    dl_modulation_order: int = 2
    dl_coding_label: str = "Viterbi k=7 r = 1/2 coding rate"
    dl_coding_rate: float = 0.5
    dl_rolloff: float = 0.5
    dl_symbol_rate_ksps: float = 2000.0
    dl_tx_power_dBW: float = 0.0
    dl_tx_harness_loss_dB: float = 4.0
    dl_tx_antenna_gain_slant_dBi: float = -1.0
    dl_tx_antenna_gain_nadir_dBi: float = 6.0
    dl_pointing_loss_dB: float = 0.0
    dl_implementation_loss_dB: float = 1.5
    dl_atmospheric_loss_slant_dB: float | None = 1.17
    dl_atmospheric_loss_nadir_dB: float | None = 0.07
    dl_atmospheric_temperature_K: float = 266.0
    dl_background_noise_temp_K: float = 50.0
    dl_rx_antenna_gain_dBi: float = 34.66
    dl_polarization_loss_dB: float = 0.5
    dl_mispointing_loss_dB: float = 0.0
    dl_system_noise_temperature_K: float = 187.0
    # NOTE: the "Eb/N0" rows in this exporter are computed as
    # C/N0 - 10log10(symbol_rate), which is Es/N0 (energy per *symbol*).  For
    # QPSK Viterbi r=1/2 the net spectral efficiency is 1.0 (0 dB) so Es/N0 ==
    # Eb/N0 and this 4.5 dB threshold matches
    # link_budget_math.QPSK_VITERBI_REQUIRED_ESN0_DB.  If you change to a mode
    # with spectral efficiency != 1.0, this value must be the Es/N0 threshold,
    # not the raw Eb/N0.
    dl_required_ebn0_dB: float = 4.5
    include_pfd: bool = True

    # ── Construction helpers ────────────────────────────────────────────────
    @classmethod
    def from_dict(cls, data: Mapping[str, Any]) -> "LinkBudgetExportConfig":
        """Build a config from a mapping, ignoring unknown keys gracefully."""
        known = {f.name for f in fields(cls)}
        unknown = set(data) - known
        if unknown:
            raise ValueError(
                "Unknown link-budget config keys: " + ", ".join(sorted(unknown))
            )
        return cls(**{k: v for k, v in data.items() if k in known})

    def to_dict(self) -> Dict[str, Any]:
        """Return a plain-dict representation (useful for serialising)."""
        return asdict(self)


def load_config(path: str | Path) -> LinkBudgetExportConfig:
    """Load a :class:`LinkBudgetExportConfig` from a YAML or JSON file.

    ``.yaml``/``.yml`` files are parsed with PyYAML; everything else is parsed
    as JSON.
    """
    path = Path(path)
    text = path.read_text(encoding="utf-8")
    if path.suffix.lower() in {".yaml", ".yml"}:
        import yaml  # local import; PyYAML is a core dependency

        data = yaml.safe_load(text) or {}
    else:
        data = json.loads(text)
    if not isinstance(data, Mapping):
        raise ValueError(f"Config file {path} must contain a mapping at the top level.")
    return LinkBudgetExportConfig.from_dict(data)


# ════════════════════════════════════════════════════════════════════════════
# Link-budget math (ports of the reference spreadsheet formulas)
# ════════════════════════════════════════════════════════════════════════════
def _slant_range_km(
    elevation_deg: float, sat_altitude_km: float, ground_altitude_km: float
) -> float:
    """Geometric slant range (km) — matches the spreadsheet's SQRT formula."""
    el = math.radians(elevation_deg)
    r_gs = EARTH_RADIUS_KM + ground_altitude_km
    r_sat = EARTH_RADIUS_KM + sat_altitude_km
    return math.sqrt(r_sat**2 - r_gs**2 * math.cos(el) ** 2) - r_gs * math.sin(el)


def _fspl_dB(slant_km: float, wavelength_m: float) -> float:
    """Free-space path loss (dB) using 20*log10(4*pi*d/lambda)."""
    return 20.0 * math.log10(4.0 * math.pi * slant_km * 1000.0 / wavelength_m)


def _brightness_temp_K(
    atm_loss_dB: float, background_temp_K: float, atmospheric_temp_K: float
) -> float:
    a = 10.0 ** (-atm_loss_dB / 10.0)
    return background_temp_K * a + atmospheric_temp_K * (1.0 - a)


def _pfd_limit_dBW_m2_4kHz(elevation_deg: float, frequency_MHz: float) -> float:
    """ITU Article 21 surface PFD mask, delegating to the shared ``pfd_math``.

    Single source of truth for the mask (S/X/Ka) instead of a local S-band
    copy.  The workbook normalises PFD to a 4 kHz reference bandwidth, which
    matches the S- and X-band mask reference bandwidths.
    """
    from cosmic_toolbox import pfd_math

    return float(
        pfd_math.itu_surface_pfd_limit_dBW_per_m2(
            [elevation_deg], frequency_MHz=frequency_MHz
        )[0]
    )


def _maybe_itu_atmospheric_loss(
    cfg: LinkBudgetExportConfig, frequency_GHz: float, elevation_deg: float
) -> float | None:
    """Compute one-way atmospheric loss with the toolbox ITU model, or None."""
    if cfg.latitude_deg is None or cfg.longitude_deg is None:
        return None
    from cosmic_toolbox.itu_losses import estimate_slant_path_loss

    losses = estimate_slant_path_loss(
        frequency_GHz=frequency_GHz,
        elevations_deg=[max(elevation_deg, 0.1)],
        lat_deg=cfg.latitude_deg,
        lon_deg=cfg.longitude_deg,
        altitude_m=cfg.ground_altitude_km * 1000.0,
        unavailability_percent=cfg.unavailability_percent,
    )
    return float(losses[0])


def compute_uplink(cfg: LinkBudgetExportConfig) -> Dict[str, float]:
    """Compute every uplink quantity displayed on the ``Uplink`` sheet."""
    freq_ghz = cfg.ul_frequency_MHz / 1000.0
    wavelength = SPEED_OF_LIGHT_M_S / (1e6 * cfg.ul_frequency_MHz)
    k = BOLTZMANN_CONSTANT_J_K
    t0 = NOISE_FIGURE_REFERENCE_TEMP_K

    atm_slant = cfg.ul_atmospheric_loss_slant_dB
    if atm_slant is None:
        atm_slant = _maybe_itu_atmospheric_loss(cfg, freq_ghz, cfg.slant_elevation_deg)
    if atm_slant is None:
        atm_slant = 0.0
    atm_nadir = cfg.ul_atmospheric_loss_nadir_dB
    if atm_nadir is None:
        atm_nadir = _maybe_itu_atmospheric_loss(cfg, freq_ghz, 90.0)
    if atm_nadir is None:
        atm_nadir = 0.0

    info_rate_kbps = cfg.ul_symbol_rate_ksps * cfg.ul_modulation_order
    occupied_bw_khz = cfg.ul_symbol_rate_ksps * (1.0 + cfg.ul_rolloff)
    spectral_eff = info_rate_kbps / occupied_bw_khz
    effective_eirp = cfg.ul_gs_eirp_dBW - cfg.ul_pointing_loss_dB

    slant_range = _slant_range_km(
        cfg.slant_elevation_deg, cfg.satellite_altitude_km, cfg.ground_altitude_km
    )
    nadir_range = _slant_range_km(
        90.0, cfg.satellite_altitude_km, cfg.ground_altitude_km
    )
    fspl_slant = _fspl_dB(slant_range, wavelength)
    fspl_nadir = _fspl_dB(nadir_range, wavelength)

    incident_slant = effective_eirp - fspl_slant - atm_slant
    incident_nadir = effective_eirp - fspl_nadir - atm_nadir

    bt_slant = _brightness_temp_K(
        atm_slant, cfg.ul_background_noise_temp_K, cfg.ul_atmospheric_temperature_K
    )
    bt_nadir = _brightness_temp_K(
        atm_nadir, cfg.ul_background_noise_temp_K, cfg.ul_atmospheric_temperature_K
    )
    incident_noise_slant = 10.0 * math.log10(bt_slant * k * occupied_bw_khz * 1000.0)
    incident_noise_nadir = 10.0 * math.log10(bt_nadir * k * occupied_bw_khz * 1000.0)

    harness_temp = 290.0 * (1.0 - 1.0 / (10.0 ** (cfg.ul_sc_harness_loss_dB / 10.0)))
    ant_noise_slant = 0.9 * bt_slant + 0.1 * cfg.ul_atmospheric_temperature_K
    ant_noise_nadir = 0.9 * bt_nadir + 0.1 * cfg.ul_atmospheric_temperature_K
    rx_noise_temp = 290.0 * (10.0 ** (cfg.ul_sc_receiver_noise_figure_dB / 10.0) - 1.0)
    harness_lin = 10.0 ** (-cfg.ul_sc_harness_loss_dB / 10.0)
    sys_noise_slant = ant_noise_slant * harness_lin + t0 * (1 - harness_lin) + rx_noise_temp
    sys_noise_nadir = ant_noise_nadir * harness_lin + t0 * (1 - harness_lin) + rx_noise_temp

    gt_slant = (
        cfg.ul_sc_rx_antenna_gain_dBi
        - cfg.ul_sc_harness_loss_dB
        - 10.0 * math.log10(sys_noise_slant)
    )
    gt_nadir = (
        cfg.ul_sc_rx_antenna_gain_dBi
        - cfg.ul_sc_harness_loss_dB
        - 10.0 * math.log10(sys_noise_nadir)
    )
    eff_gt_slant = gt_slant - cfg.ul_polarization_loss_dB - cfg.ul_sc_mispointing_loss_dB
    eff_gt_nadir = gt_nadir - cfg.ul_polarization_loss_dB - cfg.ul_sc_mispointing_loss_dB

    rx_carrier_slant = (
        incident_slant
        + cfg.ul_sc_rx_antenna_gain_dBi
        - cfg.ul_polarization_loss_dB
        - cfg.ul_sc_mispointing_loss_dB
        - cfg.ul_sc_harness_loss_dB
    )
    rx_carrier_nadir = (
        incident_nadir
        + cfg.ul_sc_rx_antenna_gain_dBi
        - cfg.ul_polarization_loss_dB
        - cfg.ul_sc_mispointing_loss_dB
        - cfg.ul_sc_harness_loss_dB
    )
    rx_noise_power = 10.0 * math.log10(k * 1000.0 * occupied_bw_khz * sys_noise_slant)
    cn_slant = rx_carrier_slant - rx_noise_power
    cn_nadir = rx_carrier_nadir - rx_noise_power
    cn0_slant = cn_slant + 10.0 * math.log10(occupied_bw_khz * 1000.0)
    cn0_nadir = cn_nadir + 10.0 * math.log10(occupied_bw_khz * 1000.0)
    ebn0_slant = cn0_slant - 10.0 * math.log10(cfg.ul_symbol_rate_ksps * 1000.0)
    ebn0_nadir = cn0_nadir - 10.0 * math.log10(cfg.ul_symbol_rate_ksps * 1000.0)
    excess_slant = ebn0_slant - cfg.ul_required_ebn0_dB - cfg.system_margin_dB
    excess_nadir = ebn0_nadir - cfg.ul_required_ebn0_dB - cfg.system_margin_dB
    rx_sensitivity = -156.0 + 10.0 * math.log10(cfg.ul_symbol_rate_ksps * 1000.0) - 30.0
    sensitivity_margin = rx_carrier_slant - rx_sensitivity - cfg.system_margin_dB

    return {
        "wavelength_m": wavelength,
        "info_rate_kbps": info_rate_kbps,
        "occupied_bw_khz": occupied_bw_khz,
        "spectral_eff": spectral_eff,
        "effective_eirp_dBW": effective_eirp,
        "slant_range_km": slant_range,
        "nadir_range_km": nadir_range,
        "fspl_slant_dB": fspl_slant,
        "fspl_nadir_dB": fspl_nadir,
        "atm_slant_dB": atm_slant,
        "atm_nadir_dB": atm_nadir,
        "incident_slant_dBW": incident_slant,
        "incident_nadir_dBW": incident_nadir,
        "brightness_temp_slant_K": bt_slant,
        "brightness_temp_nadir_K": bt_nadir,
        "incident_noise_slant_dBW": incident_noise_slant,
        "incident_noise_nadir_dBW": incident_noise_nadir,
        "harness_temp_K": harness_temp,
        "antenna_noise_slant_K": ant_noise_slant,
        "antenna_noise_nadir_K": ant_noise_nadir,
        "rx_noise_temp_K": rx_noise_temp,
        "sys_noise_slant_K": sys_noise_slant,
        "sys_noise_nadir_K": sys_noise_nadir,
        "gt_slant_dB_K": gt_slant,
        "gt_nadir_dB_K": gt_nadir,
        "eff_gt_slant_dB_K": eff_gt_slant,
        "eff_gt_nadir_dB_K": eff_gt_nadir,
        "rx_carrier_slant_dBW": rx_carrier_slant,
        "rx_carrier_nadir_dBW": rx_carrier_nadir,
        "rx_noise_power_dBW": rx_noise_power,
        "cn_slant_dB": cn_slant,
        "cn_nadir_dB": cn_nadir,
        "cn0_slant_dBHz": cn0_slant,
        "cn0_nadir_dBHz": cn0_nadir,
        "ebn0_slant_dB": ebn0_slant,
        "ebn0_nadir_dB": ebn0_nadir,
        "excess_slant_dB": excess_slant,
        "excess_nadir_dB": excess_nadir,
        "rx_sensitivity_dBW": rx_sensitivity,
        "sensitivity_margin_dB": sensitivity_margin,
    }


def compute_downlink(cfg: LinkBudgetExportConfig) -> Dict[str, float]:
    """Compute every downlink quantity displayed on the ``Downlink`` sheet."""
    freq_ghz = cfg.dl_frequency_MHz / 1000.0
    wavelength = SPEED_OF_LIGHT_M_S / (1e6 * cfg.dl_frequency_MHz)
    k = BOLTZMANN_CONSTANT_J_K

    atm_slant = cfg.dl_atmospheric_loss_slant_dB
    if atm_slant is None:
        atm_slant = _maybe_itu_atmospheric_loss(cfg, freq_ghz, cfg.slant_elevation_deg)
    if atm_slant is None:
        atm_slant = 0.0
    atm_nadir = cfg.dl_atmospheric_loss_nadir_dB
    if atm_nadir is None:
        atm_nadir = _maybe_itu_atmospheric_loss(cfg, freq_ghz, 90.0)
    if atm_nadir is None:
        atm_nadir = 0.0

    info_rate_kbps = cfg.dl_symbol_rate_ksps * cfg.dl_coding_rate * cfg.dl_modulation_order
    occupied_bw_khz = cfg.dl_symbol_rate_ksps * (1.0 + cfg.dl_rolloff)
    spectral_eff = info_rate_kbps / occupied_bw_khz

    eirp_slant = (
        cfg.dl_tx_power_dBW
        - cfg.dl_tx_harness_loss_dB
        + cfg.dl_tx_antenna_gain_slant_dBi
        - cfg.dl_pointing_loss_dB
        - cfg.dl_implementation_loss_dB
    )
    eirp_nadir = (
        cfg.dl_tx_power_dBW
        - cfg.dl_tx_harness_loss_dB
        + cfg.dl_tx_antenna_gain_nadir_dBi
        - cfg.dl_pointing_loss_dB
        - cfg.dl_implementation_loss_dB
    )

    slant_range = _slant_range_km(
        cfg.slant_elevation_deg, cfg.satellite_altitude_km, cfg.ground_altitude_km
    )
    nadir_range = _slant_range_km(
        90.0, cfg.satellite_altitude_km, cfg.ground_altitude_km
    )
    fspl_slant = _fspl_dB(slant_range, wavelength)
    fspl_nadir = _fspl_dB(nadir_range, wavelength)

    incident_slant = eirp_slant - fspl_slant - atm_slant
    incident_nadir = eirp_nadir - fspl_nadir - atm_nadir

    bt_slant = _brightness_temp_K(
        atm_slant, cfg.dl_background_noise_temp_K, cfg.dl_atmospheric_temperature_K
    )
    bt_nadir = _brightness_temp_K(
        atm_nadir, cfg.dl_background_noise_temp_K, cfg.dl_atmospheric_temperature_K
    )
    incident_noise_slant = 10.0 * math.log10(bt_slant * k * occupied_bw_khz * 1000.0)
    incident_noise_nadir = 10.0 * math.log10(bt_nadir * k * occupied_bw_khz * 1000.0)

    sys_noise = cfg.dl_system_noise_temperature_K
    gt = cfg.dl_rx_antenna_gain_dBi - 10.0 * math.log10(sys_noise)
    eff_gt = gt - cfg.dl_polarization_loss_dB - cfg.dl_mispointing_loss_dB

    rx_carrier_slant = (
        incident_slant
        + cfg.dl_rx_antenna_gain_dBi
        - cfg.dl_polarization_loss_dB
        - cfg.dl_mispointing_loss_dB
    )
    rx_carrier_nadir = (
        incident_nadir
        + cfg.dl_rx_antenna_gain_dBi
        - cfg.dl_polarization_loss_dB
        - cfg.dl_mispointing_loss_dB
    )
    rx_noise_power = 10.0 * math.log10(k * 1000.0 * occupied_bw_khz * sys_noise)
    cn_slant = rx_carrier_slant - rx_noise_power
    cn_nadir = rx_carrier_nadir - rx_noise_power
    cn0_slant = cn_slant + 10.0 * math.log10(occupied_bw_khz * 1000.0)
    cn0_nadir = cn_nadir + 10.0 * math.log10(occupied_bw_khz * 1000.0)
    ebn0_slant = cn0_slant - 10.0 * math.log10(cfg.dl_symbol_rate_ksps * 1000.0)
    ebn0_nadir = cn0_nadir - 10.0 * math.log10(cfg.dl_symbol_rate_ksps * 1000.0)
    excess_slant = ebn0_slant - cfg.dl_required_ebn0_dB - cfg.system_margin_dB
    excess_nadir = ebn0_nadir - cfg.dl_required_ebn0_dB - cfg.system_margin_dB

    spreading_slant = 10.0 * math.log10(4.0 * math.pi * (slant_range * 1000.0) ** 2)
    spreading_nadir = 10.0 * math.log10(4.0 * math.pi * (nadir_range * 1000.0) ** 2)
    pfd_slant = eirp_slant - spreading_slant - 10.0 * math.log10(occupied_bw_khz / 4.0)
    pfd_nadir = eirp_nadir - spreading_nadir - 10.0 * math.log10(occupied_bw_khz / 4.0)
    pfd_limit_slant = _pfd_limit_dBW_m2_4kHz(
        cfg.slant_elevation_deg, cfg.dl_frequency_MHz
    )
    pfd_limit_nadir = _pfd_limit_dBW_m2_4kHz(90.0, cfg.dl_frequency_MHz)
    pfd_margin_slant = pfd_limit_slant - pfd_slant
    pfd_margin_nadir = pfd_limit_nadir - pfd_nadir

    return {
        "wavelength_m": wavelength,
        "info_rate_kbps": info_rate_kbps,
        "occupied_bw_khz": occupied_bw_khz,
        "spectral_eff": spectral_eff,
        "eirp_slant_dBW": eirp_slant,
        "eirp_nadir_dBW": eirp_nadir,
        "slant_range_km": slant_range,
        "nadir_range_km": nadir_range,
        "fspl_slant_dB": fspl_slant,
        "fspl_nadir_dB": fspl_nadir,
        "atm_slant_dB": atm_slant,
        "atm_nadir_dB": atm_nadir,
        "incident_slant_dBW": incident_slant,
        "incident_nadir_dBW": incident_nadir,
        "brightness_temp_slant_K": bt_slant,
        "brightness_temp_nadir_K": bt_nadir,
        "incident_noise_slant_dBW": incident_noise_slant,
        "incident_noise_nadir_dBW": incident_noise_nadir,
        "sys_noise_K": sys_noise,
        "gt_dB_K": gt,
        "eff_gt_dB_K": eff_gt,
        "rx_carrier_slant_dBW": rx_carrier_slant,
        "rx_carrier_nadir_dBW": rx_carrier_nadir,
        "rx_noise_power_dBW": rx_noise_power,
        "cn_slant_dB": cn_slant,
        "cn_nadir_dB": cn_nadir,
        "cn0_slant_dBHz": cn0_slant,
        "cn0_nadir_dBHz": cn0_nadir,
        "ebn0_slant_dB": ebn0_slant,
        "ebn0_nadir_dB": ebn0_nadir,
        "excess_slant_dB": excess_slant,
        "excess_nadir_dB": excess_nadir,
        "spreading_slant_dBm2": spreading_slant,
        "spreading_nadir_dBm2": spreading_nadir,
        "pfd_slant": pfd_slant,
        "pfd_nadir": pfd_nadir,
        "pfd_limit_slant": pfd_limit_slant,
        "pfd_limit_nadir": pfd_limit_nadir,
        "pfd_margin_slant": pfd_margin_slant,
        "pfd_margin_nadir": pfd_margin_nadir,
    }


# ════════════════════════════════════════════════════════════════════════════
# Workbook styling helpers
# ════════════════════════════════════════════════════════════════════════════
_THIN = Side(style="thin", color="000000")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_BOLD = Font(bold=True)
_PLAIN = Font(bold=False)
_CENTER = Alignment(horizontal="center", vertical="center")
_LEFT_CENTER = Alignment(horizontal="left", vertical="center")
_VTOP = Alignment(vertical="top")


def _put(
    ws: Worksheet,
    coord: str,
    value: Any,
    *,
    bold: bool = False,
    number_format: str | None = None,
    align: Alignment | None = None,
    border: bool = True,
) -> None:
    """Write a value to a cell and apply the standard table styling."""
    cell = ws[coord]
    cell.value = value
    cell.font = _BOLD if bold else _PLAIN
    if number_format is not None:
        cell.number_format = number_format
    if align is not None:
        cell.alignment = align
    if border:
        cell.border = _BORDER


# ════════════════════════════════════════════════════════════════════════════
# Sheet builders
# ════════════════════════════════════════════════════════════════════════════
def _build_block(
    ws: Worksheet,
    rows: list[dict],
    *,
    base_col: int,
    start_row: int,
    title: str,
) -> None:
    """Render a single Parameters | Inputs | Calculations | Units block.

    ``rows`` is an ordered list of row specs.  Each spec is a dict with keys:
      label, input, calc, unit (any may be omitted), number_format,
      and optional ``sub`` (the Slant/Nadir label in the column after label).
    A spec may carry ``merge_next`` to merge its label cell with the row below
    (used for Slant/Nadir pairs).
    """
    c_param = base_col          # B-style: parameter label
    c_sub = base_col + 1        # C-style: Slant/Nadir
    c_input = base_col + 2      # D-style: inputs
    c_calc = base_col + 3       # E-style: calculations
    c_unit = base_col + 4       # F-style: units

    pl = get_column_letter(c_param)
    sl = get_column_letter(c_sub)
    il = get_column_letter(c_input)
    cl = get_column_letter(c_calc)
    ul = get_column_letter(c_unit)

    # Title row
    _put(ws, f"{pl}{start_row}", title, bold=True, border=False)
    # Header row
    hdr = start_row + 2
    _put(ws, f"{pl}{hdr}", "Parameters", bold=True)
    _put(ws, f"{il}{hdr}", "Inputs", bold=True, align=_CENTER)
    _put(ws, f"{cl}{hdr}", "Calculations", bold=True, align=_CENTER)
    _put(ws, f"{ul}{hdr}", "Units", bold=True, align=_CENTER)

    r = hdr + 1
    for spec in rows:
        label = spec.get("label", "")
        sub = spec.get("sub")
        fmt = spec.get("number_format", "0.00")
        merge = spec.get("merge_next", False)
        if label:
            _put(ws, f"{pl}{r}", label, align=_LEFT_CENTER if merge else None)
        if sub is not None:
            _put(ws, f"{sl}{r}", sub, align=_VTOP)
        if "input" in spec and spec["input"] is not None:
            _put(ws, f"{il}{r}", spec["input"], number_format=fmt, align=_CENTER)
        if "calc" in spec and spec["calc"] is not None:
            _put(ws, f"{cl}{r}", spec["calc"], number_format=fmt, align=_CENTER)
        if "unit" in spec and spec["unit"]:
            _put(ws, f"{ul}{r}", spec["unit"], align=_CENTER)
        if merge:
            ws.merge_cells(f"{pl}{r}:{pl}{r + 1}")
        r += 1

    # Column widths to mirror the reference sheet.
    ws.column_dimensions[pl].width = 38.1
    ws.column_dimensions[sl].width = 7.5
    ws.column_dimensions[il].width = 12.0
    ws.column_dimensions[cl].width = 12.0
    ws.column_dimensions[ul].width = 13.8


def _build_summary_block(
    ws: Worksheet, rows: list[dict], *, base_col: int, start_row: int, title: str
) -> None:
    """Render the right-hand Parameters | Value | Units summary block."""
    c_param = base_col          # H-style
    c_sub = base_col + 1        # I-style: Slant/Nadir
    c_value = base_col + 2      # J-style
    c_unit = base_col + 3       # K-style
    pl = get_column_letter(c_param)
    sl = get_column_letter(c_sub)
    vl = get_column_letter(c_value)
    ul = get_column_letter(c_unit)

    _put(ws, f"{pl}{start_row}", title, bold=True, border=False)
    hdr = start_row + 2
    _put(ws, f"{pl}{hdr}", "Parameters", bold=True)
    _put(ws, f"{vl}{hdr}", "Value", bold=True, align=_CENTER)
    _put(ws, f"{ul}{hdr}", "Units", bold=True, align=_CENTER)

    r = hdr + 1
    for spec in rows:
        if spec.get("label"):
            _put(ws, f"{pl}{r}", spec["label"])
        if spec.get("sub") is not None:
            _put(ws, f"{sl}{r}", spec["sub"], align=_VTOP)
        _put(ws, f"{vl}{r}", spec["value"], number_format=spec.get("number_format", "0.00"), align=_CENTER)
        if spec.get("unit"):
            _put(ws, f"{ul}{r}", spec["unit"], align=_CENTER)
        r += 1
    ws.column_dimensions[pl].width = 30.0
    ws.column_dimensions[vl].width = 12.0
    ws.column_dimensions[ul].width = 9.0


def _uplink_rows(cfg: LinkBudgetExportConfig, u: Dict[str, float]) -> list[dict]:
    """Ordered uplink parameter rows mirroring the reference Uplink sheet."""
    return [
        {"label": "Earth Radius", "input": EARTH_RADIUS_KM, "unit": "km", "number_format": "0"},
        {"label": "Boltzmann Constant", "input": BOLTZMANN_CONSTANT_J_K, "unit": "J/K", "number_format": "0.00E+00"},
        {"label": "Noise Figure Reference Temp.", "input": NOISE_FIGURE_REFERENCE_TEMP_K, "unit": "K", "number_format": "0"},
        {"label": "Speed of Light", "input": SPEED_OF_LIGHT_M_S, "unit": "m/s", "number_format": "0"},
        {"label": "Uplink Frequency", "input": cfg.ul_frequency_MHz, "unit": "MHz", "number_format": "0.0"},
        {"label": "Uplink Wavelength", "calc": u["wavelength_m"], "unit": "m", "number_format": "0.000"},
        {"label": "Modulation Type", "input": cfg.ul_modulation_type, "unit": "", "number_format": "General"},
        {"label": "Modulation Order", "input": cfg.ul_modulation_order, "unit": "", "number_format": "0"},
        {"label": cfg.ul_coding_label, "input": cfg.ul_coding_rate, "unit": "", "number_format": "0.000"},
        {"label": "SRRC Roll-off", "input": cfg.ul_rolloff, "unit": "", "number_format": "0.00"},
        {"label": "Symbol Rate", "input": cfg.ul_symbol_rate_ksps, "unit": "ksps", "number_format": "0"},
        {"label": "Information Rate", "calc": u["info_rate_kbps"], "unit": "kbps", "number_format": "0"},
        {"label": "Occupied Bandwidth", "calc": u["occupied_bw_khz"], "unit": "kHz", "number_format": "0.0"},
        {"label": "Spectral Efficiency", "calc": u["spectral_eff"], "unit": "bit/sym/Hz", "number_format": "0.000"},
        {"label": "TX EIRP", "input": cfg.ul_gs_eirp_dBW, "unit": "dBW", "number_format": "0.00"},
        {"label": "Pointing Loss", "input": cfg.ul_pointing_loss_dB, "unit": "dB", "number_format": "0.00"},
        {"label": "Effective EIRP", "calc": u["effective_eirp_dBW"], "unit": "dBW", "number_format": "0.00"},
        {"label": "Orbit Altitude", "input": cfg.satellite_altitude_km, "unit": "km", "number_format": "0"},
        {"label": "Slant Elevation", "input": cfg.slant_elevation_deg, "unit": "deg", "number_format": "0.0"},
        {"label": "Ground Station Name", "input": cfg.ground_station_name, "unit": "", "number_format": "General"},
        {"label": "Ground Station Altitude", "input": cfg.ground_altitude_km, "unit": "km", "number_format": "0.000"},
        {"label": "Slant Range", "sub": "Slant", "calc": u["slant_range_km"], "unit": "km", "number_format": "0.0", "merge_next": True},
        {"sub": "Nadir", "calc": u["nadir_range_km"], "unit": "km", "number_format": "0.0"},
        {"label": "Free Space Path Loss", "sub": "Slant", "calc": u["fspl_slant_dB"], "unit": "dB", "merge_next": True},
        {"sub": "Nadir", "calc": u["fspl_nadir_dB"], "unit": "dB"},
        {"label": "Atmospheric Loss", "sub": "Slant", "input": u["atm_slant_dB"], "unit": "dB", "merge_next": True},
        {"sub": "Nadir", "input": u["atm_nadir_dB"], "unit": "dB"},
        {"label": "Incident Channel Power", "sub": "Slant", "calc": u["incident_slant_dBW"], "unit": "dBW", "merge_next": True},
        {"sub": "Nadir", "calc": u["incident_nadir_dBW"], "unit": "dBW"},
        {"label": "Atmospheric Temperature", "input": cfg.ul_atmospheric_temperature_K, "unit": "K", "number_format": "0.0"},
        {"label": "Background Noise Temp", "input": cfg.ul_background_noise_temp_K, "unit": "K", "number_format": "0.0"},
        {"label": "Brightness Temperature", "sub": "Slant", "calc": u["brightness_temp_slant_K"], "unit": "K", "number_format": "0.0", "merge_next": True},
        {"sub": "Nadir", "calc": u["brightness_temp_nadir_K"], "unit": "K", "number_format": "0.0"},
        {"label": "Incident Noise Power", "sub": "Slant", "calc": u["incident_noise_slant_dBW"], "unit": "dBW", "merge_next": True},
        {"sub": "Nadir", "calc": u["incident_noise_nadir_dBW"], "unit": "dBW"},
        {"label": "Rx Antenna Gain", "input": cfg.ul_sc_rx_antenna_gain_dBi, "unit": "dBi"},
        {"label": "Polarization Loss", "input": cfg.ul_polarization_loss_dB, "unit": "dB"},
        {"label": "Mispointing Loss", "input": cfg.ul_sc_mispointing_loss_dB, "unit": "dB"},
        {"label": "Harness Loss", "input": cfg.ul_sc_harness_loss_dB, "unit": "dB"},
        {"label": "Harness Temperature", "calc": u["harness_temp_K"], "unit": "K", "number_format": "0"},
        {"label": "Receiver Noise Figure", "input": cfg.ul_sc_receiver_noise_figure_dB, "unit": "dB"},
        {"label": "Antenna Noise Temperature", "sub": "Slant", "calc": u["antenna_noise_slant_K"], "unit": "K", "number_format": "0", "merge_next": True},
        {"sub": "Nadir", "calc": u["antenna_noise_nadir_K"], "unit": "K", "number_format": "0"},
        {"label": "Receiver Noise Temperature", "calc": u["rx_noise_temp_K"], "unit": "K", "number_format": "0"},
        {"label": "System Noise Temperature", "sub": "Slant", "calc": u["sys_noise_slant_K"], "unit": "K", "number_format": "0", "merge_next": True},
        {"sub": "Nadir", "calc": u["sys_noise_nadir_K"], "unit": "K", "number_format": "0"},
        {"label": "System G/T", "sub": "Slant", "calc": u["gt_slant_dB_K"], "unit": "dB/K", "merge_next": True},
        {"sub": "Nadir", "calc": u["gt_nadir_dB_K"], "unit": "dB/K"},
        {"label": "Effective System G/T", "sub": "Slant", "calc": u["eff_gt_slant_dB_K"], "unit": "dB/K", "merge_next": True},
        {"sub": "Nadir", "calc": u["eff_gt_nadir_dB_K"], "unit": "dB/K"},
        {"label": "Received Carrier Power (C) ", "sub": "Slant", "calc": u["rx_carrier_slant_dBW"], "unit": "dBW", "merge_next": True},
        {"sub": "Nadir", "calc": u["rx_carrier_nadir_dBW"], "unit": "dBW"},
        {"label": "Received Noise Power (N)", "calc": u["rx_noise_power_dBW"], "unit": "dBW"},
        {"label": "Carrier-to-Noise Ratio (C/N)", "sub": "Slant", "calc": u["cn_slant_dB"], "unit": "dB", "merge_next": True},
        {"sub": "Nadir", "calc": u["cn_nadir_dB"], "unit": "dB"},
        {"label": "Carrier-to-Noise-Density Ratio (C/No)", "sub": "Slant", "calc": u["cn0_slant_dBHz"], "unit": "dBHz", "merge_next": True},
        {"sub": "Nadir", "calc": u["cn0_nadir_dBHz"], "unit": "dBHz"},
        {"label": "Eb/No", "sub": "Slant", "calc": u["ebn0_slant_dB"], "unit": "dB", "merge_next": True},
        {"sub": "Nadir", "calc": u["ebn0_nadir_dB"], "unit": "dB"},
        {"label": "Required Eb/N0", "input": cfg.ul_required_ebn0_dB, "unit": "dB"},
        {"label": "System Margin", "input": cfg.system_margin_dB, "unit": "dB"},
        {"label": "Excess Margin", "sub": "Slant", "calc": u["excess_slant_dB"], "unit": "dB", "merge_next": True},
        {"sub": "Nadir", "calc": u["excess_nadir_dB"], "unit": "dB"},
        {"label": "RX Sensitivity Level @ BER = 1e-5", "calc": u["rx_sensitivity_dBW"], "unit": "dBW"},
        {"label": "Sensitivity Margin inc. System Margin", "calc": u["sensitivity_margin_dB"], "unit": "dB"},
    ]


def _uplink_summary_rows(u: Dict[str, float]) -> list[dict]:
    return [
        {"label": "EIRP", "value": u["effective_eirp_dBW"], "unit": "dBW"},
        {"label": "Effective System G/T", "sub": "Slant", "value": u["eff_gt_slant_dB_K"], "unit": "dB/K"},
        {"sub": "Nadir", "value": u["eff_gt_nadir_dB_K"], "unit": "dB/K"},
        {"label": "C/No", "sub": "Slant", "value": u["cn0_slant_dBHz"], "unit": "dBHz"},
        {"sub": "Nadir", "value": u["cn0_nadir_dBHz"], "unit": "dBHz"},
        {"label": "Eb/No", "sub": "Slant", "value": u["ebn0_slant_dB"], "unit": "dB"},
        {"sub": "Nadir", "value": u["ebn0_nadir_dB"], "unit": "dB"},
        {"label": "Req. Eb/No", "value": None, "unit": "dB"},
        {"label": "System Margin", "value": None, "unit": "dB"},
        {"label": "Excess Margin", "sub": "Slant", "value": u["excess_slant_dB"], "unit": "dB"},
        {"sub": "Nadir", "value": u["excess_nadir_dB"], "unit": "dB"},
        {"label": "Sensitivity Margin", "value": u["sensitivity_margin_dB"], "unit": "dB"},
    ]


def _downlink_rows(cfg: LinkBudgetExportConfig, d: Dict[str, float]) -> list[dict]:
    """Ordered downlink parameter rows mirroring the reference Downlink sheet."""
    rows = [
        {"label": "Earth Radius", "input": EARTH_RADIUS_KM, "unit": "km", "number_format": "0"},
        {"label": "Boltzmann Constant", "input": BOLTZMANN_CONSTANT_J_K, "unit": "J/K", "number_format": "0.00E+00"},
        {"label": "Noise Figure Reference Temp.", "input": NOISE_FIGURE_REFERENCE_TEMP_K, "unit": "K", "number_format": "0"},
        {"label": "Speed of Light", "input": SPEED_OF_LIGHT_M_S, "unit": "m/s", "number_format": "0"},
        {"label": "Downlink Frequency", "input": cfg.dl_frequency_MHz, "unit": "MHz", "number_format": "0.0"},
        {"label": "Downlink Wavelength", "calc": d["wavelength_m"], "unit": "m", "number_format": "0.000"},
        {"label": "Modulation Type", "input": cfg.dl_modulation_type, "unit": "", "number_format": "General"},
        {"label": "Modulation Order", "input": cfg.dl_modulation_order, "unit": "", "number_format": "0"},
        {"label": cfg.dl_coding_label, "input": cfg.dl_coding_rate, "unit": "", "number_format": "0.000"},
        {"label": "SRRC Roll-off", "input": cfg.dl_rolloff, "unit": "", "number_format": "0.00"},
        {"label": "Symbol Rate", "input": cfg.dl_symbol_rate_ksps, "unit": "ksps", "number_format": "0"},
        {"label": "Information Rate", "calc": d["info_rate_kbps"], "unit": "kbps", "number_format": "0"},
        {"label": "Occupied Bandwidth", "calc": d["occupied_bw_khz"], "unit": "kHz", "number_format": "0.0"},
        {"label": "Spectral Efficiency", "calc": d["spectral_eff"], "unit": "bit/sym/Hz", "number_format": "0.000"},
        {"label": "Tx Power", "input": cfg.dl_tx_power_dBW, "unit": "dBW", "number_format": "0.00"},
        {"label": "Tx Harness Loss", "input": cfg.dl_tx_harness_loss_dB, "unit": "dB"},
        {"label": "Tx Antenna Gain", "sub": "Slant", "input": cfg.dl_tx_antenna_gain_slant_dBi, "unit": "dBi", "merge_next": True},
        {"sub": "Nadir", "input": cfg.dl_tx_antenna_gain_nadir_dBi, "unit": "dBi"},
        {"label": "Pointing Loss", "input": cfg.dl_pointing_loss_dB, "unit": "dB"},
        {"label": "Implementation Loss", "input": cfg.dl_implementation_loss_dB, "unit": "dB"},
        {"label": "Effective EIRP", "sub": "Slant", "calc": d["eirp_slant_dBW"], "unit": "dBW", "merge_next": True},
        {"sub": "Nadir", "calc": d["eirp_nadir_dBW"], "unit": "dBW"},
        {"label": "Orbit Altitude", "input": cfg.satellite_altitude_km, "unit": "km", "number_format": "0"},
        {"label": "Slant Elevation", "input": cfg.slant_elevation_deg, "unit": "deg", "number_format": "0.0"},
        {"label": "Ground Station Name", "input": cfg.ground_station_name, "unit": "", "number_format": "General"},
        {"label": "Ground Station Altitude", "input": cfg.ground_altitude_km, "unit": "km", "number_format": "0.000"},
        {"label": "Slant Range", "sub": "Slant", "calc": d["slant_range_km"], "unit": "km", "number_format": "0.0", "merge_next": True},
        {"sub": "Nadir", "calc": d["nadir_range_km"], "unit": "km", "number_format": "0.0"},
        {"label": "Free Space Path Loss", "sub": "Slant", "calc": d["fspl_slant_dB"], "unit": "dB", "merge_next": True},
        {"sub": "Nadir", "calc": d["fspl_nadir_dB"], "unit": "dB"},
        {"label": "Atmospheric Loss", "sub": "Slant", "input": d["atm_slant_dB"], "unit": "dB", "merge_next": True},
        {"sub": "Nadir", "input": d["atm_nadir_dB"], "unit": "dB"},
        {"label": "Incident Channel Power", "sub": "Slant", "calc": d["incident_slant_dBW"], "unit": "dBW", "merge_next": True},
        {"sub": "Nadir", "calc": d["incident_nadir_dBW"], "unit": "dBW"},
        {"label": "Atmospheric Temperature", "input": cfg.dl_atmospheric_temperature_K, "unit": "K", "number_format": "0.0"},
        {"label": "Background Noise Temp", "input": cfg.dl_background_noise_temp_K, "unit": "K", "number_format": "0.0"},
        {"label": "Brightness Temperature", "sub": "Slant", "calc": d["brightness_temp_slant_K"], "unit": "K", "number_format": "0.0", "merge_next": True},
        {"sub": "Nadir", "calc": d["brightness_temp_nadir_K"], "unit": "K", "number_format": "0.0"},
        {"label": "Incident Noise Power", "sub": "Slant", "calc": d["incident_noise_slant_dBW"], "unit": "dBW", "merge_next": True},
        {"sub": "Nadir", "calc": d["incident_noise_nadir_dBW"], "unit": "dBW"},
        {"label": "Rx Antenna Gain", "input": cfg.dl_rx_antenna_gain_dBi, "unit": "dBi"},
        {"label": "Polarization Loss", "input": cfg.dl_polarization_loss_dB, "unit": "dB"},
        {"label": "Mispointing Loss", "input": cfg.dl_mispointing_loss_dB, "unit": "dB"},
        {"label": "System Noise Temperature", "input": d["sys_noise_K"], "unit": "K", "number_format": "0"},
        {"label": "System G/T", "calc": d["gt_dB_K"], "unit": "dB/K"},
        {"label": "Effective System G/T", "calc": d["eff_gt_dB_K"], "unit": "dB/K"},
        {"label": "Received Carrier Power (C) ", "sub": "Slant", "calc": d["rx_carrier_slant_dBW"], "unit": "dBW", "merge_next": True},
        {"sub": "Nadir", "calc": d["rx_carrier_nadir_dBW"], "unit": "dBW"},
        {"label": "Received Noise Power (N)", "calc": d["rx_noise_power_dBW"], "unit": "dBW"},
        {"label": "Carrier-to-Noise Ratio (C/N)", "sub": "Slant", "calc": d["cn_slant_dB"], "unit": "dB", "merge_next": True},
        {"sub": "Nadir", "calc": d["cn_nadir_dB"], "unit": "dB"},
        {"label": "Carrier-to-Noise-Density Ratio (C/No)", "sub": "Slant", "calc": d["cn0_slant_dBHz"], "unit": "dBHz", "merge_next": True},
        {"sub": "Nadir", "calc": d["cn0_nadir_dBHz"], "unit": "dBHz"},
        {"label": "Eb/No", "sub": "Slant", "calc": d["ebn0_slant_dB"], "unit": "dB", "merge_next": True},
        {"sub": "Nadir", "calc": d["ebn0_nadir_dB"], "unit": "dB"},
        {"label": "Required Eb/N0", "input": cfg.dl_required_ebn0_dB, "unit": "dB"},
        {"label": "System Margin", "input": cfg.system_margin_dB, "unit": "dB"},
        {"label": "Excess Margin", "sub": "Slant", "calc": d["excess_slant_dB"], "unit": "dB", "merge_next": True},
        {"sub": "Nadir", "calc": d["excess_nadir_dB"], "unit": "dB"},
    ]
    if cfg.include_pfd:
        rows += [
            {"label": "Spreading Loss", "sub": "Slant", "calc": d["spreading_slant_dBm2"], "unit": "dBm2", "merge_next": True},
            {"sub": "Nadir", "calc": d["spreading_nadir_dBm2"], "unit": "dBm2"},
            {"label": "Downlink PFD", "sub": "Slant", "calc": d["pfd_slant"], "unit": "dBW/m2/4kHz", "merge_next": True},
            {"sub": "Nadir", "calc": d["pfd_nadir"], "unit": "dBW/m2/4kHz"},
            {"label": "PFD limit", "sub": "Slant", "calc": d["pfd_limit_slant"], "unit": "dBW/m2/4kHz", "merge_next": True},
            {"sub": "Nadir", "calc": d["pfd_limit_nadir"], "unit": "dBW/m2/4kHz"},
            {"label": "PFD margin", "sub": "Slant", "calc": d["pfd_margin_slant"], "unit": "dB", "merge_next": True},
            {"sub": "Nadir", "calc": d["pfd_margin_nadir"], "unit": "dB"},
        ]
    return rows


def _downlink_summary_rows(cfg: LinkBudgetExportConfig, d: Dict[str, float]) -> list[dict]:
    rows = [
        {"label": "EIRP", "sub": "Slant", "value": d["eirp_slant_dBW"], "unit": "dBW"},
        {"sub": "Nadir", "value": d["eirp_nadir_dBW"], "unit": "dBW"},
        {"label": "Effective System G/T", "value": d["eff_gt_dB_K"], "unit": "dB/K"},
        {"label": "C/No", "sub": "Slant", "value": d["cn0_slant_dBHz"], "unit": "dBHz"},
        {"sub": "Nadir", "value": d["cn0_nadir_dBHz"], "unit": "dBHz"},
        {"label": "Eb/No", "sub": "Slant", "value": d["ebn0_slant_dB"], "unit": "dB"},
        {"sub": "Nadir", "value": d["ebn0_nadir_dB"], "unit": "dB"},
        {"label": "Req. Eb/No", "value": cfg.dl_required_ebn0_dB, "unit": "dB"},
        {"label": "System Margin", "value": cfg.system_margin_dB, "unit": "dB"},
        {"label": "Excess Margin", "sub": "Slant", "value": d["excess_slant_dB"], "unit": "dB"},
        {"sub": "Nadir", "value": d["excess_nadir_dB"], "unit": "dB"},
    ]
    if cfg.include_pfd:
        rows += [
            {"label": "PFD Margin", "sub": "Slant", "value": d["pfd_margin_slant"], "unit": "dB"},
            {"sub": "Nadir", "value": d["pfd_margin_nadir"], "unit": "dB"},
        ]
    return rows


# ════════════════════════════════════════════════════════════════════════════
# Public API
# ════════════════════════════════════════════════════════════════════════════
def build_workbook(cfg: LinkBudgetExportConfig) -> Workbook:
    """Build and return an in-memory :class:`openpyxl.Workbook`."""
    u = compute_uplink(cfg)
    d = compute_downlink(cfg)

    wb = Workbook()

    # ── Uplink sheet ────────────────────────────────────────────────────────
    ws_ul = wb.active
    ws_ul.title = "Uplink"
    ws_ul.sheet_view.showGridLines = False
    _build_block(ws_ul, _uplink_rows(cfg, u), base_col=2, start_row=2, title="UPLINK")
    _build_summary_block(
        ws_ul, _uplink_summary_rows(u), base_col=8, start_row=2, title="UPLINK SUMMARY"
    )

    # ── Downlink sheet ──────────────────────────────────────────────────────
    ws_dl = wb.create_sheet("Downlink")
    ws_dl.sheet_view.showGridLines = False
    _build_block(ws_dl, _downlink_rows(cfg, d), base_col=2, start_row=2, title="DOWNLINK")
    _build_summary_block(
        ws_dl, _downlink_summary_rows(cfg, d), base_col=8, start_row=2, title="DOWNLINK SUMMARY"
    )

    # ── Combined "Link Budget" sheet (Downlink | Uplink side by side) ────────
    ws_lb = wb.create_sheet("Link Budget")
    ws_lb.sheet_view.showGridLines = False
    _build_block(ws_lb, _downlink_rows(cfg, d), base_col=2, start_row=2, title="DOWNLINK")
    # Uplink block to the right, separated by a spacer column.
    _build_block(ws_lb, _uplink_rows(cfg, u), base_col=8, start_row=2, title="UPLINK")
    ws_lb.column_dimensions[get_column_letter(7)].width = 3.0

    return wb


def write_link_budget_xlsx(
    cfg: LinkBudgetExportConfig, path: str | Path
) -> Path:
    """Compute the link budget and write a styled ``.xlsx`` workbook.

    Returns the resolved output :class:`~pathlib.Path`.
    """
    path = Path(path)
    if path.suffix.lower() != ".xlsx":
        path = path.with_suffix(".xlsx")
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = build_workbook(cfg)
    wb.save(str(path))
    return path


__all__ = [
    "LinkBudgetExportConfig",
    "load_config",
    "compute_uplink",
    "compute_downlink",
    "build_workbook",
    "write_link_budget_xlsx",
]

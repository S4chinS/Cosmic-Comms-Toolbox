"""Tests for the link-budget XLSX export (headless).

The reference values below are the cached Excel results from the "N-STAR"
link-budget workbook this exporter reproduces.  The Uplink sheet evaluates the
slant case at 5 deg elevation; the Downlink sheet at 10 deg.
"""

from __future__ import annotations

import json

import pytest
from openpyxl import load_workbook

from cosmic_toolbox.services.link_budget_xlsx import (
    LinkBudgetExportConfig,
    build_workbook,
    compute_downlink,
    compute_uplink,
    load_config,
    write_link_budget_xlsx,
)


def test_compute_uplink_matches_reference_at_5deg() -> None:
    cfg = LinkBudgetExportConfig(slant_elevation_deg=5.0)
    u = compute_uplink(cfg)
    assert u["slant_range_km"] == pytest.approx(1329.3931, abs=1e-2)
    assert u["fspl_slant_dB"] == pytest.approx(161.3900, abs=1e-2)
    assert u["effective_eirp_dBW"] == pytest.approx(44.0, abs=1e-6)
    assert u["incident_slant_dBW"] == pytest.approx(-118.5300, abs=1e-2)
    assert u["ebn0_slant_dB"] == pytest.approx(29.0048, abs=1e-2)
    assert u["excess_slant_dB"] == pytest.approx(13.4048, abs=1e-2)
    assert u["sensitivity_margin_dB"] == pytest.approx(7.8773, abs=1e-2)


def test_compute_downlink_matches_reference_at_10deg() -> None:
    cfg = LinkBudgetExportConfig(slant_elevation_deg=10.0)
    d = compute_downlink(cfg)
    assert d["slant_range_km"] == pytest.approx(1007.134, abs=1e-2)
    assert d["fspl_slant_dB"] == pytest.approx(159.4052, abs=1e-2)
    assert d["eirp_slant_dBW"] == pytest.approx(-6.5, abs=1e-6)
    assert d["eirp_nadir_dBW"] == pytest.approx(0.5, abs=1e-6)
    assert d["ebn0_slant_dB"] == pytest.approx(9.9552, abs=1e-2)
    assert d["excess_slant_dB"] == pytest.approx(-0.5448, abs=1e-2)
    assert d["pfd_margin_slant"] == pytest.approx(14.8045, abs=1e-2)


def test_pfd_limit_breakpoints() -> None:
    # Downlink PFD limit mask: -154 below 5 deg, ramps to -144 by 25 deg.
    low = compute_downlink(LinkBudgetExportConfig(slant_elevation_deg=5.0))
    mid = compute_downlink(LinkBudgetExportConfig(slant_elevation_deg=15.0))
    assert low["pfd_limit_slant"] == pytest.approx(-154.0)
    assert mid["pfd_limit_slant"] == pytest.approx(-149.0)
    # Nadir (90 deg) is always at the -144 ceiling.
    assert low["pfd_limit_nadir"] == pytest.approx(-144.0)


def test_workbook_structure_and_styling() -> None:
    wb = build_workbook(LinkBudgetExportConfig())
    assert wb.sheetnames == ["Uplink", "Downlink", "Link Budget"]

    ws = wb["Uplink"]
    assert ws["B2"].value == "UPLINK"
    assert ws["H2"].value == "UPLINK SUMMARY"
    assert ws["B4"].value == "Parameters"
    assert ws["D4"].value == "Inputs"
    assert ws["E4"].value == "Calculations"
    assert ws["F4"].value == "Units"
    # Header cells are bold; gridlines are hidden like the reference.
    assert ws["B4"].font.bold is True
    assert ws.sheet_view.showGridLines is False
    # Data cells carry the thin border grid.
    assert ws["B5"].border.left.style == "thin"
    # Slant/Nadir parameter labels are merged across two rows.
    assert len(ws.merged_cells.ranges) > 0

    dl = wb["Downlink"]
    assert dl["B2"].value == "DOWNLINK"
    # PFD rows present when include_pfd is True.
    labels = {c.value for col in dl.iter_cols(min_col=2, max_col=2) for c in col}
    assert "PFD margin" in labels

    combined = wb["Link Budget"]
    assert combined["B2"].value == "DOWNLINK"
    assert combined["H2"].value == "UPLINK"


def test_include_pfd_toggle() -> None:
    wb = build_workbook(LinkBudgetExportConfig(include_pfd=False))
    dl = wb["Downlink"]
    labels = {c.value for col in dl.iter_cols(min_col=2, max_col=2) for c in col}
    assert "PFD margin" not in labels


def test_config_from_dict_rejects_unknown_keys() -> None:
    with pytest.raises(ValueError):
        LinkBudgetExportConfig.from_dict({"not_a_real_key": 1})


def test_load_config_yaml_and_json_roundtrip(tmp_path) -> None:
    cfg = LinkBudgetExportConfig(ground_station_name="Test GS", satellite_altitude_km=600.0)
    data = cfg.to_dict()

    json_path = tmp_path / "cfg.json"
    json_path.write_text(json.dumps(data), encoding="utf-8")
    loaded_json = load_config(json_path)
    assert loaded_json.ground_station_name == "Test GS"
    assert loaded_json.satellite_altitude_km == 600.0

    yaml = pytest.importorskip("yaml")
    yaml_path = tmp_path / "cfg.yaml"
    yaml_path.write_text(yaml.safe_dump(data), encoding="utf-8")
    loaded_yaml = load_config(yaml_path)
    assert loaded_yaml.ground_station_name == "Test GS"


def test_write_link_budget_xlsx_creates_file(tmp_path) -> None:
    out = write_link_budget_xlsx(LinkBudgetExportConfig(), tmp_path / "out")
    assert out.exists()
    assert out.suffix == ".xlsx"
    wb = load_workbook(out)
    assert wb.sheetnames == ["Uplink", "Downlink", "Link Budget"]
